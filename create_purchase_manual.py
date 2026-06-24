import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()

# ── Color palette / styles ──────────────────────────────────
DARK_BLUE   = "1F3864"
MED_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"
YELLOW_BG   = "FFF2CC"
GREEN_BG    = "E2EFDA"
RED_BG      = "FCE4EC"

header_font  = Font(name="Calibri", bold=True, color=WHITE, size=11)
title_font   = Font(name="Calibri", bold=True, color=DARK_BLUE, size=14)
sub_font     = Font(name="Calibri", bold=True, color=DARK_BLUE, size=11)
normal_font  = Font(name="Calibri", size=10)
bold_font    = Font(name="Calibri", bold=True, size=10)
small_font   = Font(name="Calibri", size=9, italic=True, color="555555")

header_fill  = PatternFill("solid", fgColor=MED_BLUE)
alt_fill     = PatternFill("solid", fgColor=LIGHT_GRAY)

thin_border  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
wrap_align   = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row, cols, fill=None):
    """Apply header styling to a row."""
    f = fill or header_fill
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = f
        cell.alignment = center_align
        cell.border = thin_border


def write_title_row(ws, row, col_count, text):
    """Write a title across col_count columns (no merge_cells — writes to A)."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="center")


def write_sub_title_row(ws, row, text):
    """Write a sub-title in column A."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = sub_font


def write_bullet_rows(ws, start_row, col_count, items):
    """Write bullet items starting at start_row."""
    r = start_row
    for item in items:
        cell = ws.cell(row=r, column=1, value=f"• {item}")
        cell.font = normal_font
        cell.alignment = wrap_align
        r += 1
    return r


def write_table(ws, headers, data, start_row=2, col_widths=None):
    """Write a table with header row + data rows. Returns next empty row."""
    ncols = len(headers)
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=i)
        cell.value = h
    style_header_row(ws, start_row, ncols)

    for r_idx, row_data in enumerate(data, start_row + 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = val
            cell.font = normal_font
            cell.alignment = wrap_align
            cell.border = thin_border
            if (r_idx - start_row) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    return start_row + len(data) + 1


def write_note_row(ws, row, col_count, text):
    """Write a small-font note in column A spanning col_count visually."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = small_font
    cell.alignment = wrap_align


# =============================================================
# SHEET 1 — สารบัญ / Index
# =============================================================
ws_idx = wb.active
ws_idx.title = "สารบัญ"
ws_idx.sheet_properties.tabColor = DARK_BLUE

write_title_row(ws_idx, 1, 3, "คู่มือการจัดซื้อ — ISO 9001:2015 / Purchasing Manual")

info_text = f"เอกสารหมายเลข: PM-001  |  ฉบับที่: 01  |  วันที่ประกาศใช้: {date.today().strftime('%d/%m/%Y')}"
cell = ws_idx.cell(row=2, column=1, value=info_text)
cell.font = small_font
cell.alignment = Alignment(horizontal="center")

write_table(ws_idx, ["ลำดับ", "หัวข้อ", "หมายเหตุ"], [
    [1, "วัตถุประสงค์ (Objective)", "ข้อ 1"],
    [2, "ขอบเขต (Scope)", "ข้อ 2"],
    [3, "เอกสารอ้างอิง (Reference Documents)", "ข้อ 3"],
    [4, "คำจำกัดความ (Definitions)", "ข้อ 4"],
    [5, "ความรับผิดชอบ (Responsibilities)", "ข้อ 5"],
    [6, "ขั้นตอนการจัดซื้อ (Purchasing Process)", "ข้อ 6"],
    [7, "การประเมินและคัดเลือกผู้ขาย (Supplier Evaluation)", "ข้อ 7"],
    [8, "ใบสั่งซื้อ / ข้อตกลงการจัดซื้อ (Purchase Order)", "ข้อ 8"],
    [9, "การตรวจสอบสินค้าที่ได้รับ (Incoming Inspection)", "ข้อ 9"],
    [10, "บันทึกคุณภาพ (Quality Records)", "ข้อ 10"],
    [11, "การแก้ไขและการปรับปรุง (Corrective Actions)", "ข้อ 11"],
], start_row=4, col_widths=[8, 55, 15])

ws_idx.sheet_view.showGridLines = False

# =============================================================
# SHEET 2 — วัตถุประสงค์
# =============================================================
ws1 = wb.create_sheet("วัตถุประสงค์")
ws1.sheet_properties.tabColor = MED_BLUE

write_title_row(ws1, 1, 2, "1. วัตถุประสงค์ (Objective)")

write_bullet_rows(ws1, 3, 2, [
    "เพื่อกำหนดแนวทางและขั้นตอนการดำเนินงานจัดซื้อให้เป็นมาตรฐาน สอดคล้องกับข้อกำหนด ISO 9001:2015 ข้อ 8.4 (การควบคุมผลิตภัณฑ์และบริการที่จัดหาจากภายนอก)",
    "เพื่อให้มั่นใจว่าผลิตภัณฑ์และบริการที่จัดซื้อมีคุณภาพตรงตามข้อกำหนดที่องค์กรต้องการ",
    "เพื่อคัดเลือก ประเมิน และติดตามผลผู้ขาย/ผู้รับเหมาช่วงอย่างมีประสิทธิภาพ",
    "เพื่อสร้างความโปร่งใส ตรวจสอบได้ และลดความเสี่ยงในการจัดซื้อ",
])

ws1.column_dimensions["A"].width = 100
ws1.column_dimensions["B"].width = 20

# =============================================================
# SHEET 3 — ขอบเขต
# =============================================================
ws2 = wb.create_sheet("ขอบเขต")
ws2.sheet_properties.tabColor = MED_BLUE

write_title_row(ws2, 1, 2, "2. ขอบเขต (Scope)")

write_bullet_rows(ws2, 3, 2, [
    "คู่มือนี้ครอบคลุมกระบวนการจัดซื้อทั้งหมดขององค์กร ตั้งแต่การระบุความต้องการจนถึงการรับสินค้า/บริการ",
    "ครอบคลุมการจัดซื้อวัตถุดิบ วัสดุสิ้นเปลือง ครุภัณฑ์ และบริการภายนอกที่ส่งผลต่อคุณภาพของผลิตภัณฑ์",
    "ครอบคลุมการประเมิน คัดเลือก และการติดตามผลผู้ขาย (Supplier) และผู้รับเหมาช่วง (Subcontractor)",
    "ยกเว้น: การจัดซื้อที่ไม่ส่งผลกระทบต่อคุณภาพของผลิตภัณฑ์ (เช่น อุปกรณ์สำนักงานทั่วไป) ให้นำเฉพาะแนวปฏิบัติที่ดีมาใช้",
])

ws2.column_dimensions["A"].width = 100
ws2.column_dimensions["B"].width = 20

# =============================================================
# SHEET 4 — เอกสารอ้างอิง
# =============================================================
ws3 = wb.create_sheet("เอกสารอ้างอิง")
ws3.sheet_properties.tabColor = MED_BLUE

write_title_row(ws3, 1, 4, "3. เอกสารอ้างอิง (Reference Documents)")

write_table(ws3,
    ["ลำดับ", "รหัสเอกสาร", "ชื่อเอกสาร", "หมายเลข ISO"],
    [
        [1, "QM-001", "คู่มือคุณภาพ (Quality Manual)", "ISO 9001:2015"],
        [2, "WI-PUR-001", "ขั้นตอนการจัดซื้อ (Purchasing Procedure)", "8.4"],
        [3, "WI-PUR-002", "ขั้นตอนการประเมินผู้ขาย (Supplier Evaluation)", "8.4.1"],
        [4, "WI-QC-001", "ขั้นตอนการตรวจสอบขาเข้า (Incoming Inspection)", "8.4.3"],
        [5, "FR-PUR-001", "แบบฟอร์มใบสั่งซื้อ (Purchase Order Form)", "8.4.2"],
        [6, "FR-PUR-002", "แบบฟอร์มประเมินผู้ขาย (Supplier Evaluation Form)", "8.4.1"],
        [7, "FR-PUR-003", "แบบฟอร์มทะเบียนผู้ขาย (Approved Supplier List)", "8.4.1"],
        [8, "FR-QC-001", "รายงานตรวจสอบคุณภาพ (Inspection Report)", "8.4.3"],
        [9, "—", "ข้อกำหนดระบบบริหารคุณภาพ ISO 9001:2015", "ทั้งหมด"],
        [10, "—", "ข้อกำหนดของลูกค้า (Customer Specifications)", "7.2 / 8.2"],
    ],
    start_row=3, col_widths=[8, 18, 45, 18]
)

# =============================================================
# SHEET 5 — คำจำกัดความ
# =============================================================
ws4 = wb.create_sheet("คำจำกัดความ")
ws4.sheet_properties.tabColor = MED_BLUE

write_title_row(ws4, 1, 3, "4. คำจำกัดความ (Definitions)")

write_table(ws4,
    ["คำศัพท์", "คำอธิบาย", "เทียบเคียง ISO"],
    [
        ["ผู้ขาย / Supplier", "บุคคล หรือ นิติบุคคล ที่จัดหาผลิตภัณฑ์หรือบริการให้องค์กร", "External Provider (8.4)"],
        ["ผู้รับเหมาช่วง / Subcontractor", "ผู้ที่ได้รับมอบหมายให้ดำเนินการบางส่วนของกระบวนการผลิตแทนองค์กร", "Outsourced Process (8.4)"],
        ["ใบสั่งซื้อ / Purchase Order (PO)", "เอกสารที่ระบุรายละเอียดสินค้า ปริมาณ ราคา เงื่อนไข และกำหนดส่ง", "Purchase Order (8.4.2)"],
        ["ทะเบียนผู้ขาย / Approved Supplier List (ASL)", "รายชื่อผู้ขายที่ผ่านการประเมินและอนุมัติให้จัดซื้อได้", "Approved Supplier List (8.4.1)"],
        ["การประเมินผู้ขาย / Supplier Evaluation", "กระบวนการประเมินความสามารถของผู้ขายตามเกณฑ์ที่กำหนด", "Evaluation (8.4.1)"],
        ["สินค้าไม่เป็นไปตามข้อกำหนด / Nonconforming Product", "สินค้าหรือบริการที่ไม่ตรงตามข้อกำหนดที่ระบุใน PO หรือสเปก", "Nonconformity (8.7 / 10.2)"],
        ["ข้อกำหนดการจัดซื้อ / Purchase Requirement", "รายละเอียดคุณสมบัติ มาตรฐาน หรือสเปกที่ต้องการจากผู้ขาย", "Purchasing Requirements (8.4.2)"],
    ],
    start_row=3, col_widths=[28, 55, 22]
)

# =============================================================
# SHEET 6 — ความรับผิดชอบ
# =============================================================
ws5 = wb.create_sheet("ความรับผิดชอบ")
ws5.sheet_properties.tabColor = MED_BLUE

write_title_row(ws5, 1, 3, "5. ความรับผิดชอบ (Responsibilities)")

write_table(ws5,
    ["บทบาท / ฝ่าย", "หน้าที่ความรับผิดชอบหลัก", "เอกสารที่เกี่ยวข้อง"],
    [
        ["ผู้บริหารสูงสุด (Top Management)", "อนุมัตินโยบายการจัดซื้อ อนุมัติผู้ขายรายใหม่ อนุมัติวงเงินสูง", "QM-001, FR-PUR-002"],
        ["ฝ่ายจัดซื้อ (Purchasing Dept.)", "ดำเนินการจัดซื้อตามขั้นตอน คัดเลือกผู้ขาย จัดทำ PO ติดตามงาน", "WI-PUR-001, FR-PUR-001"],
        ["ฝ่ายควบคุมคุณภาพ (QC Dept.)", "ตรวจสอบคุณภาพสินค้าขาเข้า อนุมัติ/ปฏิเสธสินค้า", "WI-QC-001, FR-QC-001"],
        ["ฝ่ายคลังสินค้า (Warehouse)", "รับสินค้า ตรวจทานจำนวน จัดเก็บ และแจ้ง QC ตรวจสอบ", "WI-WH-001"],
        ["ฝ่ายวิศวกรรม / ผลิต (Engineering)", "กำหนดสเปก มาตรฐาน และข้อกำหนดทางเทคนิคของสินค้า", "Specification Sheet"],
        ["ผู้ร้องขอซื้อ (Requestor)", "ระบุความต้องการ จัดทำ PR (Purchase Requisition)", "FR-PR-001"],
    ],
    start_row=3, col_widths=[30, 55, 25]
)

# =============================================================
# SHEET 7 — ขั้นตอนการจัดซื้อ
# =============================================================
ws6 = wb.create_sheet("ขั้นตอนการจัดซื้อ")
ws6.sheet_properties.tabColor = MED_BLUE

write_title_row(ws6, 1, 5, "6. ขั้นตอนการจัดซื้อ (Purchasing Process)")

write_table(ws6,
    ["ขั้นตอน", "ผู้รับผิดชอบ", "รายละเอียดการปฏิบัติงาน", "เอกสาร/บันทึก", "ข้อกำหนด ISO"],
    [
        ["1. ระบุความต้องการ", "ผู้ร้องขอซื้อ (Requestor)", "ระบุรายละเอียดสินค้า/บริการ ปริมาณ กำหนดส่ง และสเปก", "PR (FR-PR-001)", "8.4.2"],
        ["2. ขออนุมัติจัดซื้อ", "หัวหน้าฝ่ายที่เกี่ยวข้อง", "ตรวจสอบความจำเป็น, งบประมาณ, และอนุมัติ PR", "PR อนุมัติ (FR-PR-001)", "5.1 / 7.1"],
        ["3. คัดเลือกผู้ขาย", "ฝ่ายจัดซื้อ", "เลือกจาก ASL หรือดำเนินการประเมินผู้ขายรายใหม่ (ถ้าจำเป็น)", "ASL (FR-PUR-003)", "8.4.1"],
        ["4. ขอ quotation / สอบราคา", "ฝ่ายจัดซื้อ", "ขอใบเสนอราคาอย่างน้อย 3 ราย (สำหรับมูลค่าสูง) เทียบราคาและเงื่อนไข", "Quotation / RFQ", "8.4.2"],
        ["5. จัดทำ PO", "ฝ่ายจัดซื้อ", "จัดทำใบสั่งซื้อระบุสินค้า ปริมาณ ราคา วันที่ส่ง และเงื่อนไขชัดเจน", "PO (FR-PUR-001)", "8.4.2"],
        ["6. อนุมัติ PO", "ผู้มีอำนาจอนุมัติ", "ตรวจสอบความถูกต้องของ PO ก่อนส่งให้ผู้ขาย", "PO อนุมัติ", "5.1"],
        ["7. สั่งซื้อ / ส่ง PO", "ฝ่ายจัดซื้อ", "ส่ง PO ให้ผู้ขายทาง Email / Fax / ระบบ EDI", "PO ที่ส่ง (Record)", "8.4.2"],
        ["8. ติดตาม / expediting", "ฝ่ายจัดซื้อ", "ติดตามความคืบหน้า ยืนยันกำหนดส่ง แจ้งเตือนหากล่าช้า", "Email / Expediting Log", "8.4.2"],
        ["9. รับสินค้า", "คลังสินค้า", "นับจำนวน ตรวจสอบเอกสารส่งของ (Delivery Note) เทียบกับ PO", "DN / GRN", "8.4.3"],
        ["10. ตรวจสอบคุณภาพ", "QC", "ตรวจสอบสินค้าตามแผนการตรวจสอบ (Inspection Plan)", "Inspection Report (FR-QC-001)", "8.4.3"],
        ["11. รับเข้าระบบ / reject", "คลัง / QC", "OK → รับเข้าคลัง  /  NG → ดำเนินการตาม WI สำหรับสินค้า Nonconforming", "GRN / NCR", "8.4.3 / 8.7"],
        ["12. วางบิล / จ่ายเงิน", "ฝ่ายบัญชี", "ตรวจสอบเอกสารครบถ้วนแล้วดำเนินการจ่ายเงินตามเงื่อนไข", "Invoice / Payment Record", "—"],
        ["13. บันทึก / วัดผล", "ฝ่ายจัดซื้อ / QC", "บันทึกผลการส่งมอบ, คุณภาพ, และ Feedback เพื่อใช้ประเมินผู้ขาย", "Supplier Scorecard", "8.4.1 / 9.1"],
    ],
    start_row=3, col_widths=[18, 22, 55, 28, 18]
)

# =============================================================
# SHEET 8 — การประเมินผู้ขาย
# =============================================================
ws7 = wb.create_sheet("การประเมินผู้ขาย")
ws7.sheet_properties.tabColor = MED_BLUE

write_title_row(ws7, 1, 3, "7. การประเมินและคัดเลือกผู้ขาย (Supplier Evaluation)")

# Part A
write_sub_title_row(ws7, 3, "7.1 เกณฑ์การประเมินผู้ขายรายใหม่ (New Supplier Evaluation Criteria)")

write_table(ws7,
    ["เกณฑ์การประเมิน", "รายละเอียด", "น้ำหนัก (%)"],
    [
        ["คุณภาพ (Quality)", "ระบบคุณภาพ (ISO 9001 / มาตรฐานอื่น), การควบคุมกระบวนการ, ประวัติคุณภาพสินค้า", "35%"],
        ["ราคา (Price)", "ความสามารถในการแข่งขันด้านราคา, ความโปร่งใสของโครงสร้างราคา", "20%"],
        ["การส่งมอบ (Delivery)", "ความสามารถในการส่งมอบตรงเวลา, ความยืดหยุ่น, Lead Time", "20%"],
        ["บริการ (Service)", "การตอบสนอง, การให้คำปรึกษาทางเทคนิค, การบริการหลังการขาย", "10%"],
        ["ความน่าเชื่อถือ (Reliability)", "ประวัติบริษัท, ฐานะการเงิน, ข้อพิพาทในอดีต, การอ้างอิงลูกค้า", "10%"],
        ["สิ่งแวดล้อม/ความปลอดภัย (EHS)", "การปฏิบัติตามกฎหมาย, ใบอนุญาต, ISO 14001 / ISO 45001", "5%"],
    ],
    start_row=4, col_widths=[35, 65, 15]
)

# Part B
# Determine next row: 4 + 1 header + 6 data = 11
next_row = 12
write_sub_title_row(ws7, next_row, "7.2 ระดับผลการประเมิน (Evaluation Rating)")

write_table(ws7,
    ["คะแนนรวม", "ระดับ", "การดำเนินการ"],
    [
        ["≥ 85%", "A — ดีเยี่ยม (Excellent)", "ปรับปรุงต่อเนื่อง, พิจารณาสิทธิพิเศษ"],
        ["70% – 84%", "B — ดี (Good)", "ติดตามผลเป็นระยะ, แจ้งจุดที่ควรปรับปรุง"],
        ["50% – 69%", "C — พอใช้ (Fair)", "กำหนดแผนแก้ไข (CAP) และติดตามอย่างใกล้ชิด"],
        ["< 50%", "D — ต้องปรับปรุง (Poor)", "พิจารณาเพิกถอนจาก ASL / หาผู้ขายรายใหม่"],
    ],
    start_row=next_row + 1, col_widths=[18, 30, 55]
)

# Part C
next_row2 = next_row + 1 + 1 + 4 + 1  # title + header + 4 data + blank
write_sub_title_row(ws7, next_row2, "7.3 ความถี่ในการประเมินซ้ำ (Re-evaluation Frequency)")

write_table(ws7,
    ["ระดับผู้ขาย", "ความถี่", "วิธี"],
    [
        ["A", "ทุก 12 เดือน", "ทบทวน Performance Scorecard + สอบถาม"],
        ["B", "ทุก 6 เดือน", "ทบทวน Performance Scorecard + ส่งแบบประเมิน"],
        ["C", "ทุก 3 เดือน", "ทบทวน Performance Scorecard + อาจตรวจสอบหน้างาน"],
        ["D", "ทันที", "ดำเนินการตาม CAP หรือถอดจาก ASL"],
    ],
    start_row=next_row2 + 1, col_widths=[18, 18, 65]
)

# =============================================================
# SHEET 9 — ใบสั่งซื้อ (PO)
# =============================================================
ws8 = wb.create_sheet("ใบสั่งซื้อ (PO)")
ws8.sheet_properties.tabColor = MED_BLUE

write_title_row(ws8, 1, 6, "8. ใบสั่งซื้อ / ข้อตกลงการจัดซื้อ (Purchase Order)")

po_fields = [
    ("1. เลขที่ PO", "PO-XXXX-XXX"),
    ("2. วันที่ออก PO", "DD/MM/YYYY"),
    ("3. ชื่อผู้ขาย / Supplier", ""),
    ("4. ที่อยู่ผู้ขาย", ""),
    ("5. เลขที่ผู้เสียภาษี / Tax ID", ""),
    ("6. ผู้ติดต่อ / Contact Person", ""),
    ("7. เงื่อนไขการชำระเงิน", "☐ เงินสด  ☐ เครดิต 30 วัน  ☐ เครดิต 60 วัน  ☐ อื่นๆ ..."),
    ("8. เงื่อนไขการส่งมอบ", "Incoterm: ☐ EXW  ☐ FOB  ☐ CIF  ☐ DDP  ☐ อื่นๆ ..."),
    ("9. วันที่ต้องการสินค้า", "DD/MM/YYYY"),
]

write_table(ws8, ["รายการ", "คำอธิบาย"], po_fields,
    start_row=3, col_widths=[35, 70])

# PO Line Items
next_po = 3 + 1 + len(po_fields) + 1  # = 14
write_sub_title_row(ws8, next_po, "8.1 รายการสินค้า (Line Items)")

write_table(ws8,
    ["ลำดับ", "รหัสสินค้า", "รายละเอียดสินค้า", "จำนวน", "หน่วย", "ราคาต่อหน่วย"],
    [
        ["1", "ITEM-001", "ตัวอย่าง: วัตถุดิบ A", "100", "Kg", "50.00"],
        ["2", "ITEM-002", "", "", "", ""],
        ["3", "ITEM-003", "", "", "", ""],
    ],
    start_row=next_po + 1, col_widths=[8, 15, 40, 12, 10, 15]
)

# ISO Requirement Checklist
next_po2 = next_po + 1 + 1 + 3 + 1  # title + header + 3 data + blank
write_sub_title_row(ws8, next_po2, "8.2 ข้อกำหนด ISO ที่ต้องระบุใน PO (ISO 8.4.2 Requirement Checklist)")

write_table(ws8,
    ["ข้อ", "รายการตรวจสอบ", "☐"],
    [
        ["8.4.2 a)", "ข้อกำหนดสำหรับผลิตภัณฑ์/บริการที่ต้องการจัดซื้อ (Spec / Drawing / Standard)", "☐"],
        ["8.4.2 b)", "ข้อกำหนดสำหรับการอนุมัติ: ผลิตภัณฑ์, ขั้นตอน, กระบวนการ, อุปกรณ์", "☐"],
        ["8.4.2 c)", "ข้อกำหนดสำหรับบุคลากรที่มีคุณสมบัติ (ถ้าต้องการ)", "☐"],
        ["8.4.2 d)", "ข้อกำหนดของระบบบริหารคุณภาพ (เช่น ISO 9001)", "☐"],
    ],
    start_row=next_po2 + 1, col_widths=[15, 60, 8]
)

# Note
note_row = next_po2 + 1 + 1 + 4 + 1  # title + header + 4 data + blank
note_text = ("หมายเหตุ: ทุก PO ต้องมีลายเซ็นผู้อนุมัติก่อนส่งให้ผู้ขาย "
    "และต้องจัดเก็บ PO อย่างน้อย 3 ปี (ตามข้อกำหนด ISO 9001:2015 ข้อ 7.5.3)")
write_note_row(ws8, note_row, 6, note_text)

# =============================================================
# SHEET 10 — การตรวจสอบขาเข้า
# =============================================================
ws9 = wb.create_sheet("การตรวจสอบขาเข้า")
ws9.sheet_properties.tabColor = MED_BLUE

write_title_row(ws9, 1, 5, "9. การตรวจสอบสินค้าที่ได้รับ (Incoming Inspection)")

write_table(ws9,
    ["ขั้นตอน", "ผู้รับผิดชอบ", "รายละเอียด", "เอกสาร", "มาตรฐาน ISO"],
    [
        ["1. รับเอกสารส่งของ", "คลังสินค้า", "ตรวจสอบ Delivery Note, Invoice กับ PO", "Delivery Note / PO", "8.4.3"],
        ["2. ตรวจนับจำนวน", "คลังสินค้า", "นับจำนวนจริง เทียบกับเอกสาร", "GRN", "8.4.3"],
        ["3. ตรวจสอบสภาพภายนอก", "คลังสินค้า", "ตรวจสอบความเสียหายของบรรจุภัณฑ์ / สินค้า", "GRN / Photo", "8.4.3"],
        ["4. แจ้ง QC ตรวจสอบ", "คลังสินค้า", "ติดต่อ QC เพื่อตรวจสอบคุณภาพตามแผน", "Inspection Request", "8.4.3"],
        ["5. ตรวจสอบคุณภาพ", "QC", "ตรวจตาม Inspection Plan / Spec / Drawing / Standard", "Inspection Report (FR-QC-001)", "8.4.3"],
        ["6. ตัดสินใจ", "QC", "Accept / Reject / Conditional Accept", "Inspection Report", "8.4.3"],
        ["7. รับเข้าคลัง / คืนผู้ขาย", "คลัง / QC", "AC → รับเข้า, REJ → คืน + NCR + SCAR", "GRN / NCR / SCAR", "8.4.3 / 8.7"],
    ],
    start_row=3, col_widths=[20, 18, 50, 30, 18]
)

# Sampling Inspection Plan
next_row_s = 3 + 1 + 7 + 1  # title + header + 7 data + blank
write_sub_title_row(ws9, next_row_s, "9.1 ตัวอย่างแผนการตรวจสอบ (Sampling Inspection Plan)")

write_table(ws9,
    ["ประเภทสินค้า", "ระดับการตรวจ", "จำนวนตัวอย่าง", "เกณฑ์ยอมรับ (AQL)", "วิธีตรวจ"],
    [
        ["วัตถุดิบ A (Critical)", "100%", "ทั้งหมด", "Zero Defect (C=0)", "ตาม Spec Sheet"],
        ["วัตถุดิบ B (Major)", "สุ่ม (Random)", "√N หรือตาม MIL-STD-1916", "AQL 1.0%", "ตาม Spec Sheet"],
        ["วัตถุดิบ C (Minor)", "สุ่ม (Random)", "ตาม MIL-STD-1916 Level II", "AQL 4.0%", "Visual + Dimension"],
        ["บริการภายนอก", "—", "ตรวจสอบรายงาน / Certificate", "—", "เอกสารประกอบ"],
    ],
    start_row=next_row_s + 1, col_widths=[25, 18, 28, 22, 22]
)

# =============================================================
# SHEET 11 — บันทึกคุณภาพ
# =============================================================
ws10 = wb.create_sheet("บันทึกคุณภาพ")
ws10.sheet_properties.tabColor = MED_BLUE

write_title_row(ws10, 1, 6, "10. บันทึกคุณภาพ (Quality Records)")

write_table(ws10,
    ["รหัสบันทึก", "ชื่อบันทึก", "ผู้จัดเก็บ", "ระยะเวลาเก็บ", "สถานที่จัดเก็บ", "ISO 9001 อ้างอิง"],
    [
        ["FR-PR-001", "ใบขอซื้อ (Purchase Requisition)", "ฝ่ายจัดซื้อ", "3 ปี", "Server / Filing", "7.5.3"],
        ["FR-PUR-001", "ใบสั่งซื้อ (Purchase Order)", "ฝ่ายจัดซื้อ", "3 ปี", "Server / Filing", "8.4.2 / 7.5.3"],
        ["FR-PUR-002", "แบบประเมินผู้ขาย (Supplier Eval.)", "ฝ่ายจัดซื้อ", "5 ปี", "Server / Filing", "8.4.1"],
        ["FR-PUR-003", "ทะเบียนผู้ขายอนุมัติ (ASL)", "ฝ่ายจัดซื้อ", "ปัจจุบันเสมอ", "Server / Filing", "8.4.1"],
        ["FR-QC-001", "รายงานตรวจสอบคุณภาพ (Insp. Report)", "QC", "5 ปี", "Server / Filing", "8.4.3 / 8.7"],
        ["FR-QC-002", "รายการสินค้าไม่เป็นไปตามข้อกำหนด (NCR)", "QC", "5 ปี", "Server / Filing", "8.7 / 10.2"],
        ["FR-PUR-004", "Supplier Scorecard", "ฝ่ายจัดซื้อ", "5 ปี", "Server / Filing", "8.4.1 / 9.1"],
        ["FR-PUR-005", "ใบแจ้งแก้ไขผู้ขาย (SCAR)", "ฝ่ายจัดซื้อ", "5 ปี", "Server / Filing", "10.2"],
    ],
    start_row=3, col_widths=[15, 35, 16, 16, 20, 20]
)

# =============================================================
# SHEET 12 — การแก้ไขและปรับปรุง
# =============================================================
ws11 = wb.create_sheet("การแก้ไขและปรับปรุง")
ws11.sheet_properties.tabColor = MED_BLUE

write_title_row(ws11, 1, 4, "11. การแก้ไขและการปรับปรุง (Corrective & Preventive Actions)")

write_table(ws11,
    ["กรณี / Event", "การดำเนินการ", "เอกสารที่เกี่ยวข้อง", "ระยะเวลา"],
    [
        ["สินค้าไม่เป็นไปตามข้อกำหนด (Nonconforming Product)", "แยก / ติดป้าย / คืน / ขอ Credit Note", "NCR (FR-QC-002)", "ภายใน 24 ชม."],
        ["ผู้ขายส่งของล่าช้า (>7 วัน)", "บันทึกใน Scorecard, ส่ง SCAR ให้ผู้ขาย", "SCAR (FR-PUR-005)", "ภายใน 7 วัน"],
        ["ผู้ขายส่งของไม่ได้คุณภาพซ้ำ (>2 ครั้ง)", "ถอดจาก ASL ชั่วคราว, หา Supplier รายอื่น", "ASL Update, SCAR", "ทันที"],
        ["ผู้ขายถูกถอดจาก ASL", "ดำเนินการหาและประเมินผู้ขายรายใหม่แทนที่", "FR-PUR-002, FR-PUR-003", "ภายใน 30 วัน"],
        ["ข้อร้องเรียนจากลูกค้า (Customer Complaint)", "ตรวจสอบย้อนกลับถึงการจัดซื้อ, แก้ไขกระบวนการ", "Customer Complaint Log", "ตาม WI-CR-001"],
        ["โอกาสปรับปรุง (Improvement Opportunity)", "เสนอแนวทางลดต้นทุน / เพิ่มคุณภาพผ่านกระบวนการ Kaizen", "Kaizen / CI Report", "ตามรอบ CI"],
    ],
    start_row=3, col_widths=[35, 55, 30, 18]
)

# =============================================================
# SHEET 13 — ตัวอย่างฟอร์ม (Template Preview)
# =============================================================
ws12 = wb.create_sheet("ตัวอย่างฟอร์ม")
ws12.sheet_properties.tabColor = YELLOW_BG

write_title_row(ws12, 1, 5, "ตัวอย่างแบบฟอร์ม — Supplier Evaluation Form")

info_eval = ws12.cell(row=3, column=1, value="วันที่ประเมิน: __________  |  ผู้ประเมิน: __________  |  ครั้งที่: ____")
info_eval.font = normal_font

# Supplier info
eval_info = [
    ("ชื่อผู้ขาย:", "_________________________"),
    ("ที่อยู่:", "_________________________"),
    ("ผู้ติดต่อ:", "_________  โทร: _________"),
    ("ผลิตภัณฑ์/บริการ:", "_________________________"),
]
for i, (label, val) in enumerate(eval_info, 4):
    c1 = ws12.cell(row=i, column=1, value=label)
    c1.font = bold_font
    c2 = ws12.cell(row=i, column=2, value=val)
    c2.font = normal_font

# Evaluation table
ev_headers = ["เกณฑ์", "น้ำหนัก", "คะแนน\n(1-10)", "คะแนนถ่วงน้ำหนัก", "หมายเหตุ"]
ev_data = [
    ["คุณภาพ (Quality)", "35%", "", "", ""],
    ["ราคา (Price)", "20%", "", "", ""],
    ["การส่งมอบ (Delivery)", "20%", "", "", ""],
    ["บริการ (Service)", "10%", "", "", ""],
    ["ความน่าเชื่อถือ (Reliability)", "10%", "", "", ""],
    ["สิ่งแวดล้อม/ความปลอดภัย (EHS)", "5%", "", "", ""],
    ["รวม", "100%", "", "", ""],
]

write_table(ws12, ev_headers, ev_data, start_row=9, col_widths=[35, 12, 14, 20, 25])

# Result
result_cell = ws12.cell(row=18, column=1, value="ผลการประเมิน:  ☐ ผ่าน  ☐ ไม่ผ่าน  ☐ ผ่านแบบมีเงื่อนไข")
result_cell.font = bold_font

ws12.cell(row=19, column=1, value="ลายเซ็นผู้ประเมิน: ______________").font = normal_font
ws12.cell(row=19, column=3, value="ลายเซ็นผู้อนุมัติ: ______________").font = normal_font

# ── Apply print settings to all sheets ──────
for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.zoomScale = 90

# ── Save ──────────────────────────────────────────────────
output_path = "C:\\Working_Space\\OpenWork\\คู่มือการสั่งซื้อ_ISO_Purchasing_Manual.xlsx"
wb.save(output_path)
print(f"OK - file saved")
