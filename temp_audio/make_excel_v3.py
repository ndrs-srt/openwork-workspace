import os, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

out_dir = "E:\\01.NDRS_WORKON\\OpenWork\\ถอดไฟล์เสียง\\BITEC_COLD CHAIN_2026-07-02\\"

with open(os.path.join(out_dir, "transcript", "session1_small.txt"), "r", encoding="utf-8") as f:
    s1_text = f.read()
with open(os.path.join(out_dir, "transcript", "session2_small.txt"), "r", encoding="utf-8") as f:
    s2_text = f.read()
with open(os.path.join(out_dir, "transcript", "session1_segments_small.txt"), "r", encoding="utf-8") as f:
    s1_seg = [l.rstrip('\n\r') for l in f.readlines()]
with open(os.path.join(out_dir, "transcript", "session2_segments_small.txt"), "r", encoding="utf-8") as f:
    s2_seg = [l.rstrip('\n\r') for l in f.readlines()]

wb = Workbook()

tf = Font(name="TH SarabunPSK", size=16, bold=True, color="1F4E79")
hf = Font(name="TH SarabunPSK", size=13, bold=True, color="FFFFFF")
h2f = Font(name="TH SarabunPSK", size=12, bold=True, color="1F4E79")
cf = Font(name="TH SarabunPSK", size=12)
sf = Font(name="TH SarabunPSK", size=10)
hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
afill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
gfill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
yfill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ofill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
tb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

def style_cell(cell, font=cf, fill=None, wrap=True, halign='top'):
    cell.font = font
    if fill: cell.fill = fill
    cell.alignment = Alignment(wrap_text=wrap, vertical=halign)
    cell.border = tb

def write_header_row(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=i, value=h)
        style_cell(cell, hf, hfill, True, 'center')
        ws.column_dimensions[chr(64+i) if i <= 26 else 'A'+chr(64+i-26)].width = w

def write_data_rows(ws, start_row, data_rows):
    for r, row_data in enumerate(data_rows, start_row):
        fill = afill if (r - start_row) % 2 == 1 else None
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            f = Font(name="TH SarabunPSK", size=12, bold=True) if c == 1 else cf
            style_cell(cell, f, fill, True)

# ====================================================================
# SESSION INFO
# ====================================================================
s1_title = "Session 1: โลจิสติกส์ยาและเวชภัณฑ์ — มาตรฐาน GDP, Serialization และ Temperature Monitoring"
s1_time = "11:00 – 12:00"
s1_moderator = "Mr. Ivan Sham, Founder/CEO, Red Wheels Trading"
s1_panelists = [
    "Mr. Ashwin Gurjar, Managing Director, Win Life Pharma",
    "Mr. Mark Scheepers, Founder & Innovation Leader, Thai Thermo Dynamics Co., Ltd",
    "Dr. Uttam Krishna, QA and RA Director, Pharmalytics MQ Bio Tech"
]
s1_panel_str = "\n".join(s1_panelists)

s2_title = "Session 2: ระบบอัตโนมัติควบคุมอุณหภูมิ และคลังสินค้าเย็นอัจฉริยะ"
s2_time = "13:15 – 14:15"
s2_moderator = "Mr. Pengtao Zhang, Managing Director, VentuNext"
s2_panelists = [
    "Ms. Paipin Sa-ngiam, Logistics and Quality Manager, Aviagen",
    "Dr. Uttam Krishna, QA and RA Director, Pharmalytics MQ Bio Tech",
    "Mr. Sethapong Katanyukunanon, Chief Operating Officer, Ruxchai Cold Storage Co., Ltd.",
    "Mr. Manoj Maruti Chavan, Director & CEO, Metecno Pannelli (Thailand) Co., Ltd."
]
s2_panel_str = "\n".join(s2_panelists)

# ====================================================================
# SESSION 1 TOPIC SUMMARIES
# ====================================================================
s1_topics = [
    [1, "เปิดงานและแนะนำวิทยากร",
     "Ivan Sham",
     "กล่าวเปิด: อุตสาหกรรมโลจิสติกส์เติบโตเร็ว มีกฎระเบียบเข้มงวดขึ้น โดยเฉพาะ GDP, Serialization และ Temperature Monitoring",
     "Ivan เปิดประเด็นว่าการปฏิบัติตามกฎระเบียบ (compliance), การมองเห็น (visibility), และการบริหารความเสี่ยง (risk) เป็นหัวใจของ cold chain logistics ในปัจจุบัน"],
    [2, "แนะนำตัว Ashwin Gurjar — Win Life Pharma",
     "Ashwin Gurjar",
     "CEO บริษัท one-stop service ด้านทะเบียนยา ขาย และการตลาดในไทย มีประสบการณ์ >15 ปี",
     "เน้นว่าตลาดยามีกฎระเบียบซับซ้อน Logistics เป็นปัจจัยสำคัญในการทำธุรกิจยาในไทย"],
    [3, "แนะนำตัว Mark Scheepers — Thai Thermo Dynamics",
     "Mark Scheepers",
     "ผู้เชี่ยวชาญด้าน Passive Cold Chain Solutions เริ่มจากธุรกิจอาหารสู่ Pharma",
     "การทำ cold chain ที่ดีต้อง optimize ทั้งต้นทุนและคุณภาพ — passive cooling เป็นทางเลือกที่คุ้มค่า"],
    [4, "แนะนำตัว Dr. Uttam Krishna — Pharmalytics",
     "Dr. Uttam Krishna",
     "QA/RA Director มีประสบการณ์ Pfizer และ GMP/PIC/s ด้านยา",
     "มุมมองด้านคุณภาพและกฎระเบียบ การตรวจสอบและการรับรองมาตรฐาน"],
    [5, "ภาพรวมตลาด Pharma ในไทย",
     "Ashwin Gurjar",
     "Big Pharma ครองตลาด 70% มีระบบพร้อม SME มีข้อจำกัดด้านต้นทุน compliance",
     "DKSH, Zuellig มีโครงสร้างพร้อม SME เจ้าใหม่แบกรับต้นทุน GDP ไม่ไหว"],
    [6, "GDP คือกระบวนการทั้งหมด ไม่ใช่แค่อุปกรณ์",
     "Dr. Uttam, Mark",
     "GDP ต้องครอบคลุม Supply Chain ทั้งหมด รวมถึง Transportation, SOPs, Validation",
     "คนมัก focus แค่ manufacturing แต่การกระจายและขนส่ง (2-8°C, 15-25°C) ก็ต้องได้มาตรฐาน"],
    [7, "90% ของยาเป็น Ambient, 10% เท่านั้นที่ต้อง Cold Chain",
     "Ashwin Gurjar",
     "ยาส่วนใหญ่ >25°C ได้ Cold Chain จริงๆ แค่ยาฉีดและวัคซีน",
     "ความเข้าใจผิดว่ายาทุกตัวต้อง cold chain เป็นปัญหาของการ over-invest โดยไม่จำเป็น"],
    [8, "ความท้าทายของ Data Logger",
     "Ashwin, Mark",
     "จุดติดตั้งต่างกัน, ความเสี่ยง Data Loss, Battery หมด",
     "แต่ละบริษัทมีนโยบายต่างกัน — data logger error หรือ battery หมดทำให้ไม่มี temperature history"],
    [9, "Last-mile Delivery — จุดเสี่ยงที่สุด",
     "Mark Scheepers",
     "การเปลี่ยนถ่ายหลายจุดทำให้ Temperature Break แต่ตรวจไม่พบ",
     "Smart monitoring แบบ real-time ช่วยให้ proactive แทน reactive"],
    [10, "Right First Time & Validation",
     "Dr. Uttam, Mark",
     "ต้องทำถูกตั้งแต่แรก + ระบบที่ป้องกัน human error",
     "SOP และ Procedure ต้องออกแบบให้ 'Monkey-proof' ป้องกัน human error ไม่ใช่แค่มีวินัย"],
    [11, "ระบบ Audit และ Data Integrity",
     "Ivan, Dr. Uttam",
     "Audit ต้องรวม Temperature Monitoring และ Data Integrity",
     "การ temperature abuse รู้ได้จาก monitoring เท่านั้น — ต้องตรวจสอบ process validation"],
    [12, "ประสบการณ์ Data Logger ล้มเหลวและ Backup",
     "Ashwin Gurjar",
     "การมี Data Logger Backup สามารถช่วยชีวิต",
     "ส่งสินค้าไป Latin America ใช้ 2 ตัว/ชุด ตัวแรก error ตัวที่ 2 ทำงาน Backup = สำคัญ"],
]

# ====================================================================
# SESSION 2 TOPIC SUMMARIES (corrected based on actual content)
# ====================================================================
s2_topics = [
    [1, "เปิด Session — Cold Storage Automation",
     "Pengtao Zhang (Moderator)",
     "แนะนำหัวข้อ: Cold room ต่างจาก dry warehouse — ทุกองศามีต้นทุน ทุกวินาทีที่หยุดทำงานคือความเสี่ยง",
     "Moderator เปิดประเด็น automation ใน cold storage ASEAN กำลังเติบโต — ต้องมองทั้ง efficiency และ risk reduction"],
    [2, "3 เงื่อนไขหลักสำหรับ Automation",
     "Sethapong Katanyukunanon (Ruxchai Cold Storage)",
     "Facility ที่เหมาะกับ Automation: (1) High stock storage (2) High turnover (3) Limited land/space",
     "แนะนำ 3 facility types: 1. Public cold storage (หลายลูกค้า ใหญ่) 2. Distribution Center (เร็ว แม่นยำ) 3. Food manufacturer (เชื่อม raw material→production→storage) — automation ต้องเริ่มจาก process design ก่อน"],
    [3, "มุมมองจาก Live Animals Logistics",
     "Paipin Sa-ngiam (Aviagen)",
     "Automation ไม่ใช่แค่ลดแรงงาน แต่คือคุณภาพและความสม่ำเสมอของผลิตภัณฑ์ที่มีมูลค่าสูง",
     "เน้น operation consistency สำหรับ biological/high-value products — automation ช่วยให้ deliver สินค้าสภาพดีที่สุดถึงลูกค้า"],
    [4, "GDP Compliance และ Protocol Preparation",
     "Dr. Uttam Krishna (Pharmalytics)",
     "GDP เพิ่งบังคับใช้ในไทย ต้องมี protocol สำหรับทุกผลิตภัณฑ์ ทุก storage condition",
     "ต้องเตรียม how to monitor, control, validation — กรณีต่อกรณี (oncology product ต้องการ precision สูง) sensors ต้อง calibrated และ process ต้อง validated"],
    [5, "บทบาทของ Insulation ต่อ Energy Efficiency",
     "Manoj Maruti Chavan (Metecno Pannelli)",
     "K-value ของ insulation ไม่คงที่ — อายุการใช้งาน 10-20 ปี การเลือก insulation ดีส่งผลต่อ refrigeration cost ระยะยาว",
     "หลายคน focus แค่ CAPEX ตอนสร้าง แต่ insulation เสื่อมตามเวลา → refrigeration load เพิ่ม → energy cost สูงขึ้นทุกปี การเลือก insulation คุณภาพดี = long-term savings"],
    [6, "ความท้าทายใน Cold Storage Automation",
     "Manoj, Sethapong",
     "Fog, Ice Formation, Condensation, Sensor Failure ที่อุณหภูมิต่ำ (-20°C ถึง -35°C)",
     "Sensor error, visibility ปัญหาจาก fog, condensation ทำลายระบบ electronic, GMP ต้องการสภาพแวดล้อมที่แห้ง — การออกแบบ door และ airlock สำคัญ"],
    [7, "การเปลี่ยนแปลงของบุคลากร (People Transformation)",
     "Sethapong Katanyukunanon",
     "Automation ไม่ได้แทนที่คน แต่เปลี่ยนบทบาทจาก manual → system control, problem solving",
     "ความท้าทาย: ต้องมี maintenance team พร้อม, spare parts, skill development — การเชื่อมต่อ WMS ↔ WCS ต้อง stable, ถ้า downtime → ส่งผลถึง product quality"],
    [8, "IoT Monitoring และ Real-time Visibility",
     "Paipin Sa-ngiam",
     "IoT > แค่ sensor — ต้องเข้าใจปัจจัยวิกฤตของธุรกิจ: temperature, humidity, ventilation, lighting, door opening",
     "IoT ช่วย visibility แบบ real-time โดยเฉพาะ transition point (warehouse→transport→airport) — ช่วยให้ respond ทันทีเมื่อมีปัญหา และใช้ data วิเคราะห์ improve operation"],
    [9, "IoT Validation ตามมาตรฐาน FDA 21 CFR Part 11",
     "Dr. Uttam Krishna",
     "FDA/IoT ต้อง validated — ถ้าไม่มี validation error risk สูง ไม่แนะนำให้ใช้",
     "21 CFR Part 11 กำหนด guideline สำหรับ electronic records — ถ้า IoT ไม่ผ่าน validation ไม่ควรใช้ ต้อง calibrate sensor, มีระบบ audit trail"],
    [10, "System Integration และแนวทาง ISP",
     "ทุกคน",
     "Integration ≠ เฉพาะระบบ — ต้องรวม people, process, equipment, building เข้าด้วยกัน",
     "ISP Chapter 3.2 กำหนดวิธี integrate mechanical, electrical, building infrastructure — ต้องทำ risk assessment, verification checkpoint ทุกขั้นตอน ผู้ใช้ (user) ต้องรับผิดชอบ specification และ quality"],
    [11, "ESG และ Energy Efficiency Investment",
     "Manoj, Sethapong, Paipin",
     "Energy cost เป็น main topic 2026 — การออกแบบ insulation, high-speed door, airlock, VFD, solar PV",
     "ลูกค้าพร้อมจ่ายเพิ่มถ้า warehouse มี ESG certification — automation ช่วย energy saving แต่ต้องใช้ automation ที่เหมาะกับ product ไม่ใช่ใช้ของเดียวกันหมด"],
    [12, "สรุป Session — Automation คือการเชื่อมต่อคน กระบวนการ และเทคโนโลยี",
     "Pengtao Zhang (Moderator)",
     "Automation ไม่ใช่ trend แต่เป็น necessity — ต้องถามก่อนว่า process ไหนควร automate ก่อนและวัด value ได้",
     "ผู้ใช้ต้องรับผิดชอบ requirement specification และ quality — การ integration ต้อง verification checkpoint ทุกขั้นตอน อนาคตของ cold chain = IoT + AI + API"],
]

# ====================================================================
# SHEET 1: SESSION 1 SUMMARY
# ====================================================================
ws_s1 = wb.active
ws_s1.title = "Session 1 Pharma"

# Title
ws_s1.merge_cells('A1:E1')
ws_s1['A1'] = s1_title
ws_s1['A1'].font = tf
ws_s1['A1'].alignment = Alignment(horizontal='center')

ws_s1.merge_cells('A2:E2')
ws_s1['A2'] = f"BITEC Cold Chain Conference | 02 July 2026 | {s1_time}"
ws_s1['A2'].font = Font(name="TH SarabunPSK", size=11, italic=True, color="555555")
ws_s1['A2'].alignment = Alignment(horizontal='center')

ws_s1.merge_cells('A4:E4')
ws_s1['A4'] = f"Moderator: {s1_moderator}\nPanellists:\n{s1_panel_str}"
ws_s1['A4'].font = Font(name="TH SarabunPSK", size=11, bold=True)
ws_s1['A4'].alignment = Alignment(wrap_text=True, vertical='top')

write_header_row(ws_s1, 6, ["#", "หัวข้อ", "วิทยากร", "ประเด็นสำคัญ", "รายละเอียด"], [5, 22, 20, 35, 65])
write_data_rows(ws_s1, 7, s1_topics)

# Full transcript
sr = 7 + len(s1_topics) + 1
ws_s1.merge_cells(f'A{sr}:E{sr}')
ws_s1.cell(row=sr, column=1, value="บทถอดเสียงเต็ม (Full Transcript)").font = Font(name="TH SarabunPSK", size=14, bold=True, color="1F4E79")
sr += 1
ws_s1.merge_cells(f'A{sr}:E{sr}')
ws_s1.cell(row=sr, column=1, value=s1_text).font = sf
ws_s1.cell(row=sr, column=1).alignment = Alignment(wrap_text=True)

# ====================================================================
# SHEET 2: SESSION 2 SUMMARY
# ====================================================================
ws_s2 = wb.create_sheet("Session 2 Cold Storage")
ws_s2.merge_cells('A1:E1')
ws_s2['A1'] = s2_title
ws_s2['A1'].font = tf
ws_s2['A1'].alignment = Alignment(horizontal='center')

ws_s2.merge_cells('A2:E2')
ws_s2['A2'] = f"BITEC Cold Chain Conference | 02 July 2026 | {s2_time}"
ws_s2['A2'].font = Font(name="TH SarabunPSK", size=11, italic=True, color="555555")
ws_s2['A2'].alignment = Alignment(horizontal='center')

ws_s2.merge_cells('A4:E4')
ws_s2['A4'] = f"Moderator: {s2_moderator}\nPanellists:\n{s2_panel_str}"
ws_s2['A4'].font = Font(name="TH SarabunPSK", size=11, bold=True)
ws_s2['A4'].alignment = Alignment(wrap_text=True, vertical='top')

write_header_row(ws_s2, 6, ["#", "หัวข้อ", "วิทยากร", "ประเด็นสำคัญ", "รายละเอียด"], [5, 22, 20, 35, 65])
write_data_rows(ws_s2, 7, s2_topics)

sr = 7 + len(s2_topics) + 1
ws_s2.merge_cells(f'A{sr}:E{sr}')
ws_s2.cell(row=sr, column=1, value="บทถอดเสียงเต็ม (Full Transcript)").font = Font(name="TH SarabunPSK", size=14, bold=True, color="1F4E79")
sr += 1
ws_s2.merge_cells(f'A{sr}:E{sr}')
ws_s2.cell(row=sr, column=1, value=s2_text).font = sf
ws_s2.cell(row=sr, column=1).alignment = Alignment(wrap_text=True, vertical='top')

# ====================================================================
# SHEET 3: SESSION 1 TIMESTAMPS
# ====================================================================
ws_t1 = wb.create_sheet("Timestamps S1")
write_header_row(ws_t1, 1, ["เวลา (นาที)", "เนื้อหา"], [14, 120])
for r, line in enumerate(s1_seg, 2):
    line = line.strip()
    if not line: continue
    parts = line.split("] ", 1)
    if len(parts) == 2:
        ws_t1.cell(row=r, column=1, value=parts[0].lstrip("["))
        ws_t1.cell(row=r, column=2, value=parts[1])
    for c in [1,2]:
        ws_t1.cell(row=r, column=c).font = Font(name="TH SarabunPSK", size=9)
        ws_t1.cell(row=r, column=c).border = tb
        ws_t1.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical='top')

# ====================================================================
# SHEET 4: SESSION 2 TIMESTAMPS
# ====================================================================
ws_t2 = wb.create_sheet("Timestamps S2")
write_header_row(ws_t2, 1, ["เวลา (นาที)", "เนื้อหา"], [14, 120])
for r, line in enumerate(s2_seg, 2):
    line = line.strip()
    if not line: continue
    parts = line.split("] ", 1)
    if len(parts) == 2:
        ws_t2.cell(row=r, column=1, value=parts[0].lstrip("["))
        ws_t2.cell(row=r, column=2, value=parts[1])
    for c in [1,2]:
        ws_t2.cell(row=r, column=c).font = Font(name="TH SarabunPSK", size=9)
        ws_t2.cell(row=r, column=c).border = tb
        ws_t2.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical='top')

# ====================================================================
# SHEET 5: SESSION 1 KEY TAKEAWAYS (standalone)
# ====================================================================
ws5 = wb.create_sheet("S1 Key Takeaways")
ws5.merge_cells('A1:B1')
ws5['A1'] = "Session 1 — โลจิสติกส์ยาและเวชภัณฑ์: มาตรฐาน GDP, Serialization, Temperature Monitoring"
ws5['A1'].font = tf
ws5['A1'].alignment = Alignment(horizontal='center')

ws5.merge_cells('A2:B2')
ws5['A2'] = f"{s1_time} | Moderator: {s1_moderator}"
ws5['A2'].font = Font(name="TH SarabunPSK", size=11, italic=True, color="555555")
ws5['A2'].alignment = Alignment(horizontal='center')

takeaways_s1 = [
    ["ประเด็น", "รายละเอียด"],
    ["GDP Compliance", "ยาที่ต้อง cold chain จริงๆ มีแค่ ~10% (ยาฉีด วัคซีน) — 90% เก็บที่ ambient ได้ ต้องใช้ Risk-based Approach (ICH Q9) ไม่ใช่ over-invest"],
    ["Serialization", "การตรวจสอบย้อนกลับเป็นสิ่งจำเป็น โดยเฉพาะสำหรับ pharma — ต้องมี Data Integrity และ Audit Trail ที่สมบูรณ์"],
    ["Temperature Monitoring", "Data Logger ต้องมี Backup — Ashwin เล่าประสบการณ์ส่ง Latin America ตัวแรก error ตัว backup ช่วยชีวิต ต้องมี Real-time Monitoring แทน Reactive"],
    ["Last-mile Delivery", "จุดเสี่ยงที่สุดใน cold chain — การเปลี่ยนถ่ายสินค้าหลายจุดอาจ temperature break โดยที่ปลายทางตรวจไม่พบ ต้องใช้ Smart Monitoring"],
    ["SME Challenges", "Big Pharma ครอง 70% ของตลาด — SME รับภาระต้นทุน GDP สูง (ค่าแรงไทย ~$1,500/เดือน สูงสุดใน ASEAN) ต้องการ Shared Infrastructure"],
    ["Right First Time", "SOP ต้องออกแบบให้ 'Monkey-proof' ป้องกัน human error — Validation ต้องครอบคลุมทั้ง Equipment, Process, และ People"],
    ["Closing", "GDP ไม่ใช่ยาวิเศษ แต่เป็นเครื่องมือพัฒนาธุรกิจ — การใช้เทคโนโลยี + ระบบที่ดี + Training = Cold Chain ที่มีประสิทธิภาพ"],
]

for r, (h, v) in enumerate(takeaways_s1, 4):
    for c, val in enumerate([h, v], 1):
        cell = ws5.cell(row=r, column=c, value=val)
        if r == 4:
            style_cell(cell, hf, hfill, True, 'center')
        elif c == 1:
            style_cell(cell, Font(name="TH SarabunPSK", size=12, bold=True, color="1F4E79"), gfill, True, 'top')
        else:
            style_cell(cell, cf, afill if (r-4)%2==1 else None, True, 'top')

ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 85

# ====================================================================
# SHEET 6: SESSION 2 KEY TAKEAWAYS (standalone)
# ====================================================================
ws6 = wb.create_sheet("S2 Key Takeaways")
ws6.merge_cells('A1:B1')
ws6['A1'] = "Session 2 — ระบบอัตโนมัติควบคุมอุณหภูมิ และคลังสินค้าเย็นอัจฉริยะ"
ws6['A1'].font = tf
ws6['A1'].alignment = Alignment(horizontal='center')

ws6.merge_cells('A2:B2')
ws6['A2'] = f"{s2_time} | Moderator: {s2_moderator}"
ws6['A2'].font = Font(name="TH SarabunPSK", size=11, italic=True, color="555555")
ws6['A2'].alignment = Alignment(horizontal='center')

takeaways_s2 = [
    ["ประเด็น", "รายละเอียด"],
    ["Automation Criteria", "3 เงื่อนไข: High stock storage + High turnover + Limited space — Public cold storage, DC, Food manufacturer มี business case แรงที่สุด"],
    ["Cold Storage Challenges", "Fog, Ice formation, Condensation ที่อุณหภูมิต่ำ (-20°C ถึง -35°C) ทำให้ sensor error, visibility ปัญหา — การออกแบบ airlock และ door สำคัญ"],
    ["Insulation & K-value", "K-value ไม่คงที่ — insulation เสื่อมตามเวลา → refrigeration load เพิ่มขึ้นทุกปี → energy cost สูง ต้องเลือก insulation คุณภาพดีตั้งแต่วันแรก (CAPEX > OPEX)"],
    ["Energy Efficiency", "High-speed door, VFD, Airlock system, Solar PV — ลูกค้ายินดีจ่ายเพิ่มถ้า warehouse มี ESG Certification"],
    ["IoT Monitoring", "IoT > เซนเซอร์ — ต้องเข้าใจปัจจัยวิกฤตของธุรกิจ (temperature, humidity, ventilation, lighting, door opening) ช่วย Real-time visibility โดยเฉพาะ transition point"],
    ["IoT Validation", "FDA 21 CFR Part 11 — IoT ต้อง validated ถ้าไม่มี chances of error สูง — ต้อง calibrate sensor, audit trail, qualification"],
    ["System Integration", "ISP Chapter 3.2 — ต้อง integrate mechanical, electrical, building infrastructure + People Transformation (manual→system control) ต้องมี verification checkpoint ทุกขั้นตอน"],
    ["Closing", "Automation ≠ trend แต่ necessity — User ต้องรับผิดชอบ Requirement Specification และ Quality — อนาคต = IoT + AI + API"],
]

for r, (h, v) in enumerate(takeaways_s2, 4):
    for c, val in enumerate([h, v], 1):
        cell = ws6.cell(row=r, column=c, value=val)
        if r == 4:
            style_cell(cell, hf, hfill, True, 'center')
        elif c == 1:
            style_cell(cell, Font(name="TH SarabunPSK", size=12, bold=True, color="1F4E79"), gfill, True, 'top')
        else:
            style_cell(cell, cf, afill if (r-4)%2==1 else None, True, 'top')

ws6.column_dimensions['A'].width = 25
ws6.column_dimensions['B'].width = 85

# ====================================================================
# SHEET 7: ภาพรวมการประชุม (2 sessions แยกกัน คนละหัวข้อ)
# ====================================================================
ws7 = wb.create_sheet("ภาพรวมการประชุม")
ws7.merge_cells('A1:B1')
ws7['A1'] = "BITEC Cold Chain Conference | 02 July 2026"
ws7['A1'].font = tf
ws7['A1'].alignment = Alignment(horizontal='center')

ov = [
    ["งาน", "BITEC Cold Chain Conference 2026"],
    ["วันที่", "02 July 2026"],
    ["สถานที่", "BITEC, Bangkok"],
]
for r, (h, v) in enumerate(ov, 3):
    for c, val in enumerate([h, v], 1):
        cell = ws7.cell(row=r, column=c, value=val)
        cell.font = Font(name="TH SarabunPSK", size=11, bold=c==1)
        cell.border = tb

r0 = 3 + len(ov) + 1
ws7.merge_cells(f'A{r0}:B{r0}')
ws7.cell(row=r0, column=1, value="Session 1: Pharma Logistics").font = h2f
r0 += 1
s1o = [
    ["เวลา", s1_time],
    ["หัวข้อ", s1_title],
    ["Moderator", s1_moderator],
    ["Panelists", s1_panel_str],
]
for r, (h, v) in enumerate(s1o, r0):
    for c, val in enumerate([h, v], 1):
        cell = ws7.cell(row=r, column=c, value=val)
        cell.font = Font(name="TH SarabunPSK", size=11, bold=c==1)
        cell.border = tb; cell.alignment = Alignment(wrap_text=True, vertical='top')

r0 = r0 + len(s1o) + 1
ws7.merge_cells(f'A{r0}:B{r0}')
ws7.cell(row=r0, column=1, value="Session 2: Cold Storage Automation").font = h2f
r0 += 1
s2o = [
    ["เวลา", s2_time],
    ["หัวข้อ", s2_title],
    ["Moderator", s2_moderator],
    ["Panelists", s2_panel_str],
]
for r, (h, v) in enumerate(s2o, r0):
    for c, val in enumerate([h, v], 1):
        cell = ws7.cell(row=r, column=c, value=val)
        cell.font = Font(name="TH SarabunPSK", size=11, bold=c==1)
        cell.border = tb; cell.alignment = Alignment(wrap_text=True, vertical='top')

ws7.column_dimensions['A'].width = 40
ws7.column_dimensions['B'].width = 80

# ====================================================================
# SHEET 8 : TOPIC INDEX
# ====================================================================
ws8 = wb.create_sheet("สารบัญหัวข้อทั้งหมด")
write_header_row(ws8, 1, ["Session", "#", "หัวข้อ", "วิทยากร", "เวลา"], [8, 4, 30, 20, 14])

all_topics = []
for t in s1_topics:
    all_topics.append(["S1", t[0], t[1], t[2], s1_time])
for t in s2_topics:
    all_topics.append(["S2", t[0], t[1], t[2], s2_time])

write_data_rows(ws8, 2, all_topics)

# ====================================================================
# SAVE
# ====================================================================
xlsx_path = os.path.join(out_dir, "Cold_Chain_Summary_v3_TH.xlsx")
wb.save(xlsx_path)
print("Saved:", xlsx_path)
