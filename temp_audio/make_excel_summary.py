import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

out_dir = "E:\\01.NDRS_WORKON\\OpenWork\\ถอดไฟล์เสียง\\BITEC_COLD CHAIN_2026-07-02\\"

# Read transcripts
with open(os.path.join(out_dir, "transcript", "session1_transcript_en.txt"), "r", encoding="utf-8") as f:
    s1 = f.read()
with open(os.path.join(out_dir, "transcript", "session2_transcript_en.txt"), "r", encoding="utf-8") as f:
    s2 = f.read()

# Read segments
with open(os.path.join(out_dir, "transcript", "session1_segments_en.txt"), "r", encoding="utf-8") as f:
    s1_seg = f.read()
with open(os.path.join(out_dir, "transcript", "session2_segments_en.txt"), "r", encoding="utf-8") as f:
    s2_seg = f.read()

wb = Workbook()

# ============= SHEET 1: Summary =============
ws = wb.active
ws.title = "สรุปภาพรวม"

# Styles
title_font = Font(name="TH SarabunPSK", size=16, bold=True, color="1F4E79")
header_font = Font(name="TH SarabunPSK", size=13, bold=True, color="FFFFFF")
content_font = Font(name="TH SarabunPSK", size=12)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Title
ws.merge_cells('A1:D1')
ws['A1'] = "บทสรุปการอภิปรายกลุ่ม: โลจิสติกส์ยาและเวชภัณฑ์"
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:D2')
ws['A2'] = "BITEC Cold Chain Conference | 02 July 2026 | 11:00-12:00 & 13:15-14:14"
ws['A2'].font = Font(name="TH SarabunPSK", size=11, italic=True, color="555555")
ws['A2'].alignment = Alignment(horizontal='center')

# Panelists
ws.merge_cells('A4:D4')
ws['A4'] = "วิทยากร: Mr. Ivan Sham (Moderator, Red Wheels Trading) | Mr. Ashwin Gurjar (Win life pharma) | Mr. Mark Scheepers (Thai Thermo Dynamics) | Dr. Uttam Krishna (Pharmalytics MQ Bio Tech)"
ws['A4'].font = Font(name="TH SarabunPSK", size=11, bold=True)
ws['A4'].alignment = Alignment(wrap_text=True)

# Headers
headers = ["หัวข้อ", "วิทยากร", "ประเด็นสำคัญ", "รายละเอียด"]
col_widths = [25, 22, 35, 60]
for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=6, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    ws.column_dimensions[chr(64+i)].width = w

# Data
data = [
    [
        "แนะนำวิทยากรและที่มา",
        "Ivan Sham, Ashwin Gurjar, Mark Scheepers, Dr. Uttam Krishna",
        "วิทยากร 4 ท่านจาก Malaysia, India, Thailand แนะนำตัวและประสบการณ์",
        "Ivan Sham (Red Wheels) เชี่ยวชาญ冷链ขนส่ง since 2013 | Ashwin Gurjar (Win Life Pharma) บริการครบวงจรสำหรับ Pharma ในไทย 1.5 ทศวรรษ | Mark Scheepers (Thai Thermo Dynamics) เติบโตจากธุรกิจอาหารสู่ Phama cold chain solutions | Dr. Uttam Krishna (Pharmalytics) ผู้เชี่ยวชาญ GDP และเคยทำงานกับ Pfizer"
    ],
    [
        "ภาพรวมตลาด Pharma ในไทย",
        "Ashwin Gurjar",
        "ตลาด Pharma ไทยแบ่งเป็น Big Pharma (~70%) และ SME (~30%)",
        "บริษัทใหญ่เช่น DKSH, Zuellig Pharma มีโครงสร้างพื้นฐานดี มีระบบและ procedure ชัดเจน ส่วน SME มักเผชิญความท้าทายด้านต้นทุน เนื่องจากค่าใช้จ่ายสูงในการปฏิบัติตามมาตรฐานต่าง ๆ"
    ],
    [
        "GDP และการปฏิบัติตามมาตรฐาน",
        "Dr. Uttam Krishna, Mark Scheepers",
        "GDP คือกระบวนการทำงาน ไม่ใช่แค่การมีอุปกรณ์",
        "การปฏิบัติตาม GDP ต้องครอบคลุมทั้งกระบวนการผลิตจนถึงขนส่ง โดยเฉพาะจุดที่มักพลาดคือการขนส่ง (transportation) ที่มีโซ่ความเย็นขาดตอน ต้องมี SOP, validation, และ temperature monitoring ที่เหมาะสม"
    ],
    [
        "สัดส่วนผลิตภัณฑ์ที่ต้องควบคุมอุณหภูมิ",
        "Ashwin Gurjar",
        "90% เป็นสินค้าทั่วไป (อุณหภูมิห้อง <25°C) มี 10% เท่านั้นที่เป็น cold chain",
        "ใน Cold chain 10% ส่วนใหญ่มีข้อกำหนด 2-8°C หรือ 15-25°C โดยเฉพาะยาฉีด (injections) และวัคซีน (vaccines) ที่มีความไวต่ออุณหภูมิสูง"
    ],
    [
        "ความท้าทายในการติดตามอุณหภูมิ",
        "Ashwin Gurjar, Mark Scheepers",
        "การติดตั้ง data logger มีหลายแนวทาง ขึ้นกับนโยบายแต่ละบริษัท",
        "บางบริษัทติดตั้งตอน container เตรียมออก บางแห่งติดตั้งตอน dispatch ปัญหาคือเมื่อใดที่ data logger ไม่ทำงานหรือ error จะไม่มีข้อมูลแสดง Temperature excursion ซึ่งเป็น red flag ในการ audit"
    ],
    [
        "Last-mile Delivery และ Awareness",
        "Mark Scheepers",
        "ปัญหาใหญ่คือการขนส่งช่วงสุดท้าย (last-mile) ที่ขาด awareness",
        "การขนส่งจากรถใหญ่ไปรถเล็ก การเปลี่ยนถ่ายสินค้าหลายจุดเพิ่มความเสี่ยง การมี temperature break ระหว่างทางแต่ปลายทางวัดได้อุณหภูมิปกติ => ความเสียหายถูกซ่อนเร้น ผู้รับไม่รู้ว่าสินค้าเสียหายระหว่างทาง"
    ],
    [
        "Right First Time & Validation",
        "Dr. Uttam Krishna",
        "หลักการสำคัญ: ทำถูกต้องตั้งแต่ครั้งแรก (right first time) และการ validation",
        "การทำ right first time ช่วยลดความเสี่ยง โดยเฉพาะใน Pharma ที่ผลกระทบถึงชีวิตผู้ป่วย การ validation ต้องทำทั้งกระบวนการรวมถึง transportation validation"
    ],
    [
        "Proactive vs Reactive Monitoring",
        "Mark Scheepers",
        "ต้องเปลี่ยนจาก Reactive (ดูข้อมูลย้อนหลัง) เป็น Proactive",
        "การใช้ smart tags / IoT ช่วยให้เห็นความผิดปกติ实时 สามารถจัดการ exception ได้ทันทีก่อนสินค้าถึงมือผู้ป่วย แทนที่จะมาเห็นข้อมูลย้อนหลังเมื่อสายเกินไป",
    ],
    [
        "คน vs เทคโนโลยี",
        "Ashwin Gurjar, Mark Scheepers",
        "ต้องออกแบบระบบให้ 'Monkey-proof' ป้องกันความผิดพลาดของมนุษย์",
        "การมี SOP และการฝึกอบรมพนักงาน至关重要 แต่ต้องออกแบบระบบที่ไม่เปิดช่องให้เกิด error ต้องมี data logger, cloud solutions, real-time monitoring เป็นตัวช่วย"
    ],
    [
        "Audit และการตรวจสอบ",
        "Ivan Sham, Dr. Uttam Krishna",
        "การ Audit ใน Pharma logistics ต้องละเอียด ครอบคลุมทุกจุด",
        "Audit รวมถึงการตรวจสอบ temperature monitoring, data integrity, การปฏิบัติตาม GDP, การตรวจสอบ冷链 logistics process และการตรวจสอบย้อนกลับ (traceability)"
    ],
    [
        "ต้นทุนและ SME",
        "Ashwin Gurjar, Mark Scheepers",
        "มาตรฐาน GDP มีค่าใช้จ่ายสูง โดยเฉพาะสำหรับ SME",
        "ค่าแรงในประเทศไทยสูงกว่าใน ASEAN อื่น ๆ ทำให้ SME แบกรับต้นทุนไม่ไหว ต้องมีโซลูชันแบบ subscription หรือ shared infrastructure เพื่อให้ SME เข้าถึงเทคโนโลยี cold chain ได้"
    ],
    [
        "การขนส่งข้ามประเทศ",
        "Ashwin Gurjar, Mark Scheepers, Dr. Uttam Krishna",
        "ความท้าทายของการขนส่ง Pharma ข้ามประเทศใน ASEAN",
        "แต่ละประเทศมีข้อกำหนดต่างกัน การจัดการ temperature excursion, การตรวจสอบที่ด่านศุลกากร, การใช้ data logger ซ้ำซ้อน (backup) เพื่อความมั่นใจ บริษัทอย่าง DKSH มีระบบข้ามประเทศที่ดี"
    ],
    [
        "อนาคตของ Cold Chain ในไทย",
        "ทุกคน",
        "ต้องใช้เทคโนโลยีและนวัตกรรมลดต้นทุน เพิ่มประสิทธิภาพ",
        "แนวโน้ม: IoT sensors, cloud-based monitoring, shared infrastructure, subscription model เพื่อให้ทุกขนาดธุรกิจเข้าถึง cold chain ได้ Thailand มี potential แต่ต้องพัฒนา framework และโครงสร้างพื้นฐานร่วมกัน"
    ],
]

for r, row_data in enumerate(data, 7):
    fill = alt_fill if (r-7) % 2 == 1 else None
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = content_font
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border
        if fill:
            cell.fill = fill

ws.row_dimensions[6].height = 30
for r in range(7, 7+len(data)):
    ws.row_dimensions[r].height = 55

# ============= SHEET 2: Transcript EN =============
ws2 = wb.create_sheet("Transcript (English)")

ws2.cell(row=1, column=1, value="Session 1: 11:00-12:00").font = Font(bold=True, size=14)
ws2.merge_cells('A1:B1')
ws2.cell(row=2, column=1, value=s1).font = Font(size=10)
ws2.column_dimensions['A'].width = 200
ws2.cell(row=2, column=1).alignment = Alignment(wrap_text=True)

ws2.cell(row=4, column=1, value="Session 2: 13:15-14:14").font = Font(bold=True, size=14)
ws2.merge_cells('A4:B4')
ws2.cell(row=5, column=1, value=s2).font = Font(size=10)
ws2.cell(row=5, column=1).alignment = Alignment(wrap_text=True)

# ============= SHEET 3: Segments with Timestamps =============
ws3 = wb.create_sheet("Timestamps (Session 1)")

for c, h in enumerate(["เวลา", "เนื้อหา"], 1):
    cell = ws3.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
ws3.column_dimensions['A'].width = 28
ws3.column_dimensions['B'].width = 120

for r, line in enumerate(s1_seg.split("\n"), 2):
    if not line.strip():
        continue
    parts = line.split("] ", 1)
    if len(parts) == 2:
        ws3.cell(row=r, column=1, value=parts[0].strip("["))
        ws3.cell(row=r, column=2, value=parts[1].strip())
    for c in [1, 2]:
        ws3.cell(row=r, column=c).font = Font(size=9)
        ws3.cell(row=r, column=c).border = thin_border
        ws3.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical='top')

# Save
xlsx_path = os.path.join(out_dir, "Cold_Chain_Summary_TH.xlsx")
wb.save(xlsx_path)
import sys
sys.stdout.reconfigure(encoding='utf-8')
print("Saved:", xlsx_path)
