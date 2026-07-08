import os, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

out_dir = "E:\\01.NDRS_WORKON\\OpenWork\\ถอดไฟล์เสียง\\BITEC_COLD CHAIN_2026-07-02\\"

with open(os.path.join(out_dir, "transcript", "session1_small.txt"), "r", encoding="utf-8") as f:
    s1_text = f.read()
with open(os.path.join(out_dir, "transcript", "session2_small.txt"), "r", encoding="utf-8") as f:
    s2_text = f.read()
with open(os.path.join(out_dir, "transcript", "session1_segments_small.txt"), "r", encoding="utf-8") as f:
    s1_seg = f.readlines()
with open(os.path.join(out_dir, "transcript", "session2_segments_small.txt"), "r", encoding="utf-8") as f:
    s2_seg = f.readlines()

wb = Workbook()

tf = Font(name="TH SarabunPSK", size=16, bold=True, color="1F4E79")
hf = Font(name="TH SarabunPSK", size=13, bold=True, color="FFFFFF")
cf = Font(name="TH SarabunPSK", size=12)
hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
afill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
gfill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
tb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

def style_cell(cell, font=cf, fill=None, wrap=True, align='top'):
    cell.font = font
    if fill: cell.fill = fill
    cell.alignment = Alignment(wrap_text=wrap, vertical=align)
    cell.border = tb

def write_sheet(ws, title, session_label, datetime_str, text, segments, data_rows, speaker_info):
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = title
    ws['A1'].font = tf
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:E2')
    ws['A2'] = f"BITEC Cold Chain Conference | 02 July 2026 | {session_label}"
    ws['A2'].font = Font(name="TH SarabunPSK", size=11, italic=True, color="555555")
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A4:E4')
    ws['A4'] = speaker_info
    ws['A4'].font = Font(name="TH SarabunPSK", size=11, bold=True)
    ws['A4'].alignment = Alignment(wrap_text=True)

    # Summary table
    ws.cell(row=6, column=1, value="ลำดับ").font = hf
    ws.cell(row=6, column=1).fill = hfill
    ws.cell(row=6, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=6, column=1).border = tb
    headers = ["หัวข้อ", "วิทยากร", "ประเด็นสำคัญ", "รายละเอียด"]
    widths = [7, 22, 22, 35, 60]
    for i, (h, w) in enumerate(zip(headers, widths), 2):
        cell = ws.cell(row=6, column=i, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = tb
        ws.column_dimensions[chr(64+i-1) if i <= 26 else 'A'+chr(64+i-26)].width = w
    ws.column_dimensions['A'].width = 7

    for r, row_data in enumerate(data_rows, 7):
        fill = afill if (r-7) % 2 == 1 else None
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            font = Font(name="TH SarabunPSK", size=12, bold=True) if c == 1 else cf
            style_cell(cell, font, fill)
            if c == 1:
                cell.alignment = Alignment(horizontal='center', vertical='top')
    ws.row_dimensions[6].height = 30

    start_row = 7 + len(data_rows) + 1

    # Full transcript
    ws.merge_cells(f'A{start_row}:E{start_row}')
    ws.cell(row=start_row, column=1, value="บทถอดเสียงเต็ม (Full Transcript)").font = Font(name="TH SarabunPSK", size=14, bold=True, color="1F4E79")
    start_row += 1
    ws.merge_cells(f'A{start_row}:E{start_row}')
    ws.cell(row=start_row, column=1, value=text).font = Font(name="TH SarabunPSK", size=10)
    ws.cell(row=start_row, column=1).alignment = Alignment(wrap_text=True)

# ===== DATA =====

s1_topics = [
    [1, "เปิดเวทีและแนะนำวิทยากร", "Ivan Sham (Moderator)",
     "แนะนำหัวข้อ GDP Compliance, Serialization, Temperature Monitoring",
     "Ivan Sham กล่าวเปิด: อุตสาหกรรม Logistics กำลังเติบโตเร็ว มีกฎระเบียบเข้มงวดขึ้น โดยเฉพาะด้าน GDP, การติดตามอุณหภูมิ และ Serialization แนะนำ panelists ทั้ง 4 ท่าน"],
    [2, "แนะนำตัว Ashwin Gurjar (Win Life Pharma)", "Ashwin Gurjar",
     "CEO Win Life Pharma ให้บริการครบวงจร 1.5 ทศวรรษในไทย",
     "บริษัท one-stop service ด้าน registration, sales, marketing สำหรับ Pharma ในไทย มีความรู้และประสบการณ์ตลาดยาไทย"],
    [3, "แนะนำตัว Mark Scheepers (Thai Thermo Dynamics)", "Mark Scheepers",
     "ผู้เชี่ยวชาญด้าน Passive Cold Chain Solutions",
     "เริ่มจากธุรกิจ food logistics แล้วพัฒนาไปสู่ passive cold chain solutions สำหรับ Pharma โดยเน้น Optimize ต้นทุนและคุณภาพ"],
    [4, "แนะนำตัว Dr. Uttam Krishna (Pharmalytics)", "Dr. Uttam Krishna",
     "QA/RA Director มีประสบการณ์ Pfizer และเป็น Registered RDP สำหรับ Cannabis",
     "ทำงานกับ Pfizer, เป็นผู้เชี่ยวชาญด้าน GMP/PIC/s และทำงานกับผู้ผลิต Pharma ทุกราย"],
    [5, "ภาพรวมตลาด Pharma ในไทย", "Ashwin Gurjar",
     "Big Pharma ครอง 70% มีระบบดี แต่ SME มีข้อจำกัดด้านต้นทุน",
     "บริษัทใหญ่เช่น DKSH, Zuellig มีโครงสร้าง Infrastructure พร้อม ส่วน SME เจ้าใหม่ประสบปัญหาต้นทุนสูงในการปฏิบัติตามข้อกำหนด"],
    [6, "GDP คือกระบวนการ ไม่ใช่แค่อุปกรณ์", "Dr. Uttam Krishna, Mark Scheepers",
     "GDP ต้องครอบคลุม Supply Chain ทั้งหมด โดยเฉพาะ Transportation",
     "คนมัก focus ที่ Manufacturing แต่ GDP จริงๆ ต้องดูทั้ง Distribution และ Transportation เช่น เงื่อนไขการจัดเก็บ 2-8°C, 15-25°C รวมถึง SOPs, Validation, Equipment"],
    [7, "90% Pharma เป็น Ambient, 10% Cold Chain", "Ashwin Gurjar",
     "Cold Chain จริงๆ มีแค่ 10% โดยเฉพาะ Injections และ Vaccines",
     "90% ของ Pharma เก็บที่อุณหภูมิห้อง (<25°C) ได้ ส่วนที่ต้อง Cold Chain 2-8°C หรือ 15-25°C ส่วนใหญ่เป็นยาฉีดและวัคซีน"],
    [8, "ความท้าทายของ Data Logger", "Ashwin Gurjar, Mark Scheepers",
     "ปัญหา: จุดติดตั้ง Data Logger ต่างกัน และความเสี่ยง Data Loss",
     "แต่ละบริษัทมีนโยบายการติดตั้ง data logger ต่างกัน บางแห่งติดก่อน container ออก บางแห่งติดตอน dispatch ปัญหาคือ data logger อาจ error หรือ battery หมด ทำให้ไม่มีข้อมูล temperature"],
    [9, "Last-mile Delivery จุดเสี่ยงที่สุด", "Mark Scheepers",
     "Awareness และการเปลี่ยนถ่ายสินค้าหลายจุดเพิ่มความเสี่ยง",
     "การขนส่งจากรถใหญ่ไปรถเล็ก (last-mile) มีหลายจุดที่ temperature break เกิดขึ้นได้แต่ปลายทางวัดได้ปกติ = ความเสียหายถูกซ่อนเร้น การมี smart monitoring แบบ real-time ช่วยให้ proactive แทน reactive"],
    [10, "Right First Time & Validation", "Dr. Uttam Krishna, Mark Scheepers",
     "ทำถูกครั้งแรก + ระบบ Monkey-proof ป้องกัน human error",
     "หลักการ Right First Time และ Validation สำคัญมาก ต้องออกแบบ SOP และ Procedure ที่ 'Monkey-proof' = ป้องกัน human error ไม่ใช่แค่มีวินัย แต่ต้องมีระบบควบคุมที่ดี"],
    [11, "ระบบ Audit และการตรวจสอบ", "Ivan Sham, Dr. Uttam Krishna",
     "Audit ต้องครอบคลุม Temperature Monitoring และ Data Integrity",
     "การขนส่งมีหลาย process การ temperature abuse จะรู้ได้ด้วยการ monitoring เท่านั้น Audit ต้องรวมถึง data integrity, การปฏิบัติตาม GDP, และ process validation"],
    [12, "ประสบการณ์ Data Logger ล้มเหลวและ Backup", "Ashwin Gurjar",
     "การมี Data Logger ตัวที่ 2 (Backup) สามารถช่วยชีวิตได้",
     "เล่าประสบการณ์ส่งสินค้าไป Latin America ใช้ data logger 2 ตัวต่อชุด ตัวแรก error แต่ตัวที่ 2 ทำงานได้ — backup สำคัญมากเพราะ human error มีได้เสมอ"],
]

s2_topics = [
    [1, "Digital Temperature Monitoring", "Ivan Sham, Mark Scheepers",
     "Data Logger เป็นมาตรฐานแต่มีข้อจำกัด ต้องใช้ Cloud Solution",
     "เล่าเรื่อง Chris จาก Freshport ที่ data logger ข้อมูลหายเพราะ laptop เสีย ทำให้ไม่มีข้อมูลแสดง client → ชี้ให้เห็นความสำคัญของ Cloud-based real-time monitoring แทนการพึ่งพา local data logger"],
    [2, "Stability Data และ Temperature Excursion", "Ashwin Gurjar",
     "ผลิตภัณฑ์ส่วนใหญ่มี Stability Data รองรับ 24 ชม.นอกเหนืออุณหภูมิ",
     "90% ของมียังคง stable หาก out of temperature ไม่เกิน 24 ชม. แต่ไม่มีใครกล้าเสี่ยง โดยเฉพาะยาฉีด — ผลกระทบอาจไม่เห็นทันทีแต่จะแสดงผลระยะหลัง ทำให้แยกยากว่ายาเสียหรือ process การจัดการเสีย"],
    [3, "Reactive → Proactive Monitoring", "Mark Scheepers",
     "ต้องเปลี่ยนจากการดูข้อมูลย้อนหลังเป็นการแจ้งเตือน实时",
     "ปัจจุบันการ monitoring ส่วนใหญ่เป็น reactive (ดูย้อนหลังเมื่อถึงปลายทาง) แต่ควรเป็น proactive ด้วย IoT smart tags, cloud-based alerts เพื่อจัดการ exception ได้ทันที"],
    [4, "Investment ใน Cold Chain", "ทุกคน",
     "Investment ขึ้นกับ Product Risk-based Approach ตาม ICH Q9",
     "ต้องใช้ Risk-based Approach (ICH Q9) แทนการทำทุกอย่าง เพราะแพงเกินไป เช่น Amlodipine มี stability data ชัดเจน — investment ต้องเหมาะกับความเสี่ยงของผลิตภัณฑ์"],
    [5, "ต้นทุนแรงงานไทยสูงที่สุดใน ASEAN", "Ashwin Gurjar",
     "ค่าแรงคนไทยสูงเมื่อเทียบกับเวียดนาม มาเลเซีย เมียนมา ฟิลิปปินส์",
     "Thailand มี labor cost สูงมากเมื่อเทียบกับ ASEAN peers — มีผลต่อต้นทุนการทำ GDP compliance โดยเฉพาะ SME ที่รับภาระไม่ไหว"],
    [6, "SME กับภาระต้นทุน GDP", "Ashwin Gurjar",
     "กฎระเบียบที่เหมือนกันสำหรับทุกขนาดธุรกิจไม่เป็นธรรมกับ SME",
     "SME ถูกกดดันให้ทำ GDP ทั้งที่สินค้าเป็น ambient (อุณหภูมิห้อง) ไม่จำเป็นต้อง cold chain ต้นทุน personel สูง การจ้าง QA/RA 1 คนในไทย ~$1,500/เดือน เทียบกับ $500 ในประเทศเพื่อนบ้าน"],
    [7, "Shared Infrastructure สำหรับ Cold Chain", "Mark Scheepers",
     "ต้องมี Shared Solution แบบ Subscription เพื่อลดต้นทุน SME",
     "แทนที่ SME แต่ละรายต้องลงทุนเอง ควรมี 3PL หรือ Shared Platform ที่ให้ใช้ cold chain infrastructure แบบ subscription/pay-per-use ลด barrier to entry"],
    [8, "การขนส่งข้ามประเทศใน ASEAN", "Ashwin Gurjar, Dr. Uttam Krishna",
     "แต่ละประเทศมีข้อกำหนดแตกต่างกัน ทำให้การขนส่งข้ามพรมแดนยุ่งยาก",
     "เช่น ขนส่งไทย→ลาว ต้องผ่านหลายจุด, ผู้รับสินค้ามักขอกด data logger ก่อน — ทำให้ document ยุ่งยาก DKSH มีระบบดีเพราะมีสาขาในหลายประเทศ"],
    [9, "อนาคต Cold Chain ในไทย", "ทุกคน",
     "ต้องใช้ Technology & Collaboration เพื่อลดต้นทุนและเพิ่มประสิทธิภาพ",
     "IoT sensors, Cloud monitoring, Smart tags, Shared infrastructure เป็นแนวโน้มสำคัญ ไทยมีศักยภาพแต่ต้องพัฒนา framework และโครงสร้างพื้นฐานร่วมกันระหว่างภาครัฐและเอกชน"],
    [10, "สรุป: GDP Compliance ต้องใช้ทั้งคนและเทคโนโลยี", "Ivan Sham",
     "GDP ไม่ใช่ยาวิเศษ แต่เป็นเครื่องมือช่วยพัฒนาธุรกิจ",
     "GDP เป็นตัวช่วยให้ธุรกิจพัฒนาขึ้น ไม่ใช่ภาระ การนำเทคโนโลยีมาใช้ + การออกแบบระบบที่ดี + การฝึกอบรมพนักงาน จะช่วยให้ cold chain มีประสิทธิภาพ"],
]

ws1 = wb.active
ws1.title = "Session 1 (11-12)"
write_sheet(ws1,
    "Session 1: มาตรฐาน GDP, Serialization และ Temperature Monitoring",
    "11:00-12:00 น.",
    "",
    s1_text, s1_seg, s1_topics,
    "Moderator: Mr. Ivan Sham (Red Wheels Trading)\nPanellists: Mr. Ashwin Gurjar (Win Life Pharma) | Mr. Mark Scheepers (Thai Thermo Dynamics) | Dr. Uttam Krishna (Pharmalytics MQ Bio Tech)"
)

ws2 = wb.create_sheet("Session 2 (13-15~14-14)")
write_sheet(ws2,
    "Session 2: การบริหารความเสี่ยง Cold Chain และอนาคต Logistics ยา",
    "13:15-14:14 น.",
    "",
    s2_text, s2_seg, s2_topics,
    "Moderator: Mr. Ivan Sham (Red Wheels Trading)\nPanellists: Mr. Ashwin Gurjar (Win Life Pharma) | Mr. Mark Scheepers (Thai Thermo Dynamics) | Dr. Uttam Krishna (Pharmalytics MQ Bio Tech)"
)

# Sheet 3: Timestamps S1
ws3 = wb.create_sheet("Timestamps Session 1")
for c, h in enumerate(["เวลา", "เนื้อหา (Timestamped)"], 1):
    cell = ws3.cell(row=1, column=c, value=h)
    cell.font = hf; cell.fill = hfill; cell.border = tb
ws3.column_dimensions['A'].width = 28
ws3.column_dimensions['B'].width = 120
for r, line in enumerate(s1_seg, 2):
    line = line.strip()
    if not line: continue
    parts = line.split("] ", 1)
    if len(parts) == 2:
        ws3.cell(row=r, column=1, value=parts[0].lstrip("["))
        ws3.cell(row=r, column=2, value=parts[1])
    for c in [1,2]:
        ws3.cell(row=r, column=c).font = Font(size=9)
        ws3.cell(row=r, column=c).border = tb
        ws3.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical='top')

# Sheet 4: Timestamps S2
ws4 = wb.create_sheet("Timestamps Session 2")
for c, h in enumerate(["เวลา", "เนื้อหา (Timestamped)"], 1):
    cell = ws4.cell(row=1, column=c, value=h)
    cell.font = hf; cell.fill = hfill; cell.border = tb
ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 120
for r, line in enumerate(s2_seg, 2):
    line = line.strip()
    if not line: continue
    parts = line.split("] ", 1)
    if len(parts) == 2:
        ws4.cell(row=r, column=1, value=parts[0].lstrip("["))
        ws4.cell(row=r, column=2, value=parts[1])
    for c in [1,2]:
        ws4.cell(row=r, column=c).font = Font(size=9)
        ws4.cell(row=r, column=c).border = tb
        ws4.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical='top')

xlsx_path = os.path.join(out_dir, "Cold_Chain_Summary_v2_TH.xlsx")
wb.save(xlsx_path)
print("Saved:", xlsx_path)
