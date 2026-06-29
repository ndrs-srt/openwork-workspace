import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook(r"E:\01.NDRS_WORKON\OpenWork\PEA_INVERTER_LIST\inverter_list.xlsx")
ws = wb["20-22kW"]

# Check if Price/Price Link columns exist
current_headers = [c.value for c in ws[1]]
print(f"Current headers: {current_headers}")

# Add Price column (col 12) and Price Link (col 13) if they don't exist
if "Price" not in current_headers:
    ws.cell(1, 12).value = "Price"
if "Price Link" not in current_headers:
    ws.cell(1, 13).value = "Price Link"

# Pricing data (found from research)
price_data = {
    1:  {"brand": "Afore", "model": "BNT020KTL", "price": "€1,599 + VAT (~฿63,000)", "link": "https://reestartsolar.com/en/product/afore-20kw-three-phase-string-inverter-bnt020ktl/"},
    2:  {"brand": "Alpha ESS", "model": "SMILE-G3-T20-INV", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://www.alpha-ess.co.uk/"},
    3:  {"brand": "ATESS", "model": "HPS30", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://www.atesspower.com/contact"},
    4:  {"brand": "AUXSOL", "model": "ASG-20TL-ZH", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://www.auxsol.com/contact"},
    5:  {"brand": "AUXSOL", "model": "ASN-20TL", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://www.auxsol.com/contact"},
    6:  {"brand": "AUXSOL", "model": "ASN-30TL-G2", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://www.auxsol.com/contact"},
    7:  {"brand": "CHINT POWER", "model": "SCA20K-T-EU", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://shop.faveos.com/String-Wechselrichter-Dreiphasen-mit-20-0-kW-von-CHINT/100040"},
    8:  {"brand": "CLEANLINE", "model": "CL 30KTL-P3", "price": "ติดต่อตัวแทนจำหน่าย", "link": ""},
    9:  {"brand": "FoxESS", "model": "P3-Pro-20.0", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://www.energrosso.com/en/collections/fox-ess-inverterek"},
    10: {"brand": "FoxESS", "model": "P3-Pro-30.0", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://www.energrosso.com/en/collections/fox-ess-inverterek"},
    11: {"brand": "FoxESS", "model": "T20-G3", "price": "ติดต่อตัวแทนจำหน่าย", "link": ""},
    12: {"brand": "FoxESS", "model": "T30-M", "price": "ติดต่อตัวแทนจำหน่าย", "link": ""},
    13: {"brand": "GOODWE", "model": "GW30K-SDT-C30", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://en.goodwe.com/Product/index/id/15"},
    14: {"brand": "GROWATT", "model": "MID20KTL3-X", "price": "ติดต่อตัวแทนจำหน่าย", "link": ""},
    15: {"brand": "HUAWEI", "model": "SUN2000-20KTL-M2", "price": "~฿62,700 - ฿92,700", "link": "https://solartouch.co.th/product/huawei-20k-3-phase-sun2000-20ktl-mb0/"},
    16: {"brand": "HYXiPOWER", "model": "HYX-H25K-HT", "price": "ติดต่อตัวแทนจำหน่าย", "link": ""},
    17: {"brand": "ISUNA", "model": "Isuna 20000T", "price": "฿102,132", "link": "https://www.r3solarcell.com/product/187/isuna-20kw-3-phase-hv-hybrid-inverter"},
    18: {"brand": "LANPWR", "model": "LANPWR 20000T", "price": "ติดต่อตัวแทนจำหน่าย", "link": ""},
    19: {"brand": "LESSO", "model": "LSBH20KTL3-OC1", "price": "ติดต่อตัวแทนจำหน่าย", "link": ""},
    20: {"brand": "LIVOLTEK", "model": "HP3-30KT2", "price": "ติดต่อตัวแทนจำหน่าย", "link": ""},
    21: {"brand": "PSI", "model": "P250", "price": "ติดต่อตัวแทนจำหน่าย", "link": ""},
    22: {"brand": "SAJ", "model": "H2-20K-LT2", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://www.saj-electric.com/product/19.html"},
    23: {"brand": "Sigenergy", "model": "SigenStor EC 20.0 TP", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://www.sigenergy.com/products/sigenstor"},
    24: {"brand": "SINENG", "model": "SN20PT", "price": "ติดต่อตัวแทนจำหน่าย", "link": ""},
    25: {"brand": "SINENG", "model": "SN30PT", "price": "ติดต่อตัวแทนจำหน่าย", "link": ""},
    26: {"brand": "SOFAR SOLAR", "model": "SOFAR 30KTLX-G3", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://www.sofar-solar.com/product/sofar-30ktlx-g3/"},
    27: {"brand": "SOLAR ENERGY", "model": "SUNLIGHT-20KTH", "price": "ติดต่อตัวแทนจำหน่าย (~80,000-100,000 บาท)", "link": "https://b2hsolar.com/%e0%b8%ad%e0%b8%b4%e0%b8%99%e0%b9%80%e0%b8%a7%e0%b8%ad%e0%b8%a3%e0%b9%8c%e0%b9%80%e0%b8%95%e0%b8%ad%e0%b8%a3%e0%b9%8c/"},
    28: {"brand": "SOLAX POWER", "model": "X3-MGA-40K-G2", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://www.solaxpower.com/product/x3-mga-g2/"},
    29: {"brand": "SOLAX POWER", "model": "X3-PRO-20K-G2", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://www.solaxpower.com/product/x3-pro-g2/"},
    30: {"brand": "SOLAX POWER", "model": "X3-PRO-30K-G2", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://www.solaxpower.com/product/x3-pro-g2/"},
    31: {"brand": "SOLAX POWER", "model": "X3-ULT-20K", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://www.solaxpower.com/product/x3-ultimate/"},
    32: {"brand": "SOLAX POWER", "model": "X3-ULT-30K", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://www.solaxpower.com/product/x3-ultimate/"},
    33: {"brand": "solis", "model": "S6-EH3P20K-H", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://www.ginlong.com/product/s6-eh3p20k-h.html"},
    34: {"brand": "Sungrow", "model": "SG20RT-P2", "price": "ติดต่อตัวแทนจำหน่าย", "link": "https://en.sungrowpower.com/productDetail/593"},
    35: {"brand": "SUNGROW", "model": "SH25T", "price": "฿140,000-170,000 (~€3,355)", "link": "https://panpower.eu/en/inverters/4714-sungrow-sh25t.html"},
    36: {"brand": "TMDA", "model": "SUN-20K-G03", "price": "ติดต่อตัวแทนจำหน่าย", "link": ""},
    37: {"brand": "TMDA", "model": "SUN-30K-G03", "price": "ติดต่อตัวแทนจำหน่าย", "link": ""},
    38: {"brand": "TMDA", "model": "SUN-40K-G04", "price": "ติดต่อตัวแทนจำหน่าย", "link": ""},
}

# Update specific links for the 3 requested models
specific_updates = {
    17: {  # Isuna 20000T
        "link": "https://en.sinexcel.com/HomeEnergy/Three-phase-hybrid-inverter.php",
        "where_buy": "https://www.r3solarcell.com/product/187/isuna-20kw-3-phase-hv-hybrid-inverter (ไทย)",
    },
    27: {  # SUNLIGHT-20KTH
        "link": "https://en.eastups.com/ResidentialEnergyStorage/8281.jhtml",
        "where_buy": "https://b2hsolar.com (B2H Solar ไทย)",
    },
    35: {  # SH25T
        "link": "https://aus.sungrowpower.com/productDetail/3158/mv-power-converter-hybrid-inverter-sh15-20-25t",
        "where_buy": "https://panpower.eu/en/inverters/4714-sungrow-sh25t.html / ติดต่อตัวแทน Sungrow ไทย",
    },
}

for r in range(2, ws.max_row + 1):
    cell_no = ws.cell(r, 1).value
    if cell_no in price_data:
        pd = price_data[cell_no]
        ws.cell(r, 12).value = pd["price"]
        ws.cell(r, 13).value = pd["link"]
    
    # Update link/where_buy for the 3 specific models
    if cell_no in specific_updates:
        su = specific_updates[cell_no]
        if "link" in su:
            ws.cell(r, 9).value = su["link"]
        if "where_buy" in su:
            ws.cell(r, 11).value = su["where_buy"]

wb.save(r"E:\01.NDRS_WORKON\OpenWork\PEA_INVERTER_LIST\inverter_list.xlsx")
print("Excel updated with Price and Price Link columns!")
print("\n=== สรุป 3 รุ่นที่อัปเดตลิงก์ ===")
print("17. Isuna 20000T: https://en.sinexcel.com/HomeEnergy/Three-phase-hybrid-inverter.php")
print("    ซื้อได้ที่: https://www.r3solarcell.com/product/187/isuna-20kw-3-phase-hv-hybrid-inverter")
print("27. SUNLIGHT-20KTH: https://en.eastups.com/ResidentialEnergyStorage/8281.jhtml")
print("    ซื้อได้ที่: https://b2hsolar.com (B2H Solar)")
print("35. SH25T: https://aus.sungrowpower.com/productDetail/3158")
print("    ซื้อได้ที่: https://panpower.eu/en/inverters/4714-sungrow-sh25t.html / ตัวแทน Sungrow ไทย")
