import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()

# ── Styles ──
DARK_BLUE = "1F3864"
MED_BLUE  = "2E75B6"
LIGHT_BLUE = "D6E4F0"
WHITE     = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
YELLOW_BG = "FFF2CC"
GREEN_BG  = "E2EFDA"
RED_BG    = "FCE4EC"

hdr_font   = Font(name="Calibri", bold=True, color=WHITE, size=11)
title_font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=14)
sub_font   = Font(name="Calibri", bold=True, color=DARK_BLUE, size=11)
norm_font  = Font(name="Calibri", size=10)
bold_font  = Font(name="Calibri", bold=True, size=10)
small_font = Font(name="Calibri", size=9, italic=True, color="555555")

hdr_fill   = PatternFill("solid", fgColor=MED_BLUE)
light_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
yellow_fill = PatternFill("solid", fgColor=YELLOW_BG)

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))
wrap_align  = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align  = Alignment(vertical="top", wrap_text=True)


def setup_sheet(ws, title_text, col_count=6):
    ws.sheet_properties.tabColor = MED_BLUE
    c = ws.cell(row=1, column=1, value=title_text)
    c.font = title_font
    c.alignment = Alignment(horizontal="center", vertical="center")


def hdr_row(ws, row, headers):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center_align
        cell.border = thin_border


def data_cell(ws, row, col, value="", font=None, align=None, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or norm_font
    cell.alignment = align or wrap_align
    cell.border = thin_border
    if fill:
        cell.fill = fill
    return cell


def write_field(ws, row, col, label, value="", label_width=30):
    data_cell(ws, row, col, label, bold_font, left_align)
    data_cell(ws, row, col + 1, value, norm_font, left_align)


def write_empty_table(ws, start_row, headers, row_count, col_widths=None):
    """Write header row + empty data rows."""
    ncols = len(headers)
    hdr_row(ws, start_row, headers)
    for r in range(start_row + 1, start_row + 1 + row_count):
        for c in range(1, ncols + 1):
            data_cell(ws, r, c, "", norm_font, left_align)
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return start_row + 1 + row_count + 1


# ──── Remove default sheet ────
default = wb.active
wb.remove(default)

# =====================================================================
# FORM 1 — FR-PR-001 ใบขอซื้อ (Purchase Requisition)
# =====================================================================
ws1 = wb.create_sheet("FR-PR-001 ใบขอซื้อ")
setup_sheet(ws1, "FR-PR-001 — ใบขอซื้อ (Purchase Requisition)")

info1 = [
    ("PR No.:", "PR-_____-___"),
    ("วันที่ขอซื้อ:", "____/____/______"),
    ("ฝ่ายที่ขอ:", "__________________"),
    ("ผู้ขอซื้อ:", "__________________"),
]
for i, (l, v) in enumerate(info1, 3):
    write_field(ws1, i, 1, l, v)
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 30

write_empty_table(ws1, 8,
    ["ลำดับ", "รหัสสินค้า", "รายละเอียดสินค้า / สเปก", "จำนวน", "หน่วย", "วันที่ต้องการ"],
    5, [6, 16, 40, 12, 10, 16])

r = 8 + 1 + 5 + 1  # 15
data_cell(ws1, r, 1, "เหตุผลในการขอซื้อ:", bold_font, left_align)
data_cell(ws1, r, 2, "☐ ใช้งานทั่วไป  ☐ เปลี่ยนของเสีย  ☐ โครงการใหม่  ☐ สำรองStock  ☐ อื่นๆ ......", norm_font, left_align)

data_cell(ws1, r+1, 1, "อนุมัติโดย:", bold_font, left_align)
data_cell(ws1, r+1, 2, "_______________  วันที่ ____/____/______", norm_font, left_align)
data_cell(ws1, r+1, 4, "อนุมัติงบประมาณโดย:", bold_font, left_align)
data_cell(ws1, r+1, 5, "_______________  วันที่ ____/____/______", norm_font, left_align)

# =====================================================================
# FORM 2 — FR-PUR-001 ใบสั่งซื้อ (Purchase Order)
# =====================================================================
ws2 = wb.create_sheet("FR-PUR-001 ใบสั่งซื้อ")
setup_sheet(ws2, "FR-PUR-001 — ใบสั่งซื้อ (Purchase Order)")

po_info = [
    ("PO No.:", "PO-_____-___", "วันที่:", "____/____/______"),
    ("ชื่อผู้ขาย:", "_________________________", "Tax ID:", "_________________"),
    ("ที่อยู่:", "_________________________", "ผู้ติดต่อ:", "_________________"),
    ("เงื่อนไขการชำระเงิน:", "☐ เงินสด  ☐ 30 วัน  ☐ 60 วัน  ☐ อื่นๆ......", "Incoterm:", "☐ EXW ☐ FOB ☐ CIF ☐ DDP"),
    ("วันที่ต้องการ:", "____/____/______", "สถานที่จัดส่ง:", "_________________"),
]
for i, (l1, v1, l2, v2) in enumerate(po_info, 3):
    write_field(ws2, i, 1, l1, v1)
    write_field(ws2, i, 3, l2, v2)
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 22
ws2.column_dimensions["D"].width = 30

write_empty_table(ws2, 9,
    ["ลำดับ", "รหัสสินค้า", "รายละเอียดสินค้า / Specification", "จำนวน", "หน่วย", "ราคาต่อหน่วย"],
    5, [6, 16, 40, 12, 10, 16])

r = 9 + 1 + 5  # 15
sum_row = r + 1
data_cell(ws2, sum_row, 4, "รวมเงิน (ก่อน VAT):", bold_font, Alignment(horizontal="right", vertical="top"))
data_cell(ws2, sum_row, 5, "", norm_font, left_align)
data_cell(ws2, sum_row, 6, "", norm_font, left_align)
data_cell(ws2, sum_row+1, 4, "VAT ___%:", bold_font, Alignment(horizontal="right", vertical="top"))
data_cell(ws2, sum_row+1, 5, "", norm_font, left_align)
data_cell(ws2, sum_row+1, 6, "", norm_font, left_align)
data_cell(ws2, sum_row+2, 4, "รวมทั้งสิ้น:", bold_font, Alignment(horizontal="right", vertical="top"))
data_cell(ws2, sum_row+2, 5, "", norm_font, left_align)
data_cell(ws2, sum_row+2, 6, "", norm_font, left_align)

r2 = sum_row + 3
data_cell(ws2, r2, 1, "ผู้อนุมัติ PO:", bold_font)
data_cell(ws2, r2, 2, "_______________  วันที่ ____/____/______", norm_font)
data_cell(ws2, r2, 4, "ผู้จัดซื้อ:", bold_font)
data_cell(ws2, r2, 5, "_______________  วันที่ ____/____/______", norm_font)

note2 = ws2.cell(row=r2+1, column=1,
    value="หมายเหตุ: ใบสั่งซื้อนี้มีผลผูกพันทางกฎหมายเมื่อมีลายเซ็นผู้อนุมัติ  กรุณาส่ง PO ยืนยันกลับภายใน 3 วันทำการ")
note2.font = small_font
note2.alignment = wrap_align

# =====================================================================
# FORM 3 — FR-PUR-002 แบบประเมินผู้ขาย (Supplier Evaluation)
# =====================================================================
ws3 = wb.create_sheet("FR-PUR-002 ประเมินผู้ขาย")
setup_sheet(ws3, "FR-PUR-002 — แบบประเมินผู้ขาย (Supplier Evaluation Form)", 6)

info3 = [
    ("Supplier Name:", "_________________________"),
    ("Address:", "_________________________"),
    ("Contact Person:", "_________  Tel: _________  Email: _________"),
    ("Product / Service:", "_________________________"),
    ("Type:", "☐ New Supplier  ☐ Re-evaluation  ☐ Annual Review"),
    ("Evaluation Date:", "____/____/______  |  Evaluator: _________________"),
]
for i, (l, v) in enumerate(info3, 3):
    write_field(ws3, i, 1, l, v)
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 55
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 20
ws3.column_dimensions["F"].width = 20

# Scoring table
ev_hdrs = ["เกณฑ์ (Criteria)", "น้ำหนัก", "คะแนน\n(1-10)", "คะแนนถ่วงน้ำหนัก", "หลักฐาน / หมายเหตุ"]
ev_data = [
    ["1. คุณภาพ (Quality)", "35%", "", "", ""],
    ["   - ระบบคุณภาพ / ISO Certificate", "", "", "", ""],
    ["   - ประวัติ Defect / PPM", "", "", "", ""],
    ["   - การควบคุมกระบวนการ", "", "", "", ""],
    ["2. ราคา (Price)", "20%", "", "", ""],
    ["   - ระดับราคาเทียบตลาด", "", "", "", ""],
    ["   - โครงสร้างราคาโปร่งใส", "", "", "", ""],
    ["3. การส่งมอบ (Delivery)", "20%", "", "", ""],
    ["   - On-Time Delivery (OTD) %", "", "", "", ""],
    ["   - Lead Time / ความยืดหยุ่น", "", "", "", ""],
    ["4. บริการ (Service)", "10%", "", "", ""],
    ["5. ความน่าเชื่อถือ (Reliability)", "10%", "", "", ""],
    ["   - ฐานะการเงิน / ประวัติบริษัท", "", "", "", ""],
    ["   - การอ้างอิงจากลูกค้า", "", "", "", ""],
    ["6. EHS (สิ่งแวดล้อม/ความปลอดภัย)", "5%", "", "", ""],
]
start3 = 10
for i, h in enumerate(ev_hdrs, 1):
    c = ws3.cell(row=start3, column=i, value=h)
hdr_row(ws3, start3, ev_hdrs)

for ri, row_data in enumerate(ev_data, start3 + 1):
    for ci, val in enumerate(row_data, 1):
        fill = light_fill if "คะแนน" in str(val) or val == "" else None
        data_cell(ws3, ri, ci, val, bold_font if val and not val.startswith("   -") else norm_font, left_align, fill)

# Set column widths for scoring table
for i, w in enumerate([35, 12, 14, 20, 35], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

tot_row = start3 + 1 + len(ev_data)  # row after last data row
data_cell(ws3, tot_row, 1, "คะแนนรวม (Total Score):", bold_font, left_align)
data_cell(ws3, tot_row, 2, "100%", bold_font, center_align)
data_cell(ws3, tot_row, 3, "", norm_font, left_align)
data_cell(ws3, tot_row, 4, "", bold_font, center_align, yellow_fill)

res_row = tot_row + 1
data_cell(ws3, res_row, 1, "ผลการประเมิน:", bold_font)
data_cell(ws3, res_row, 2, "☐ A (≥85%) ดีเยี่ยม  ☐ B (70-84%) ดี  ☐ C (50-69%) พอใช้  ☐ D (<50%) ต้องปรับปรุง", norm_font)

data_cell(ws3, res_row+1, 1, "ข้อเสนอแนะ:", bold_font)
data_cell(ws3, res_row+1, 2, "_______________________________________________", norm_font)

data_cell(ws3, res_row+2, 1, "ผู้ประเมิน:", bold_font)
data_cell(ws3, res_row+2, 2, "_______________  วันที่ ____/____/______", norm_font)
data_cell(ws3, res_row+2, 4, "ผู้อนุมัติ:", bold_font)
data_cell(ws3, res_row+2, 5, "_______________  วันที่ ____/____/______", norm_font)

ws3.column_dimensions["F"].width = 20

# =====================================================================
# FORM 4 — FR-PUR-003 ทะเบียนผู้ขายอนุมัติ (ASL)
# =====================================================================
ws4 = wb.create_sheet("FR-PUR-003 ASL")
setup_sheet(ws4, "FR-PUR-003 — ทะเบียนผู้ขายอนุมัติ (Approved Supplier List — ASL)", 9)

info4 = [
    ("Revision No.:", "____",  "Date:", date.today().strftime("%d/%m/%Y")),
    ("Prepared by:", "_________________", "Approved by:", "_________________"),
]
for i, (l1, v1, l2, v2) in enumerate(info4, 3):
    write_field(ws4, i, 1, l1, v1)
    write_field(ws4, i, 3, l2, v2)

ws4.column_dimensions["A"].width = 18
ws4.column_dimensions["B"].width = 25
ws4.column_dimensions["C"].width = 16
ws4.column_dimensions["D"].width = 25
ws4.column_dimensions["E"].width = 22
ws4.column_dimensions["F"].width = 18
ws4.column_dimensions["G"].width = 16
ws4.column_dimensions["H"].width = 16
ws4.column_dimensions["I"].width = 16

write_empty_table(ws4, 6,
    ["ลำดับ", "ชื่อผู้ขาย", "ผลิตภัณฑ์/บริการ", "วันที่ประเมิน", "ระดับ\n(A/B/C/D)", "วันหมดอายุ", "สถานะ\n(Active/Suspend)", "หมายเหตุ"],
    10, [6, 25, 22, 16, 12, 16, 16, 16, 20])

# =====================================================================
# FORM 5 — FR-QC-001 รายงานตรวจสอบคุณภาพ (Inspection Report)
# =====================================================================
ws5 = wb.create_sheet("FR-QC-001 ตรวจสอบ QC")
setup_sheet(ws5, "FR-QC-001 — รายงานตรวจสอบคุณภาพ (Inspection Report)", 8)

info5 = [
    ("Report No.:", "QC-_____-___", "Date:", date.today().strftime("%d/%m/%Y")),
    ("PO No.:", "_________________", "Supplier:", "_________________"),
    ("Item / Material:", "_________________", "Lot / Batch No.:", "_________________"),
    ("Quantity Received:", "_________", "Unit:", "_________"),
]
for i, (l1, v1, l2, v2) in enumerate(info5, 3):
    write_field(ws5, i, 1, l1, v1)
    write_field(ws5, i, 4, l2, v2)

ws5.column_dimensions["A"].width = 20
ws5.column_dimensions["B"].width = 25
ws5.column_dimensions["C"].width = 5
ws5.column_dimensions["D"].width = 18
ws5.column_dimensions["E"].width = 25
ws5.column_dimensions["F"].width = 12
ws5.column_dimensions["G"].width = 12
ws5.column_dimensions["H"].width = 12

insp_hdrs = ["Check Point", "Spec / Standard", " Method", "Result", "Accept?\n(Pass/Fail)", "Inspector", "Remark"]
write_empty_table(ws5, 8, insp_hdrs, 6, [30, 25, 15, 15, 12, 15, 20])

r5 = 8 + 1 + 6 + 1  # 16
data_cell(ws5, r5, 1, "Sampling Plan:", bold_font)
data_cell(ws5, r5, 2, "☐ 100% Inspection  ☐ AQL ___%  ☐ MIL-STD-1916  ☐ Other: _______", norm_font)

data_cell(ws5, r5+1, 1, "Total Inspected:", bold_font)
data_cell(ws5, r5+1, 2, "_________  pcs", norm_font)
data_cell(ws5, r5+1, 4, "Defect Found:", bold_font)
data_cell(ws5, r5+1, 5, "_________  pcs", norm_font)

data_cell(ws5, r5+2, 1, "ผลการตรวจสอบ:", bold_font)
data_cell(ws5, r5+2, 2, "☐ Accepted  ☐ Rejected  ☐ Conditional Accepted (ระบุ) _________________", norm_font)

data_cell(ws5, r5+3, 1, "QC Inspector:", bold_font)
data_cell(ws5, r5+3, 2, "_______________  วันที่ ____/____/______", norm_font)
data_cell(ws5, r5+3, 5, "QC Supervisor:", bold_font)
data_cell(ws5, r5+3, 6, "_______________  วันที่ ____/____/______", norm_font)

# =====================================================================
# FORM 6 — FR-QC-002 รายการสินค้าไม่เป็นไปตามข้อกำหนด (NCR)
# =====================================================================
ws6 = wb.create_sheet("FR-QC-002 NCR")
setup_sheet(ws6, "FR-QC-002 — รายการสินค้าไม่เป็นไปตามข้อกำหนด (Nonconformance Report — NCR)", 8)

info6 = [
    ("NCR No.:", "NCR-_____-___", "Date:", date.today().strftime("%d/%m/%Y")),
    ("PO No.:", "_________________", "Supplier:", "_________________"),
    ("Item:", "_________________", "Quantity Rejected:", "_________"),
]
for i, (l1, v1, l2, v2) in enumerate(info6, 3):
    write_field(ws6, i, 1, l1, v1)
    write_field(ws6, i, 4, l2, v2)

data_cell(ws6, 6, 1, "รายละเอียดข้อบกพร่อง (Description of Nonconformance):", bold_font)
data_cell(ws6, 6, 4, "__________________________________________________________________", norm_font)
# merge cols 4-8 for description
for c in range(5, 9):
    data_cell(ws6, 6, c, "", norm_font)

data_cell(ws6, 7, 1, "ตรวจพบโดย (Detected by):", bold_font)
data_cell(ws6, 7, 2, "☐ QC Incoming  ☐ Production  ☐ Warehouse  ☐ Customer  ☐ Other: _____", norm_font)

data_cell(ws6, 8, 1, "ประเภทข้อบกพร่อง (Defect Category):", bold_font)
data_cell(ws6, 8, 2, "☐ Critical  ☐ Major  ☐ Minor", norm_font)

data_cell(ws6, 9, 1, "การดำเนินการกับสินค้านี้ (Disposition):", bold_font)
data_cell(ws6, 9, 2, "☐ Return to Supplier  ☐ Rework  ☐ Use As Is  ☐ Scrap  ☐ Other: _____", norm_font)

data_cell(ws6, 10, 1, "Root Cause (จากผู้ขาย):", bold_font)
data_cell(ws6, 10, 2, "_______________________________________________", norm_font)

data_cell(ws6, 11, 1, "Corrective Action (CAR No.):", bold_font)
data_cell(ws6, 11, 2, "CAR-_____-___", norm_font)

data_cell(ws6, 12, 1, "QC Inspector:", bold_font)
data_cell(ws6, 12, 2, "_______________  วันที่ ____/____/______", norm_font)
data_cell(ws6, 12, 5, "Supplier Acknowledged:", bold_font)
data_cell(ws6, 12, 6, "_______________  วันที่ ____/____/______", norm_font)

ws6.column_dimensions["A"].width = 25
ws6.column_dimensions["B"].width = 30
ws6.column_dimensions["C"].width = 5
ws6.column_dimensions["D"].width = 20
ws6.column_dimensions["E"].width = 25
ws6.column_dimensions["F"].width = 15
ws6.column_dimensions["G"].width = 15
ws6.column_dimensions["H"].width = 15

# =====================================================================
# FORM 7 — FR-PUR-004 Supplier Scorecard
# =====================================================================
ws7 = wb.create_sheet("FR-PUR-004 Scorecard")
setup_sheet(ws7, "FR-PUR-004 — Supplier Scorecard (Performance Report)", 8)

info7 = [
    ("Scorecard Period:", "☐ Q1  ☐ Q2  ☐ Q3  ☐ Q4  Year: ______", "", ""),
    ("Supplier:", "_________________________", "Product:", "_________________"),
    ("Review Date:", date.today().strftime("%d/%m/%Y"), "Reviewed by:", "_________________"),
]
for i, (l1, v1, l2, v2) in enumerate(info7, 3):
    write_field(ws7, i, 1, l1, v1)
    write_field(ws7, i, 4, l2, v2)

ws7.column_dimensions["A"].width = 22
ws7.column_dimensions["B"].width = 25
ws7.column_dimensions["C"].width = 5
ws7.column_dimensions["D"].width = 18
ws7.column_dimensions["E"].width = 16
ws7.column_dimensions["F"].width = 16
ws7.column_dimensions["G"].width = 16
ws7.column_dimensions["H"].width = 16

sc_hdrs = ["KPI", "Weight", "Target", "Actual", "Score", "Weighted Score", "Remark"]
sc_data = [
    ["On-Time Delivery (OTD)", "30%", "≥95%", "", "", "", ""],
    ["Quality (Defect PPM)", "30%", "≤ 1000 ppm", "", "", "", ""],
    ["Price Competitiveness", "15%", "Market Level", "", "", "", ""],
    ["Service / Responsiveness", "15%", "≤ 24 hrs", "", "", "", ""],
    ["EHS / Compliance", "10%", "No Violation", "", "", "", ""],
]
start_sc = 7
for i, h in enumerate(sc_hdrs, 1):
    c = ws7.cell(row=start_sc, column=i, value=h)
hdr_row(ws7, start_sc, sc_hdrs)

for ri, row_data in enumerate(sc_data, start_sc + 1):
    for ci, val in enumerate(row_data, 1):
        data_cell(ws7, ri, ci, val, norm_font, left_align)

subtotal_r = start_sc + 1 + len(sc_data)  # 7 + 1 + 5 = 13
data_cell(ws7, subtotal_r, 1, "Total Score", bold_font, left_align)
data_cell(ws7, subtotal_r, 2, "100%", bold_font, center_align)
for c in range(3, 8):
    data_cell(ws7, subtotal_r, c, "", norm_font, left_align)
# Put formula placeholders
data_cell(ws7, subtotal_r, 6, "", bold_font, center_align, yellow_fill)
data_cell(ws7, subtotal_r, 7, "", norm_font, left_align)

data_cell(ws7, subtotal_r+1, 1, "Rating:", bold_font)
data_cell(ws7, subtotal_r+1, 2, "☐ A (≥85)  ☐ B (70-84)  ☐ C (50-69)  ☐ D (<50)", norm_font)

data_cell(ws7, subtotal_r+2, 1, "Comments / Action Plan:", bold_font)
data_cell(ws7, subtotal_r+2, 2, "_______________________________________________", norm_font)

data_cell(ws7, subtotal_r+3, 1, "Supplier Signature:", bold_font)
data_cell(ws7, subtotal_r+3, 2, "_______________  วันที่ ____/____/______", norm_font)
data_cell(ws7, subtotal_r+3, 5, "Purchasing Manager:", bold_font)
data_cell(ws7, subtotal_r+3, 6, "_______________  วันที่ ____/____/______", norm_font)


# =====================================================================
# FORM 8 — FR-PUR-005 ใบแจ้งแก้ไขผู้ขาย (SCAR)
# =====================================================================
ws8 = wb.create_sheet("FR-PUR-005 SCAR")
setup_sheet(ws8, "FR-PUR-005 — ใบแจ้งแก้ไขผู้ขาย (Supplier Corrective Action Request — SCAR)", 8)

info8 = [
    ("SCAR No.:", "SCAR-_____-___", "Date:", date.today().strftime("%d/%m/%Y")),
    ("Supplier:", "_________________________", "Contact:", "_________________"),
    ("NCR Reference:", "NCR-_____-___", "PO No.:", "_________________"),
    ("Problem Description:", "_______________________________________________", "", ""),
]
for i, (l1, v1, l2, v2) in enumerate(info8, 3):
    write_field(ws8, i, 1, l1, v1)
    if l2:
        write_field(ws8, i, 4, l2, v2)

data_cell(ws8, 7, 1, "Severity:", bold_font)
data_cell(ws8, 7, 2, "☐ Critical  ☐ Major  ☐ Minor", norm_font)
data_cell(ws8, 7, 4, "Require 8D Report?", bold_font)
data_cell(ws8, 7, 5, "☐ Yes  ☐ No", norm_font)

# 8D / Root Cause section
data_cell(ws8, 9, 1, "Root Cause (วิเคราะห์โดยผู้ขาย):", bold_font)
data_cell(ws8, 9, 2, "__________________________________________________________________", norm_font)
for c in range(3, 9):
    data_cell(ws8, 9, c, "", norm_font)

data_cell(ws8, 10, 1, "Corrective Action Plan:", bold_font)
data_cell(ws8, 10, 2, "__________________________________________________________________", norm_font)
for c in range(3, 9):
    data_cell(ws8, 10, c, "", norm_font)

data_cell(ws8, 11, 1, "Preventive Action:", bold_font)
data_cell(ws8, 11, 2, "__________________________________________________________________", norm_font)
for c in range(3, 9):
    data_cell(ws8, 11, c, "", norm_font)

data_cell(ws8, 12, 1, "Implementation Date:", bold_font)
data_cell(ws8, 12, 2, "____/____/______", norm_font)
data_cell(ws8, 12, 4, "Verification Date:", bold_font)
data_cell(ws8, 12, 5, "____/____/______", norm_font)

data_cell(ws8, 13, 1, "Status:", bold_font)
data_cell(ws8, 13, 2, "☐ Open  ☐ In Progress  ☐ Closed  ☐ Rejected", norm_font)

data_cell(ws8, 14, 1, "Verification Result:", bold_font)
data_cell(ws8, 14, 2, "☐ Effective — Closed  ☐ Partially Effective — Follow up  ☐ Not Effective — Re-open", norm_font)

data_cell(ws8, 15, 1, "Supplier Representative:", bold_font)
data_cell(ws8, 15, 2, "_______________  วันที่ ____/____/______", norm_font)
data_cell(ws8, 15, 5, "Purchasing / QA:", bold_font)
data_cell(ws8, 15, 6, "_______________  วันที่ ____/____/______", norm_font)

ws8.column_dimensions["A"].width = 25
ws8.column_dimensions["B"].width = 30
ws8.column_dimensions["C"].width = 5
ws8.column_dimensions["D"].width = 20
ws8.column_dimensions["E"].width = 22
ws8.column_dimensions["F"].width = 15
ws8.column_dimensions["G"].width = 15
ws8.column_dimensions["H"].width = 15

# ═══════════════════════════════════════════════════════════════
# Apply print settings
# ═══════════════════════════════════════════════════════════════
for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.zoomScale = 85

# ═══════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════
output = "C:\\Working_Space\\OpenWork\\แบบฟอร์ม_Forms_ISO_Purchasing.xlsx"
wb.save(output)
print("OK")
