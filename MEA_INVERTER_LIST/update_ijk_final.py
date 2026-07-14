"""
Update columns I (Web Link), J (Where to Buy), K (Price) for all models.
Keep good data, fix issues found during research.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

EXCEL_FILE = r"MEA_Inverter_list_pass_standard_20-50kW.xlsx"

wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active

changes = []

def update_cell(row, col, new_value, reason):
    old_value = ws.cell(row=row, column=col).value
    if str(old_value).strip() != str(new_value).strip():
        ws.cell(row=row, column=col).value = new_value
        changes.append(f"Row {row} Col {col}: {reason}")
        return True
    return False

# ============================================================
# CORRECTIONS FOR COLUMNS I, J, K
# ============================================================

# Row 32: CHUPHOTIC STT-20KTL-P
# This is actually a Sunways OEM product - link should go to Sunways datasheet
update_cell(32, 9, 
    "https://www.sunways-tech.com/sunways-website/products/grid/SWS%20Datasheet%20STT%204~25K-P_EN.pdf",
    "CHUPHOTIC STT-20KTL-P: Updated to Sunways official datasheet (OEM manufacturer)")

# Row 76-77: Haier HS3P series
# ENF Solar has the complete datasheet
update_cell(76, 9,
    "https://www.enfsolar.com/pv/inverter-datasheet/20436",
    "Haier HS3P-20KA: Updated to ENF Solar datasheet with full specs")
update_cell(77, 9,
    "https://www.enfsolar.com/pv/inverter-datasheet/20436",
    "Haier HS3P-50KA: Updated to ENF Solar datasheet with full specs")

# Row 33: CLEANLINE CL 30KTL-P3
# Thai Watsadu link is actually correct for this Thai brand
# Keep current link as it's the official retail channel

# Row 49: Deye SUN-40K-G04
# Current link is generic deye.com - update to specific product page
update_cell(49, 9,
    "https://www.deyeinverter.com/product/three-phase-string-inverter/sun40-50ktl-g04/",
    "Deye SUN-40K-G04: Updated to specific product page")

# Row 51: Deye SUN-50K-G04
update_cell(51, 9,
    "https://www.deyeinverter.com/product/three-phase-string-inverter/sun40-50ktl-g04/",
    "Deye SUN-50K-G04: Updated to specific product page")

# Row 72: Growatt 20000UE
# Current link is generic products page - update to specific
update_cell(72, 9,
    "https://en.growatt.com/products/ue-series",
    "Growatt 20000UE: Updated to specific UE series page")

# Row 99-101: INVT iMars series
# Update to specific product page
update_cell(99, 9,
    "https://www.invt.com/products/solar-inverters/imars-bg-series",
    "INVT iMars BG20KTR: Updated to specific BG series page")
update_cell(100, 9,
    "https://www.invt.com/products/solar-inverters/imars-bg-series",
    "INVT iMars BG30KTR: Updated to specific BG series page")
update_cell(101, 9,
    "https://www.invt.com/products/solar-inverters/imars-bg-series",
    "INVT iMars BG50KTR: Updated to specific BG series page")

# Row 162: SolarEdge SE27.6K
# Current link goes to single-phase - should be three-phase
update_cell(162, 9,
    "https://www.solaredge.com/products/inverters/three-phase-inverters",
    "SolarEdge SE27.6K: Updated to three-phase inverters page")

# Row 192: Thai Tabuchi
# No public link - add MEA/PEA smart list reference
update_cell(192, 9,
    "N/A (Thai Tabuchi; Factory Bang Pakong, Chachoengsao; MEA/PEA approved; B2B only)",
    "Thai Tabuchi: No public product page available")

# Save
wb.save(EXCEL_FILE)

print("=" * 60)
print("SUMMARY OF I J K UPDATES")
print("=" * 60)
print(f"Total changes: {len(changes)}")
print()
for c in changes:
    print(f"  ✓ {c}")

# Verify final state
print("\n" + "=" * 60)
print("VERIFICATION")
print("=" * 60)
for r in [32, 76, 77, 49, 51, 72, 99, 162]:
    brand = ws.cell(row=r, column=2).value
    model = ws.cell(row=r, column=3).value
    link = ws.cell(row=r, column=9).value
    print(f"Row {r}: {brand} {model}")
    print(f"  Link: {link[:80]}...")
    print()
