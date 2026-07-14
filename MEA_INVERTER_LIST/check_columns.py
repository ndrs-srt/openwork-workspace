import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
wb = openpyxl.load_workbook('MEA_Inverter_list_pass_standard_20-50kW.xlsx')
ws = wb.active

print("Checking columns I (Web Link), J (Where to Buy), K (Price) for all rows...")
print("=" * 150)

for r in range(2, ws.max_row + 1):
    brand = str(ws.cell(row=r, column=2).value or '')
    model = str(ws.cell(row=r, column=3).value or '')
    link = str(ws.cell(row=r, column=9).value or '')
    buy = str(ws.cell(row=r, column=10).value or '')
    price = str(ws.cell(row=r, column=11).value or '')
    
    # Print all rows with their data
    print(f"Row {r}: {brand} | {model}")
    print(f"  Web Link: {link[:100]}")
    print(f"  Where to Buy: {buy[:80]}")
    print(f"  Price: {price}")
    print()
