import sys; sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

path = r'E:\01.NDRS_WORKON\OpenWork\MEA_INVERTER_LIST\MEA_Inverter_list_pass_standard_20-50kW.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active

# Where to Buy data - Thai sellers first, then international
# Format: model -> "Seller Name, URL"
where_to_buy = {
    # ABB/FIMER - Thai: Solar-Thailand, Solaris; International: SolarTraders
    "PVS-20-TL-SX": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/, SolarTraders https://solartraders.com/",
    "PVS-20-TL-SXD": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/, SolarTraders https://solartraders.com/",
    "PVS-20-TL-SY": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/, SolarTraders https://solartraders.com/",
    "TRIO-20.0-TL-OUTD-400": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/",
    "TRIO-20.0-TL-OUTD-S2-400": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/",
    "TRIO-20.0-TL-OUTD-S2F-400": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/",
    "TRIO-20.0-TL-OUTD-S2X-400": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/",
    "TRIO-27.6-TL-OUTD-400": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/",
    "TRIO-27.6-TL-OUTD-S2-400": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/",
    "TRIO-27.6-TL-OUTD-S2F-400": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/",
    "TRIO-27.6-TL-OUTD-S2X-400": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/",
    "PVS-30-TL-SX": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/, SolarTraders https://solartraders.com/",
    "PVS-30-TL-SY": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/, SolarTraders https://solartraders.com/",
    "PVS-33-TL-SI": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/",
    "PVS-33-TL-SX": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/",
    "PVS-33-TL-SY": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/",
    "PVS-50-TL": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/",
    "TRIO-50.0-TL-OUTD": "Solar-Thailand https://solar-thailand.co.th/, Solaris Green Energy https://solaris.co.th/",
    # Afore - Thai: Electronmove; International: Panpower, ReestartSolar
    "BNT020KTL": "Electronmove https://electronmove.com/, Panpower https://panpower.eu/",
    "AF50K-TH": "Electronmove https://electronmove.com/, Alibaba https://www.alibaba.com/",
    "BNT050KTL": "Panpower https://panpower.eu/, ReestartSolar https://reestartsolar.com/",
    # AOTAI - Thai: Facebook group; International: Alibaba
    "ASP-50KTLC": "Aotai New Energy (Facebook), Alibaba https://www.alibaba.com/",
    # ATESS - Thai: Nastech Solar; International: Alibaba
    "HPS30": "Nastech Solar https://nastechsolar.com/, Alibaba https://www.alibaba.com/",
    # AUXSOL - Thai: N/A; International: Official site
    "ASG-20TL-ZH": "Auxsol Official https://www.auxsol.com/",
    "ASN-20TL": "Auxsol Official https://www.auxsol.com/",
    "ASN-30TL-G2": "Auxsol Official https://www.auxsol.com/",
    "ASN-40TL": "Auxsol Official https://www.auxsol.com/",
    "ASN-50TL-G2": "Auxsol Official https://www.auxsol.com/",
    # Bonaire
    "Bonaire XG50KTR": "N/A",
    # CHINT POWER - Thai: Solar-Thailand; International: ENF Solar
    "SCA20K-T-EU": "ENF Solar https://www.enfsolar.com/, Solar-Thailand https://solar-thailand.co.th/",
    # CHUPHOTIC
    "STT-20KTL-P": "N/A",
    # CLEANLINE
    "CL 30KTL-P3": "N/A",
    # DELTA - Thai: Delta Thailand; International: Europe-SolarStore
    "Delta 20000T": "Delta Thailand https://www.deltathailand.com/, Europe-SolarStore https://europe-solarstore.com/",
    "RPI-M20A": "Delta Thailand https://www.deltathailand.com/, Europe-SolarStore https://europe-solarstore.com/",
    "M50A_260": "Delta Thailand https://www.deltathailand.com/, Europe-SolarStore https://europe-solarstore.com/",
    "RPI M50A_020": "Delta Thailand https://www.deltathailand.com/, Europe-SolarStore https://europe-solarstore.com/",
    "RPI M50A_021": "Delta Thailand https://www.deltathailand.com/, Europe-SolarStore https://europe-solarstore.com/",
    "RPI M50A_022": "Delta Thailand https://www.deltathailand.com/, Europe-SolarStore https://europe-solarstore.com/",
    "RPI M50A_120": "Delta Thailand https://www.deltathailand.com/, Europe-SolarStore https://europe-solarstore.com/",
    "RPI M50A_121": "Delta Thailand https://www.deltathailand.com/, Europe-SolarStore https://europe-solarstore.com/",
    "RPI M50A_122": "Delta Thailand https://www.deltathailand.com/, Europe-SolarStore https://europe-solarstore.com/",
    "RPI M50A_12s": "Delta Thailand https://www.deltathailand.com/, Europe-SolarStore https://europe-solarstore.com/",
    # Deye - Thai: Shopee, Thaisolarsystem, Solaris; International: Alibaba
    "SUN-20K-G05": "Shopee https://shopee.co.th/, Solar-Thailand https://solar-thailand.co.th/, Thaisolarsystem https://www.thaisolarsystem.com/",
    "SUN-20K-SG05LP3-EU-SM2": "Shopee https://shopee.co.th/, Thaisolarsystem https://www.thaisolarsystem.com/",
    "SUN-30K-G04": "Shopee https://shopee.co.th/, Solar-Thailand https://solar-thailand.co.th/, R3 Solar Cell",
    "SUN-30K-SG01HP3-EU-BM3": "Shopee https://shopee.co.th/, Thaisolarsystem https://www.thaisolarsystem.com/",
    "SUN-30K-SG02HP3-EU-AM3": "Shopee https://shopee.co.th/, Thaisolarsystem https://www.thaisolarsystem.com/",
    "SUN-40K-G04": "Shopee https://shopee.co.th/, Solar-Thailand https://solar-thailand.co.th/, R3 Solar Cell",
    "SUN-40K-SG01HP3-EU-BM4": "Shopee https://shopee.co.th/, Thaisolarsystem https://www.thaisolarsystem.com/",
    "SUN-50K-G04": "Shopee https://shopee.co.th/, Solar-Thailand https://solar-thailand.co.th/, R3 Solar Cell",
    "SUN-50K-SG01HP3-EU-BM4": "Shopee https://shopee.co.th/, Thaisolarsystem https://www.thaisolarsystem.com/",
    # EVE
    "EP20KTR-2M3P": "N/A",
    "EP30KTR-2M3P": "N/A",
    "EP50KTR-2M3P": "N/A",
    # FOX ESS - Thai: N/A; International: Eco-Asset, SunBeam369
    "P3-Pro-20.0": "Eco-Asset https://eco-assemble.com/, SunBeam369 https://sunbeam369.com/",
    "T20-G3": "Eco-Asset https://eco-assemble.com/, Fox ESS Official https://www.fox-ess.com/",
    "P3-Pro-30.0": "Eco-Asset https://eco-assemble.com/, SunBeam369 https://sunbeam369.com/",
    "T30-M": "Eco-Asset https://eco-assemble.com/, Fox ESS Official https://www.fox-ess.com/",
    "V50": "Fox ESS Official https://www.fox-ess.com/",
    # FRECON
    "F025i-4PVb": "N/A",
    # FRONIUS - Thai: Electronmove, Lirik Solar; International: Wccsolar, Energy Freedom
    "Fronius Symo 20.0-3-M": "Electronmove https://electronmove.com/, Energy Freedom https://energyfreedom.com.au/, Wccsolar https://wccsolar.com/",
    "Fronius Eco 27.0-3-S": "Electronmove https://electronmove.com/, Wccsolar https://wccsolar.com/, Lirik Solar",
    # GOODWE - Thai: QES Energy, Solar Touch, Solar-Thailand; International: Official
    "GW20K-ET-L-G10": "QES Energy https://www.qes.co.th/, Solar Touch https://solartouch.co.th/, Solar-Thailand https://solar-thailand.co.th/",
    "GW20K-SDT-20": "QES Energy https://www.qes.co.th/, Solar-Thailand https://solar-thailand.co.th/",
    "GW30K-MT": "QES Energy https://www.qes.co.th/, Solar Touch https://solartouch.co.th/",
    "GW30K-SDT-C30": "QES Energy https://www.qes.co.th/, Solar Touch https://solartouch.co.th/",
    "GW50K-ET-10": "QES Energy https://www.qes.co.th/, Solar-Thailand https://solar-thailand.co.th/",
    "GW50K-SDT-C30": "QES Energy https://www.qes.co.th/, Solar-Thailand https://solar-thailand.co.th/",
    "GW50KS-MT": "QES Energy https://www.qes.co.th/, Shopee https://shopee.co.th/",
    # Growatt - Thai: Solar-Thailand, JPS Solar, Solar Tech Center; International: Official
    "Growatt 20000TL3-HE-TH": "Solar-Thailand https://solar-thailand.co.th/, JPS Solar",
    "Growatt 20000UE": "Solar-Thailand https://solar-thailand.co.th/",
    "MID 20KTL3-X": "Solar-Thailand https://solar-thailand.co.th/, Solar Tech Center https://solartech-center.com/",
    "Growatt 40000TL3-NS": "Solar-Thailand https://solar-thailand.co.th/, JPS Solar",
    "MID 40KTL3-X": "Solar-Thailand https://solar-thailand.co.th/, Solar Tech Center https://solartech-center.com/, JPS Solar",
    # Haier
    "HS3P-20KA": "N/A",
    "HS3P-50KA": "N/A",
    # hoymiles
    "HIT-20L-G3": "Hoymiles Official https://www.hoymiles.com/",
    "HIT-20L-G3S": "Hoymiles Official https://www.hoymiles.com/",
    "HiOne-20T-G3": "Hoymiles Official https://www.hoymiles.com/",
    # Huawei - Thai: Supersolarz, Thaisolarsystem, Double N Solar, Solaris; International: Official
    "SUN2000-20K-MB0": "Supersolarz https://www.supersolarz.com/, Thaisolarsystem https://www.thaisolarsystem.com/, Solaris https://solaris.co.th/",
    "SUN2000-20KTL": "Supersolarz https://www.supersolarz.com/, Thaisolarsystem https://www.thaisolarsystem.com/",
    "SUN2000-20KTL-M0": "Supersolarz https://www.supersolarz.com/, Thaisolarsystem https://www.thaisolarsystem.com/",
    "SUN2000-20KTL-M2": "Supersolarz https://www.supersolarz.com/, Thaisolarsystem https://www.thaisolarsystem.com/",
    "SUN2000-20KTL-M5": "Supersolarz https://www.supersolarz.com/, Thaisolarsystem https://www.thaisolarsystem.com/",
    "SUN2000-30K-MC0": "Supersolarz https://www.supersolarz.com/, Thaisolarsystem https://www.thaisolarsystem.com/",
    "SUN2000-30KTL-M3": "Supersolarz https://www.supersolarz.com/, Thaisolarsystem https://www.thaisolarsystem.com/",
    "SUN2000-33KTL": "Supersolarz https://www.supersolarz.com/, Double N Solar",
    "SUN2000-36KTL": "Supersolarz https://www.supersolarz.com/, Double N Solar",
    "SUN2000-36KTL-M3": "Supersolarz https://www.supersolarz.com/, Double N Solar",
    "SUN2000-40K-MC0": "Supersolarz https://www.supersolarz.com/, Thaisolarsystem https://www.thaisolarsystem.com/",
    "SUN2000-40KTL-M3": "Supersolarz https://www.supersolarz.com/, Thaisolarsystem https://www.thaisolarsystem.com/",
    "SUN2000-42KTL": "Supersolarz https://www.supersolarz.com/, Double N Solar",
    "SUN2000-50K-MC0": "Supersolarz https://www.supersolarz.com/, Thaisolarsystem https://www.thaisolarsystem.com/",
    "SUN2000-50KTL-M3": "Supersolarz https://www.supersolarz.com/, Thaisolarsystem https://www.thaisolarsystem.com/",
    # HYXiPOWER
    "HYX-H25K-HT": "N/A",
    "HYX-S30K-T": "N/A",
    "HYX-S50K-T": "N/A",
    # INVT
    "iMars BG20KTR": "INVT Official https://www.invt.com/",
    "iMars BG30KTR": "INVT Official https://www.invt.com/",
    "iMars BG50KTR": "INVT Official https://www.invt.com/",
    # ISUNA
    "Isuna 20000T": "N/A",
    # JFY
    "SUNTREE 20000TL": "N/A",
    "SUNTREE 30000TL": "N/A",
    # KACO
    "Blueplanet 20.0 TL3 M2 WM | OD IIG0": "KACO Official https://www.kaco-newenergy.com/",
    "Powador 30.0 TL3 - M – INT": "KACO Official https://www.kaco-newenergy.com/",
    "Powador 30.0 TL3 - XL - F - | INT": "KACO Official https://www.kaco-newenergy.com/",
    "Powador 30.0 TL3 - XL - F - | SPD 1+2": "KACO Official https://www.kaco-newenergy.com/",
    "Powador 30.0 TL3 - XL - INT | - SPD 1+2": "KACO Official https://www.kaco-newenergy.com/",
    "Powador 30.0 TL3 - XL – INT": "KACO Official https://www.kaco-newenergy.com/",
    "Powador 39.0 TL3-XL-INT": "KACO Official https://www.kaco-newenergy.com/",
    "Powador 60.0 TL3-XL-INT": "KACO Official https://www.kaco-newenergy.com/",
    "Blueplanet 50.0 TL3 M1 WM | OD IIGM": "KACO Official https://www.kaco-newenergy.com/",
    "Blueplanet 50.0 TL3 M1 WM | OD IIGX": "KACO Official https://www.kaco-newenergy.com/",
    # KEHUA TECH
    "SPI20K-B": "N/A",
    # KOSTAL
    "PIKO 20": "KOSTAL Official https://www.kostal.com/",
    # KSTAR
    "BluE-20KT-M1": "N/A",
    "KSG-20K": "N/A",
    "KAC50DP": "N/A",
    "KAC50DP2": "N/A",
    "KSG-50K": "N/A",
    # LANPWR
    "LANPWR 20000T": "N/A",
    "UHC 50KT-U2": "N/A",
    # LEONICS
    "APOLLO GTP-4020TL": "N/A",
    # LESSO
    "LSBH20KTL3-OC1": "N/A",
    # Litto
    "LT 20000HD": "N/A",
    # Midea
    "MEI2-HT20H": "N/A",
    "MEI2-HT30H": "N/A",
    # PEC
    "PCS-VACON-3-20G": "N/A",
    # Pixii
    "PowerShaper 2 ACU OD AUS | Pole Top": "Pixii Official https://pixii.com/",
    # PrimeVOLT
    "PV-20000T-U": "N/A",
    # PSI
    "P250": "N/A",
    "P300": "N/A",
    "P500": "N/A",
    # RENAC
    "NAC33K-DT": "RENAC Official https://www.renac.com/",
    # SAJ
    "H2-20K-LT2": "N/A",
    "CH2-50K-T6": "N/A",
    "R6-50K-T4-32": "N/A",
    # Schneider Electric
    "Conext CL 20000 E": "Schneider Official https://www.se.com/",
    "Conext CL 25000 E": "Schneider Official https://www.se.com/",
    "Conext CL36": "Schneider Official https://www.se.com/",
    # Sigenergy
    "SigenStor EC 20.0 TP": "N/A",
    "SigenStor EC 25.0 TP": "N/A",
    "Sigen PV 50M1": "N/A",
    "Sigen PV 50M1-HYA": "N/A",
    "Sigen PV 50M1-HYB": "N/A",
    # SINENG
    "SN20PT": "N/A",
    "SN30PT": "N/A",
    "SN50PT": "N/A",
    # Sinexcel
    "PMG2-50K": "N/A",
    # SMA - Thai: Techtron, Sunnergy, Thaifactorymart, SelfConcept; International: Official
    "STP 20-50": "Techtron https://www.techtron.co.th/, Sunnergy https://www.sunnergytech.com/, SelfConcept https://shop.selfconcept-onlineshop.com/",
    "STP 20000TL-30": "Techtron https://www.techtron.co.th/, Sunnergy https://www.sunnergytech.com/",
    "STP 20000TLEE-10": "Techtron https://www.techtron.co.th/, Sunnergy https://www.sunnergytech.com/",
    "STP 25-50": "Techtron https://www.techtron.co.th/, SelfConcept https://shop.selfconcept-onlineshop.com/",
    "STP 25000TL-30": "Techtron https://www.techtron.co.th/, Sunnergy https://www.sunnergytech.com/",
    "STP 50-40": "Solar-Thailand https://solar-thailand.co.th/, Techtron https://www.techtron.co.th/",
    "STP 50-41": "Solar-Thailand https://solar-thailand.co.th/, Techtron https://www.techtron.co.th/",
    # SOFAR - Thai: Thaisolarsystem; International: Shopee
    "SOFAR 20KTLX-G3": "Thaisolarsystem https://www.thaisolarsystem.com/, Shopee https://shopee.co.th/",
    "SOFAR 20KTLX2-G3P": "Thaisolarsystem https://www.thaisolarsystem.com/, Shopee https://shopee.co.th/",
    "SOFAR 30KTLX-G3": "Thaisolarsystem https://www.thaisolarsystem.com/, Shopee https://shopee.co.th/",
    "SOFAR 50KTLX-G3": "Thaisolarsystem https://www.thaisolarsystem.com/, Shopee https://shopee.co.th/",
    # Solar Edge
    "SE27.6K": "SolarEdge Official https://www.solaredge.com/",
    "SE30K": "SolarEdge Official https://www.solaredge.com/",
    "SE33.3K": "SolarEdge Official https://www.solaredge.com/",
    "PCS050": "SolarEdge Official https://www.solaredge.com/",
    # SOLAR ENERGY - Thai: Lazada
    "Sunlight-20KTH": "Lazada https://www.lazada.co.th/",
    # SOLAX POWER - Thai: Thaisolarsystem; International: Official
    "X3-PRO-20K-G2": "Thaisolarsystem https://www.thaisolarsystem.com/, SolaX Official https://th.solaxpower.com/",
    "X3-ULT-20K": "Thaisolarsystem https://www.thaisolarsystem.com/, SolaX Official https://th.solaxpower.com/",
    "X3-PRO-30K-G2": "Thaisolarsystem https://www.thaisolarsystem.com/, SolaX Official https://th.solaxpower.com/",
    "X3-ULT-30K": "Thaisolarsystem https://www.thaisolarsystem.com/, SolaX Official https://th.solaxpower.com/",
    "X3-MGA-40K-G2": "Thaisolarsystem https://www.thaisolarsystem.com/, SolaX Official https://th.solaxpower.com/",
    "X3-MGA-50K-G2": "Thaisolarsystem https://www.thaisolarsystem.com/, SolaX Official https://th.solaxpower.com/",
    # SOLINTEG
    "M2HT-50K-150": "Alibaba https://www.alibaba.com/, SOLINTEG Official https://www.solinteg.com/",
    # Solis - Thai: SPNM, QES Energy; International: Official
    "S5-GR3P20K": "SPNM https://www.spnm1.com/, QES Energy https://www.qes.co.th/, Lazada https://www.lazada.co.th/",
    "S5-GR3P20K(21A)": "SPNM https://www.spnm1.com/, QES Energy https://www.qes.co.th/",
    "S6-GC3P30K03-NV-ND": "SPNM https://www.spnm1.com/, QES Energy https://www.qes.co.th/",
    "S6-GC3P36K03-NV-ND": "SPNM https://www.spnm1.com/, QES Energy https://www.qes.co.th/",
    "S5-GC40K": "SPNM https://www.spnm1.com/, Lazada https://www.lazada.co.th/",
    "S6-GC3P40K04-NV-ND": "SPNM https://www.spnm1.com/, QES Energy https://www.qes.co.th/",
    "S6-EH3P50K-H": "SPNM https://www.spnm1.com/, QES Energy https://www.qes.co.th/",
    "S6-EH3P50K-H(21A)": "SPNM https://www.spnm1.com/, QES Energy https://www.qes.co.th/",
    # SUNGROW - Thai: Solaris, Solar-Thailand, Supersolarz, Mammoth Energy; International: Official
    "SG20RT": "Solaris https://solaris.co.th/, Solar-Thailand https://solar-thailand.co.th/, Supersolarz https://www.supersolarz.com/",
    "SG20RT-P2": "Solaris https://solaris.co.th/, Solar-Thailand https://solar-thailand.co.th/",
    "SH25T": "Solaris https://solaris.co.th/, Supersolarz https://www.supersolarz.com/",
    "SG36CX-P2": "Solaris https://solaris.co.th/, Supersolarz https://www.supersolarz.com/",
    "SG36KTL-M": "Solaris https://solaris.co.th/, Solar-Thailand https://solar-thailand.co.th/",
    "SC 50HV": "Solaris https://solaris.co.th/",
    "SG50CX": "Solaris https://solaris.co.th/, Solar-Thailand https://solar-thailand.co.th/",
    "SG50CX-P2": "Solaris https://solaris.co.th/, Solar-Thailand https://solar-thailand.co.th/",
    # TBEA
    "TS50KTL_PLUS": "N/A",
    # TCL
    "BlueArk X5-50k-100": "N/A",
    # Thai Tabuchi
    "TPD-T250P6-TH": "N/A",
    # TMDA
    "SUN-20K-G03": "N/A",
    "SUN-30K-G03": "N/A",
    "SUN-40K-G04": "N/A",
    "SUN-50K-G03": "N/A",
    # Trannergy
    "TRM033KTL": "N/A",
    # URECO
    "IS050B1": "N/A",
    # V SOLAR
    "VSOLAR 33K": "N/A",
    # VEICHI
    "VHT-50K-100-H": "N/A",
    # ZONERGY
    "Mars 50k": "N/A",
}

# Update Where to Buy column (column J = 10)
updated = 0
for row in range(2, ws.max_row + 1):
    model = ws.cell(row=row, column=3).value
    if model:
        model = model.strip()
    buy = where_to_buy.get(model, "N/A")
    ws.cell(row=row, column=10).value = buy
    updated += 1

wb.save(path)
print(f"Done! Updated {updated} rows with Where to Buy data")
