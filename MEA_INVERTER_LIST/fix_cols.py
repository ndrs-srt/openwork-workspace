import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

path = r'E:\01.NDRS_WORKON\OpenWork\MEA_INVERTER_LIST\MEA_Inverter_list_pass_standard_20-50kW.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active

# Delete columns H through K (cols 8-11) entirely
ws.delete_cols(8, 4)

# Verify
print(f"Cols after delete: {ws.max_column}")
print("Headers:", [ws.cell(1, c).value for c in range(1, ws.max_column + 1)])

wb.save(path)
print("Saved.")
