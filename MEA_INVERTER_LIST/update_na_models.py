import sys; sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

path = r'E:\01.NDRS_WORKON\OpenWork\MEA_INVERTER_LIST\MEA_Inverter_list_pass_standard_20-50kW.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active

# Model-level updates: row_number -> {web_link, where_to_buy, price}
updates = {
    30: {  # Bonaire Energia | Bonaire XG50KTR
        "web_link": "https://www.enfsolar.com/inverter-datasheet",
        "where_to_buy": "INVT Solar (OEM); Available via ENF Solar, SolarTraders",
        "price": "N/A (B2B)",
    },
    32: {  # CHUPHOTIC | STT-20KTL-P
        "web_link": "https://www.enfsolar.com/inverter-datasheet",
        "where_to_buy": "CHUPHOTIC Thailand; Sunways (OEM); Available via Solar-Thailand",
        "price": "~฿25,000–฿35,000",
    },
    33: {  # CLEANLINE | CL 30KTL-P3
        "web_link": "https://www.thaiwatsadu.com/ (CLEANLINE brand)",
        "where_to_buy": "Thai Watsadu; HomePro; Thai installer networks",
        "price": "~฿45,000–฿65,000 (ThaiWatsadu)",
    },
    49: {  # Deye | SUN-40K-G04
        "web_link": "https://deye.com/Inverter/Three-Phase-String-Inverter/",
        "where_to_buy": "Deye Thailand; Solar-Thailand; Shopee/Lazada",
        "price": "~฿35,000–฿55,000",
    },
    53: {  # EVE | EP20KTR-2M3P
        "web_link": "N/A (Only MEA/PEA approved list; no official product page)",
        "where_to_buy": "MEA/PEA registered; EVE Energy Co., Ltd (B2B only)",
        "price": "N/A (B2B)",
    },
    54: {  # EVE | EP30KTR-2M3P
        "web_link": "N/A (Only MEA/PEA approved list; no official product page)",
        "where_to_buy": "MEA/PEA registered; EVE Energy Co., Ltd (B2B only)",
        "price": "N/A (B2B)",
    },
    55: {  # EVE | EP50KTR-2M3P
        "web_link": "N/A (Only MEA/PEA approved list; no official product page)",
        "where_to_buy": "MEA/PEA registered; EVE Energy Co., Ltd (B2B only)",
        "price": "N/A (B2B)",
    },
    61: {  # FRECON | F025i-4PVb
        "web_link": "https://www.frecon.com.cn/static/upload/file/",
        "where_to_buy": "FRECON Electric (Shenzhen); Chinese distributors",
        "price": "N/A (B2B)",
    },
    76: {  # Haier | HS3P-20KA
        "web_link": "https://www.enfsolar.com/inverter-datasheet",
        "where_to_buy": "Haier Energy Technology; ENF Solar",
        "price": "~¥5,000–¥8,000 / ~฿25,000–฿40,000",
    },
    77: {  # Haier | HS3P-50KA
        "web_link": "https://www.enfsolar.com/inverter-datasheet",
        "where_to_buy": "Haier Energy Technology; ENF Solar",
        "price": "~¥10,000–¥15,000 / ~฿50,000–฿75,000",
    },
    96: {  # HYXiPOWER | HYX-H25K-HT
        "web_link": "https://webfile.hyxipower.com/soft/20251120/DS_HYX-H(15-25)K-HT_Datasheet_V1.4-20251118_EU.pdf",
        "where_to_buy": "HYXiPOWER Official https://www.hyxipower.com/",
        "price": "N/A (B2B)",
    },
    97: {  # HYXiPOWER | HYX-S30K-T
        "web_link": "https://webfile.hyxipower.com/soft/20251120/DS_HYX-S(30-50)K-T_Datasheet_V1.4-20251118_EU.pdf",
        "where_to_buy": "HYXiPOWER Official https://www.hyxipower.com/",
        "price": "N/A (B2B)",
    },
    98: {  # HYXiPOWER | HYX-S50K-T
        "web_link": "https://webfile.hyxipower.com/soft/20251120/DS_HYX-S(30-50)K-T_Datasheet_V1.4-20251118_EU.pdf",
        "where_to_buy": "HYXiPOWER Official https://www.hyxipower.com/",
        "price": "N/A (B2B)",
    },
    102: {  # ISUNA | Isuna 20000T
        "web_link": "https://www.sinexcel-isuna.com/upload/image/",
        "where_to_buy": "ISUNA/Sinexcel Official; ENF Solar; fuseco.com.au",
        "price": "N/A (B2B)",
    },
    103: {  # JFY | SUNTREE 20000TL
        "web_link": "https://www.enfsolar.com/inverter-datasheet",
        "where_to_buy": "JFY Tech; ENF Solar; International Energy Solutions",
        "price": "N/A (B2B)",
    },
    104: {  # JFY | SUNTREE 30000TL
        "web_link": "https://www.enfsolar.com/inverter-datasheet",
        "where_to_buy": "JFY Tech; ENF Solar; International Energy Solutions",
        "price": "N/A (B2B)",
    },
    115: {  # KEHUA TECH | SPI20K-B
        "web_link": "https://solarapi.kehua.com/string-inverter/",
        "where_to_buy": "Kehua Tech Official https://www.kehua.com/download/",
        "price": "N/A (B2B)",
    },
    117: {  # KSTAR | BluE-20KT-M1
        "web_link": "https://so-en.pl/wp-content/uploads/2024/02/KSG-10K-20K.pdf",
        "where_to_buy": "KSTAR/KFR Power; ENF Solar",
        "price": "N/A (B2B)",
    },
    118: {  # KSTAR | KSG-20K
        "web_link": "https://so-en.pl/wp-content/uploads/2024/02/KSG-10K-20K.pdf",
        "where_to_buy": "KSTAR/KFR Power; ENF Solar",
        "price": "N/A (B2B)",
    },
    119: {  # KSTAR | KAC50DP
        "web_link": "https://www.enfsolar.com/inverter-datasheet",
        "where_to_buy": "KSTAR/KFR Power; ENF Solar",
        "price": "N/A (B2B)",
    },
    120: {  # KSTAR | KAC50DP2
        "web_link": "https://www.enfsolar.com/inverter-datasheet",
        "where_to_buy": "KSTAR/KFR Power; ENF Solar",
        "price": "N/A (B2B)",
    },
    121: {  # KSTAR | KSG-50K
        "web_link": "https://www.enfsolar.com/inverter-datasheet",
        "where_to_buy": "KSTAR/KFR Power; ENF Solar",
        "price": "N/A (B2B)",
    },
    122: {  # LANPWR | LANPWR 20000T
        "web_link": "https://th.lanpwr.com/",
        "where_to_buy": "LANPWR Thailand https://th.lanpwr.com/; Shopee/Lazada",
        "price": "N/A (B2B)",
    },
    123: {  # LANPWR | UHC 50KT-U2
        "web_link": "https://th.lanpwr.com/",
        "where_to_buy": "LANPWR Thailand https://th.lanpwr.com/; Shopee/Lazada",
        "price": "N/A (B2B)",
    },
    124: {  # LEONICS | APOLLO GTP-4020TL
        "web_link": "https://www.leonics.com/product/renewable/inverter/",
        "where_to_buy": "LEONICS Official https://www.leonics.com/ (Thai company)",
        "price": "N/A (B2B)",
    },
    125: {  # LESSO | LSBH20KTL3-OC1
        "web_link": "https://en.lesso.com/upfile/2023/04/Datasheets.pdf",
        "where_to_buy": "China LESSO Group https://en.lesso.com/",
        "price": "N/A (B2B)",
    },
    126: {  # Litto | LT 20000HD
        "web_link": "https://www.enfsolar.com/inverter-datasheet",
        "where_to_buy": "Litto Official; ELBA-ENERGA; EOS Energia; Thai distributors",
        "price": "N/A (B2B)",
    },
    127: {  # Midea | MEI2-HT20H
        "web_link": "https://www.midea.com/th/solar",
        "where_to_buy": "Midea Thailand; Solar-Thailand; Solaris",
        "price": "N/A (B2B)",
    },
    128: {  # Midea | MEI2-HT30H
        "web_link": "https://www.midea.com/th/solar",
        "where_to_buy": "Midea Thailand; Solar-Thailand; Solaris",
        "price": "N/A (B2B)",
    },
    129: {  # PEC | PCS-VACON-3-20G
        "web_link": "N/A (PEC Group Thailand; VACON drive-based inverter; B2B only)",
        "where_to_buy": "PEC Group Thailand (B2B)",
        "price": "N/A (B2B)",
    },
    131: {  # PrimeVOLT | PV-20000T-U
        "web_link": "https://primevolt.com.tw/uploads/2020/05/",
        "where_to_buy": "PrimeVOLT Taiwan https://primevolt.com.tw/",
        "price": "N/A (B2B)",
    },
    132: {  # PSI | P250
        "web_link": "https://psienergy.co.th/",
        "where_to_buy": "PSI Energy Thailand https://psienergy.co.th/; Shopee",
        "price": "~฿40,000–฿55,000 (Shopee)",
    },
    133: {  # PSI | P300
        "web_link": "https://psienergy.co.th/",
        "where_to_buy": "PSI Energy Thailand https://psienergy.co.th/; Shopee",
        "price": "~฿50,000–฿65,000",
    },
    134: {  # PSI | P500
        "web_link": "https://psienergy.co.th/",
        "where_to_buy": "PSI Energy Thailand https://psienergy.co.th/; Shopee",
        "price": "~฿65,000–฿90,000",
    },
    136: {  # SAJ | H2-20K-LT2
        "web_link": "https://www.saj-electric.com/hubfs/H2%208-20K-LT2.pdf",
        "where_to_buy": "SAJ Electric Official https://www.saj-electric.com/",
        "price": "N/A (B2B)",
    },
    137: {  # SAJ | CH2-50K-T6
        "web_link": "https://www.saj-electric.com/services-support-download",
        "where_to_buy": "SAJ Electric Official https://www.saj-electric.com/",
        "price": "N/A (B2B)",
    },
    138: {  # SAJ | R6-50K-T4-32
        "web_link": "https://www.saj-electric.com/services-support-download",
        "where_to_buy": "SAJ Electric Official https://www.saj-electric.com/",
        "price": "N/A (B2B)",
    },
    142: {  # Sigenergy | SigenStor EC 20.0 TP
        "web_link": "https://www.sigenergy.com/products/hybrid-inverter",
        "where_to_buy": "Sigenergy Official https://www.sigenergy.com/",
        "price": "N/A (B2B)",
    },
    143: {  # Sigenergy | SigenStor EC 25.0 TP
        "web_link": "https://www.sigenergy.com/products/hybrid-inverter",
        "where_to_buy": "Sigenergy Official https://www.sigenergy.com/",
        "price": "N/A (B2B)",
    },
    144: {  # Sigenergy | Sigen PV 50M1
        "web_link": "https://www.sigenergy.com/products/business-inverter",
        "where_to_buy": "Sigenergy Official https://www.sigenergy.com/",
        "price": "N/A (B2B)",
    },
    145: {  # Sigenergy | Sigen PV 50M1-HYA
        "web_link": "https://www.sigenergy.com/products/business-inverter",
        "where_to_buy": "Sigenergy Official https://www.sigenergy.com/",
        "price": "N/A (B2B)",
    },
    146: {  # Sigenergy | Sigen PV 50M1-HYB
        "web_link": "https://www.sigenergy.com/products/business-inverter",
        "where_to_buy": "Sigenergy Official https://www.sigenergy.com/",
        "price": "N/A (B2B)",
    },
    147: {  # SINENG | SN20PT
        "web_link": "https://en.si-neng.com/Public/uploadfile/files/",
        "where_to_buy": "Sineng Electric https://en.si-neng.com/",
        "price": "N/A (B2B)",
    },
    148: {  # SINENG | SN30PT
        "web_link": "https://en.si-neng.com/Public/uploadfile/files/",
        "where_to_buy": "Sineng Electric https://en.si-neng.com/",
        "price": "N/A (B2B)",
    },
    149: {  # SINENG | SN50PT
        "web_link": "https://en.si-neng.com/Public/uploadfile/files/",
        "where_to_buy": "Sineng Electric https://en.si-neng.com/",
        "price": "N/A (B2B)",
    },
    150: {  # Sinexcel | PMG2-50K
        "web_link": "https://www.sinexcel.com/",
        "where_to_buy": "Sinexcel Official https://www.sinexcel.com/",
        "price": "N/A (B2B)",
    },
    190: {  # TBEA | TS50KTL_PLUS
        "web_link": "https://www.tbea.com/product/solar-inverter/",
        "where_to_buy": "TBEA Official https://www.tbea.com/",
        "price": "N/A (B2B)",
    },
    191: {  # TCL | BlueArk X5-50k-100
        "web_link": "https://digitalpower.tcl.com/products/blueArk-x5",
        "where_to_buy": "TCL Digital Power https://digitalpower.tcl.com/",
        "price": "N/A (B2B)",
    },
    192: {  # Thai Tabuchi | TPD-T250P6-TH
        "web_link": "N/A (Thai Tabuchi; Factory Bang Pakong, Chachoengsao; MEA/PEA approved)",
        "where_to_buy": "Thai Tabuchi Electric; Bang Pakong Factory; Mercury Tower Bangkok",
        "price": "N/A (B2B)",
    },
    193: {  # TMDA | SUN-20K-G03
        "web_link": "https://godungfaifaa.com/ (TMDA product pages)",
        "where_to_buy": "TMDA Thailand; godungfaifaa.com; Shopee; Facebook TMDA INVERTER THAILAND",
        "price": "~฿20,000–฿30,000",
    },
    194: {  # TMDA | SUN-30K-G03
        "web_link": "https://godungfaifaa.com/ (TMDA product pages)",
        "where_to_buy": "TMDA Thailand; godungfaifaa.com; Shopee; Facebook TMDA INVERTER THAILAND",
        "price": "~฿25,000–฿40,000",
    },
    195: {  # TMDA | SUN-40K-G04
        "web_link": "https://godungfaifaa.com/ (TMDA product pages)",
        "where_to_buy": "TMDA Thailand; godungfaifaa.com; Shopee; Facebook TMDA INVERTER THAILAND",
        "price": "~฿35,000–฿50,000",
    },
    196: {  # TMDA | SUN-50K-G03
        "web_link": "https://godungfaifaa.com/ (TMDA product pages)",
        "where_to_buy": "TMDA Thailand; godungfaifaa.com; Shopee; Facebook TMDA INVERTER THAILAND",
        "price": "~฿45,000–฿60,000",
    },
    197: {  # Trannergy | TRM033KTL
        "web_link": "https://www.enfsolar.com/inverter-datasheet",
        "where_to_buy": "Trannergy Official; ENF Solar; solarsmart.ie",
        "price": "N/A (B2B)",
    },
    198: {  # URECO | IS050B1
        "web_link": "N/A (Thai brand; MEA/PEA registered; no official website)",
        "where_to_buy": "URECO Thailand (B2B only)",
        "price": "N/A (B2B)",
    },
    199: {  # V SOLAR | VSOLAR 33K
        "web_link": "N/A (V SOLAR brand; MEA/PEA registered)",
        "where_to_buy": "V SOLAR; MEA/PEA registered; Pakistani solar traders",
        "price": "N/A (B2B)",
    },
    200: {  # VEICHI | VHT-50K-100-H
        "web_link": "https://www.veichi.com/product/on-grid-inverter/",
        "where_to_buy": "VEICHI Electric Official https://www.veichi.com/",
        "price": "N/A (B2B)",
    },
    201: {  # ZONERGY | Mars 50k
        "web_link": "https://www.zonergy.com/download-category/full-series/",
        "where_to_buy": "Zonergy Official https://www.zonergy.com/",
        "price": "N/A (B2B)",
    },
}

# Styles
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
center_align = Alignment(horizontal="center", vertical="center")

updated = 0
for row_num, data in updates.items():
    web_cell = ws.cell(row=row_num, column=9)
    where_cell = ws.cell(row=row_num, column=10)
    price_cell = ws.cell(row=row_num, column=11)
    
    web_cell.value = data["web_link"]
    where_cell.value = data["where_to_buy"]
    price_cell.value = data["price"]
    
    for cell in [web_cell, where_cell, price_cell]:
        cell.border = thin_border
    
    web_cell.alignment = Alignment(wrap_text=True, vertical="center")
    where_cell.alignment = Alignment(wrap_text=True, vertical="center")
    price_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    updated += 1
    brand = ws.cell(row=row_num, column=2).value
    model = ws.cell(row=row_num, column=3).value
    print(f"Row {row_num}: {brand} | {model} -> Updated")

wb.save(path)
print(f"\nDone! Updated {updated} rows with Web Link, Where to Buy, and Price data.")
