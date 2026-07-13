import sys; sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Alignment, Border, Side

path = r'E:\01.NDRS_WORKON\OpenWork\MEA_INVERTER_LIST\MEA_Inverter_list_pass_standard_20-50kW.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

# Fix all broken/incorrect Web Links
fixes = {
    # === AUXSOL - correct model-specific links ===
    25: "https://www.auxsol.com/hybrid-inverter/hybrid-inverter-three-phase-5-20kw.html",
    26: "https://www.auxsol.com/all-in-one/2c91838092e5a9d001933e79844700fc.html",
    27: "https://www.auxsol.com/all-in-one/on-grid-three-phase-30kw-eng.html",
    28: "https://www.auxsol.com/all-in-one/three-phase-on-grid-33-40kw-eng.html",
    29: "https://www.auxsol.com/cl-inverter/on-grid-inverter-three-phase-50-80kw-eng.html",

    # === ENF Solar 404s -> use official manufacturer sites ===
    30: "N/A (Bonaire Energia / INVT Solar OEM; no public product page)",
    32: "https://www.sunways.com/products/solar-inverter/stt-4-25ktl-p/",
    33: "https://www.thaiwatsadu.com/",
    76: "https://www.haier.com/newenergy/",
    77: "https://www.haier.com/newenergy/",
    103: "https://www.jfrgroup.com/product/suntree-tl/",
    104: "https://www.jfrgroup.com/product/suntree-tl/",
    119: "https://www.kfrpower.com/product/kac50dp/",
    120: "https://www.kfrpower.com/product/kac50dp2/",
    121: "https://www.kfrpower.com/product/ksg-50k/",
    126: "https://www.enfsolar.com/pv/inverter",

    # === ISUNA ===
    102: "https://www.sinexcel-isuna.com/Products/",

    # === KSTAR - use official download center ===
    117: "https://www.kstar.com/download?category=solar",
    118: "https://www.kstar.com/download?category=solar",

    # === FRECON ===
    61: "https://www.frecon.com.cn/",

    # === KEHUA TECH ===
    115: "https://www.kehua.com/product/solar-inverter/",

    # === LANPWR ===
    122: "https://th.lanpwr.com/",
    123: "https://th.lanpwr.com/",

    # === LESSO ===
    125: "https://en.lesso.com/",

    # === Midea ===
    127: "https://www.midea.com/th/products/energy-storage-system",
    128: "https://www.midea.com/th/products/energy-storage-system",

    # === PrimeVOLT ===
    131: "https://primevolt.com.tw/product/three-phase-on-grid-pv-inverter/",

    # === RENAC ===
    135: "https://www.renac.com/product/three-phase-string-inverter/",

    # === SAJ ===
    136: "https://www.saj-electric.com/hubfs/Download/H2%208-20K-LT2.pdf",
    137: "https://www.saj-electric.com/services-support-download",
    138: "https://www.saj-electric.com/services-support-download",

    # === SINENG ===
    147: "https://en.si-neng.com/product/string-inverter/",
    148: "https://en.si-neng.com/product/string-inverter/",
    149: "https://en.si-neng.com/product/string-inverter/",

    # === Sinexcel ===
    150: "https://www.sinexcel.com/product/",

    # === SOLAR ENERGY ===
    166: "https://shopee.co.th/search?keyword=SOLAR+ENERGY+SUNLIGHT-20KTH",

    # === TBEA ===
    190: "https://www.tbea.com/product/",

    # === TMDA ===
    193: "https://www.facebook.com/TMDA-INVERTER-THAILAND/",
    194: "https://www.facebook.com/TMDA-INVERTER-THAILAND/",
    195: "https://www.facebook.com/TMDA-INVERTER-THAILAND/",
    196: "https://www.facebook.com/TMDA-INVERTER-THAILAND/",

    # === VEICHI ===
    200: "https://www.veichi.com/product/on-grid-solar-inverter/",

    # === ZONERGY ===
    201: "https://www.zonergy.com/product/on-grid-inverter/",
}

updated = 0
for row_num, new_link in fixes.items():
    cell = ws.cell(row=row_num, column=9)
    old_val = cell.value
    cell.value = new_link
    cell.border = thin_border
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    brand = ws.cell(row=row_num, column=2).value
    model = ws.cell(row=row_num, column=3).value
    updated += 1
    print(f"Row {row_num}: {brand} | {model}")
    print(f"  OLD: {old_val}")
    print(f"  NEW: {new_link}")

wb.save(path)
print(f"\nDone! Fixed {updated} Web Links.")
