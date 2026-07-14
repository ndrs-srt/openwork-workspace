import openpyxl
import sys
import json

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Working_Space\OpenWork\MEA_INVERTER_LIST\MEA_Inverter_list_pass_standard_20-50kW.xlsx')
ws = wb.active

all_models = []
for r in range(2, ws.max_row + 1):
    brand = ws.cell(r, 2).value or ""
    model = ws.cell(r, 3).value or ""
    col_g = ws.cell(r, 7).value or ""
    col_h = ws.cell(r, 8).value or ""
    all_models.append({
        "row": r,
        "brand": brand,
        "model": model,
        "iec62477": col_g,
        "iec61000": col_h
    })

# Print summary
total = len(all_models)
g_yes = sum(1 for m in all_models if m["iec62477"] == "✅")
g_no = sum(1 for m in all_models if m["iec62477"] == "❌")
h_yes = sum(1 for m in all_models if m["iec61000"] == "✅")
h_no = sum(1 for m in all_models if m["iec61000"] == "❌")

print(f"Total models: {total}")
print(f"IEC 62477: {g_yes} ✅, {g_no} ❌")
print(f"IEC 61000: {h_yes} ✅, {h_no} ❌")
print()

# Print all models
for m in all_models:
    print(f"Row {m['row']}: {m['brand']} | {m['model']} | 62477={m['iec62477']} | 61000={m['iec61000']}")
