"""
Fix Midea models: IEC 62477 and Web Link
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

EXCEL_FILE = r"MEA_Inverter_list_pass_standard_20-50kW.xlsx"

# Load workbook
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active

# Official Midea PowerX1 datasheet link
MIDEA_DATASHEET = "https://www.mideahvac.fr/wp-content/uploads/sites/2/2026/05/26FPM_FR_PowerX1-Tri.pdf"

# Midea corrections
# Row 127: MEI2-HT20H - IEC 62477 should be ❌ (not confirmed in datasheet)
# Row 128: MEI2-HT30H - IEC 62477 should be ❌ (not confirmed in datasheet)
# Both have: EN 62109-1/2, EN/IEC 61000-6-1/2/3/4

changes = []

# Row 127: MEI2-HT20H
row = 127
old_g = ws.cell(row=row, column=7).value
old_link = ws.cell(row=row, column=9).value
if '✅' in str(old_g):
    ws.cell(row=row, column=7).value = '❌'
    changes.append(f"Row {row} (MEI2-HT20H): IEC 62477 ✅ → ❌ (datasheet confirms EN 62109 only)")
ws.cell(row=row, column=9).value = MIDEA_DATASHEET
changes.append(f"Row {row} (MEI2-HT20H): Web Link updated to official Midea PowerX1 datasheet")

# Row 128: MEI2-HT30H
row = 128
old_g = ws.cell(row=row, column=7).value
old_link = ws.cell(row=row, column=9).value
if '✅' in str(old_g):
    ws.cell(row=row, column=7).value = '❌'
    changes.append(f"Row {row} (MEI2-HT30H): IEC 62477 ✅ → ❌ (datasheet confirms EN 62109 only)")
ws.cell(row=row, column=9).value = MIDEA_DATASHEET
changes.append(f"Row {row} (MEI2-HT30H): Web Link updated to official Midea PowerX1 datasheet")

# Save
wb.save(EXCEL_FILE)
print("Midea corrections applied:")
for c in changes:
    print(f"  ✓ {c}")

print(f"\nTotal changes: {len(changes)}")
