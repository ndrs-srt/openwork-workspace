import sys; sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

path = r'E:\01.NDRS_WORKON\OpenWork\MEA_INVERTER_LIST\MEA_Inverter_list_pass_standard_20-50kW.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active

# Brand-level data: web_link, where_to_buy, price, iec62477, iec61000
# Prices are approximate from Google search results (THB or USD)
data = {
    "ABB/FIMER": {
        "iec62477": "❌",
        "iec61000": "✅",
        "web_link": "https://www.fimer.com/three-phase/pvs-203033-tl",
        "where_to_buy": "https://www.fimer.com/products-and-services/solar/string-inverters",
        "price": "N/A (B2B)",
    },
    "Afore": {
        "iec62477": "❌",
        "iec61000": "✅",
        "web_link": "https://www.aforenergy.com/product/three-phase-hybrid-inverter-36-50kw/",
        "where_to_buy": "https://panpower.eu/en/inverters/6797-afore-bnt050ktl-50-kw.html",
        "price": "€1,530–€2,900 / ~฿57,000–฿108,000",
    },
    "AOTAI": {
        "iec62477": "❌",
        "iec61000": "✅",
        "web_link": "https://www.aotaiwelding.com/products/photovoltaic/inverter/asp_5060ktlc.html",
        "where_to_buy": "https://www.aotaiwelding.com/",
        "price": "$2,500–$3,500 / ~฿90,000–฿126,000",
    },
    "ATESS": {
        "iec62477": "❌",
        "iec61000": "✅",
        "web_link": "https://www.atesspower.com/hybrid-inverter/hps30-50-100-120-150",
        "where_to_buy": "https://www.atesspower.com/contact-us",
        "price": "$5,500–$6,700 / ~฿198,000–฿241,000",
    },
    "AUXSOL": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "https://www.auxsol.com/cl-inverter/on-grid-inverter-three-phase-35-75kw-en.html",
        "where_to_buy": "https://www.auxsol.com/",
        "price": "$600–$3,500 / ~฿21,600–฿126,000",
    },
    "Bonaire Energia": {
        "iec62477": "❌",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "CHINT POWER": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "https://www.chintpowersystems.com/",
        "where_to_buy": "https://www.enfsolar.com/inverter-datasheet/chint-power",
        "price": "$1,500–$2,200 / ~฿54,000–฿79,000",
    },
    "CHUPHOTIC": {
        "iec62477": "❌",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "CLEANLINE": {
        "iec62477": "❌",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "Delta": {
        "iec62477": "❌",
        "iec61000": "✅",
        "web_link": "https://www.deltathailand.com/PV-Inverter",
        "where_to_buy": "https://www.deltathailand.com/",
        "price": "€2,073–€5,000 / ~฿78,000–฿190,000",
    },
    "DELTA": {
        "iec62477": "❌",
        "iec61000": "✅",
        "web_link": "https://www.deltathailand.com/PV-Inverter",
        "where_to_buy": "https://www.deltathailand.com/",
        "price": "€2,073–€5,000 / ~฿78,000–฿190,000",
    },
    "Deye": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "https://www.deyeinverter.com/",
        "where_to_buy": "https://www.thaisolarsystem.com/",
        "price": "฿125,000–฿219,000",
    },
    "EVE": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "FOX ESS": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "https://www.fox-ess.com/",
        "where_to_buy": "https://eco-assemble.com/",
        "price": "€1,948–€3,000 / ~฿73,000–฿112,000",
    },
    "FRECON": {
        "iec62477": "❌",
        "iec61000": "✅",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "FRONIUS": {
        "iec62477": "❌",
        "iec61000": "✅",
        "web_link": "https://www.fronius.com/en/solar-energy/installers-partners/products/all-products/inverters",
        "where_to_buy": "https://www.electronmove.com/",
        "price": "฿110,000–฿155,000",
    },
    "GOODWE": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "https://en.goodwe.com/",
        "where_to_buy": "https://www.qes.co.th/",
        "price": "฿49,000–฿157,400",
    },
    "Growatt": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "https://en.growatt.com/",
        "where_to_buy": "https://www.solar-thailand.co.th/",
        "price": "฿77,000–฿168,000",
    },
    "Haier": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "hoymiles": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "https://www.hoymiles.com/",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "Huawei": {
        "iec62477": "❌",
        "iec61000": "✅",
        "web_link": "https://solar.huawei.com/en/products/inverters",
        "where_to_buy": "https://www.supersolarz.com/",
        "price": "฿69,000–฿132,000",
    },
    "HUAWEI": {
        "iec62477": "❌",
        "iec61000": "✅",
        "web_link": "https://solar.huawei.com/en/products/inverters",
        "where_to_buy": "https://www.supersolarz.com/",
        "price": "฿69,000–฿132,000",
    },
    "HYXiPOWER": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "INVT": {
        "iec62477": "❌",
        "iec61000": "❌",
        "web_link": "https://www.invt.com/",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "ISUNA": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "JFY": {
        "iec62477": "❌",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "KACO": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "https://www.kaco-newenergy.com/en/products/solar-inverters",
        "where_to_buy": "https://www.kaco-newenergy.com/en/contact",
        "price": "N/A",
    },
    "KEHUA TECH": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "KOSTAL": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "https://www.kostal.com/en/solar-electrics/solar-inverters",
        "where_to_buy": "https://www.kostal.com/en/contact",
        "price": "N/A",
    },
    "KSTAR": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "LANPWR": {
        "iec62477": "❌",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "LEONICS": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "LESSO": {
        "iec62477": "❌",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "Litto": {
        "iec62477": "❌",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "Midea": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "PEC": {
        "iec62477": "❌",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "Pixii": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "PrimeVOLT": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "PSI": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "RENAC": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "SAJ": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "Schneider Electric": {
        "iec62477": "❌",
        "iec61000": "✅",
        "web_link": "https://www.se.com/ww/en/product-range/2250-solar-inverters/",
        "where_to_buy": "https://www.se.com/ww/en/contacts/",
        "price": "N/A",
    },
    "Sigenergy": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "SINENG": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "Sinexcel": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "SMA": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "https://www.sma-solar.com/en/products/solar-inverters/",
        "where_to_buy": "https://www.techtron.co.th/",
        "price": "฿125,000–฿273,200",
    },
    "SOFAR": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "https://www.sofarsolar.com/",
        "where_to_buy": "https://www.thaisolarsystem.com/",
        "price": "฿32,000–฿85,000",
    },
    "Solar Edge": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "https://www.solaredge.com/products/inverters",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "SOLAR ENERGY": {
        "iec62477": "❌",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "SOLAX POWER": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "https://th.solaxpower.com/",
        "where_to_buy": "https://www.thaisolarsystem.com/",
        "price": "฿23,540–฿47,000",
    },
    "SOLINTEG": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "https://www.solinteg.com/",
        "where_to_buy": "https://www.alibaba.com/",
        "price": "฿81,000–฿88,000",
    },
    "solis": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "https://www.ginlong.com/",
        "where_to_buy": "https://www.spnm1.com/",
        "price": "฿55,500–฿62,500",
    },
    "SUNGROW": {
        "iec62477": "✅",
        "iec61000": "✅",
        "web_link": "https://www.sungrowpower.com/",
        "where_to_buy": "https://solaris.co.th/",
        "price": "฿63,800–฿135,000",
    },
    "TBEA": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "TCL": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "Thai Tabuchi": {
        "iec62477": "❌",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "TMDA": {
        "iec62477": "❌",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "Trannergy": {
        "iec62477": "❌",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "URECO": {
        "iec62477": "❌",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "V SOLAR": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "VEICHI": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
    "ZONERGY": {
        "iec62477": "✅",
        "iec61000": "❌",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A",
    },
}

# Styles
header_font = Font(bold=True)
header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
center_align = Alignment(horizontal="center", vertical="center")

# Set headers (G-K)
headers = ["IEC 62477", "IEC 61000", "Web Link", "Where to Buy", "Price"]
col_widths = [14, 14, 45, 40, 30]
for i, (h, w) in enumerate(zip(headers, col_widths)):
    col = i + 7
    cell = ws.cell(row=1, column=col)
    cell.value = h
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

# Populate data
matched = 0
for row in range(2, ws.max_row + 1):
    brand = ws.cell(row=row, column=2).value
    if brand:
        brand = brand.strip()
    d = data.get(brand, None)
    
    for col_idx, key in enumerate(["iec62477", "iec61000", "web_link", "where_to_buy", "price"]):
        cell = ws.cell(row=row, column=col_idx + 7)
        cell.border = thin_border
        if d:
            cell.value = d[key]
            if key in ("iec62477", "iec61000"):
                cell.alignment = center_align
            matched += 1
        else:
            cell.value = "—"

wb.save(path)
print(f"Done! {matched} cells written for {ws.max_row - 1} models")
