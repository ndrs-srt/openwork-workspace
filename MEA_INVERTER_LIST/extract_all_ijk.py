import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
import json

wb = openpyxl.load_workbook('MEA_Inverter_list_pass_standard_20-50kW.xlsx')
ws = wb.active

# Extract all data for columns I, J, K
all_data = []
for r in range(2, ws.max_row + 1):
    brand = str(ws.cell(row=r, column=2).value or '').strip()
    model = str(ws.cell(row=r, column=3).value or '').strip()
    link = str(ws.cell(row=r, column=9).value or '').strip()
    buy = str(ws.cell(row=r, column=10).value or '').strip()
    price = str(ws.cell(row=r, column=11).value or '').strip()
    
    all_data.append({
        'row': r,
        'brand': brand,
        'model': model,
        'link': link,
        'buy': buy,
        'price': price
    })

# Group by brand
brands = {}
for d in all_data:
    b = d['brand']
    if b not in brands:
        brands[b] = []
    brands[b].append(d)

# Print summary
print("=" * 120)
print("ALL BRANDS AND MODELS - COLUMNS I J K STATUS")
print("=" * 120)

for brand, models in sorted(brands.items()):
    print(f"\n{'='*80}")
    print(f"BRAND: {brand} ({len(models)} models)")
    print(f"{'='*80}")
    for m in models:
        link_short = m['link'][:70] + '...' if len(m['link']) > 70 else m['link']
        print(f"  Row {m['row']}: {m['model']}")
        print(f"    Link: {link_short}")
        print(f"    Buy:  {m['buy'][:60]}")
        print(f"    Price: {m['price']}")
        print()

# Save to file for reference
with open('all_ijk_data.json', 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

print(f"\nTotal models: {len(all_data)}")
print(f"Total brands: {len(brands)}")
print("Data saved to all_ijk_data.json")
