import sys; sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

path = r'E:\01.NDRS_WORKON\OpenWork\MEA_INVERTER_LIST\MEA_Inverter_list_pass_standard_20-50kW.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active

# Model-specific web links from Google search results
model_links = {
    # ABB/FIMER - PVS-20/30/33 series
    "PVS-20-TL-SX": "https://www.fimer.com/three-phase/pvs-203033-tl",
    "PVS-20-TL-SXD": "https://www.fimer.com/three-phase/pvs-203033-tl",
    "PVS-20-TL-SY": "https://www.fimer.com/three-phase/pvs-203033-tl",
    "PVS-30-TL-SX": "https://www.fimer.com/three-phase/pvs-203033-tl",
    "PVS-30-TL-SY": "https://www.fimer.com/three-phase/pvs-203033-tl",
    "PVS-33-TL-SI": "https://www.fimer.com/three-phase/pvs-203033-tl",
    "PVS-33-TL-SX": "https://www.fimer.com/three-phase/pvs-203033-tl",
    "PVS-33-TL-SY": "https://www.fimer.com/three-phase/pvs-203033-tl",
    "PVS-50-TL": "https://www.fimer.com/three-phase/pvs-5060-tl",
    # ABB/FIMER - TRIO series
    "TRIO-20.0-TL-OUTD-400": "https://www.fimer.com/three-phase/trio-200276-tl",
    "TRIO-20.0-TL-OUTD-S2-400": "https://www.fimer.com/three-phase/trio-200276-tl",
    "TRIO-20.0-TL-OUTD-S2F-400": "https://www.fimer.com/three-phase/trio-200276-tl",
    "TRIO-20.0-TL-OUTD-S2X-400": "https://www.fimer.com/three-phase/trio-200276-tl",
    "TRIO-27.6-TL-OUTD-400": "https://www.fimer.com/three-phase/trio-200276-tl",
    "TRIO-27.6-TL-OUTD-S2-400": "https://www.fimer.com/three-phase/trio-200276-tl",
    "TRIO-27.6-TL-OUTD-S2F-400": "https://www.fimer.com/three-phase/trio-200276-tl",
    "TRIO-27.6-TL-OUTD-S2X-400": "https://www.fimer.com/three-phase/trio-200276-tl",
    "TRIO-50.0-TL-OUTD": "https://www.fimer.com/three-phase/trio-500-tl",
    # Afore
    "BNT020KTL": "https://www.aforenergy.com/product/three-phase-string-inverter-3-25kw/",
    "AF50K-TH": "https://www.aforenergy.com/product/three-phase-hybrid-inverter-36-50kw/",
    "BNT050KTL": "https://www.aforenergy.com/product/three-phase-string-inverter-36-60kw/",
    # AOTAI
    "ASP-50KTLC": "https://www.aotaiwelding.com/products/photovoltaic/inverter/asp_5060ktlc.html",
    # ATESS
    "HPS30": "https://www.atesspower.com/hybrid-inverter/hps30-50-100-120-150",
    # AUXSOL
    "ASG-20TL-ZH": "https://www.auxsol.com/hybrid-inverter/asg-20tl-zh-en.html",
    "ASN-20TL": "https://www.auxsol.com/cl-inverter/on-grid-inverter-three-phase-35-75kw-en.html",
    "ASN-30TL-G2": "https://www.auxsol.com/cl-inverter/on-grid-inverter-three-phase-35-75kw-en.html",
    "ASN-40TL": "https://www.auxsol.com/cl-inverter/on-grid-inverter-three-phase-35-75kw-en.html",
    "ASN-50TL-G2": "https://www.auxsol.com/cl-inverter/on-grid-inverter-three-phase-35-75kw-en.html",
    # Bonaire
    "Bonaire XG50KTR": "N/A",
    # CHINT POWER
    "SCA20K-T-EU": "https://www.chintpowersystems.com/product/sca5-25k-t-eu/",
    # CHUPHOTIC
    "STT-20KTL-P": "N/A",
    # CLEANLINE
    "CL 30KTL-P3": "N/A",
    # DELTA
    "Delta 20000T": "https://www.deltathailand.com/PV-Inverter",
    "RPI-M20A": "https://www.deltathailand.com/PV-Inverter",
    "M50A_260": "https://www.deltathailand.com/PV-Inverter",
    "RPI M50A_020": "https://www.deltathailand.com/PV-Inverter",
    "RPI M50A_021": "https://www.deltathailand.com/PV-Inverter",
    "RPI M50A_022": "https://www.deltathailand.com/PV-Inverter",
    "RPI M50A_120": "https://www.deltathailand.com/PV-Inverter",
    "RPI M50A_121": "https://www.deltathailand.com/PV-Inverter",
    "RPI M50A_122": "https://www.deltathailand.com/PV-Inverter",
    "RPI M50A_12s": "https://www.deltathailand.com/PV-Inverter",
    # Deye
    "SUN-20K-G05": "https://www.deyeinverter.com/product/three-phase-string-inverter/sun18-20-25kg04-1825kw-three-phase-2-mppt-237.html",
    "SUN-20K-SG05LP3-EU-SM2": "https://www.deyeinverter.com/product/three-phase-hybrid-inverter/sun-20-25k-sg01hp3-eu-bm3-2025kw-three-phase-hybrid-140.html",
    "SUN-30K-G04": "https://www.deyeinverter.com/product/three-phase-string-inverter/sun30-33-36kg04-3036kw-three-phase-2-mppt.html",
    "SUN-30K-SG01HP3-EU-BM3": "https://www.deyeinverter.com/product/three-phase-hybrid-inverter/sun-29-9-30-35-40-50k-sg01hp3-eu-bm3-bm4-29950kw-three-phase-hybrid.html",
    "SUN-30K-SG02HP3-EU-AM3": "https://www.deyeinverter.com/product/three-phase-hybrid-inverter/sun-30k-sg02hp3-eu-am3-30kw-three-phase-hybrid-142.html",
    "SUN-40K-G04": "https://www.deyeinverter.com/product/three-phase-string-inverter/sun40-45-50kg04-4050kw-three-phase-4-mppt.html",
    "SUN-40K-SG01HP3-EU-BM4": "https://www.deyeinverter.com/product/three-phase-hybrid-inverter/sun-29-9-30-35-40-50k-sg01hp3-eu-bm3-bm4-29950kw-three-phase-hybrid.html",
    "SUN-50K-G04": "https://www.deyeinverter.com/product/three-phase-string-inverter/sun40-45-50kg04-4050kw-three-phase-4-mppt.html",
    "SUN-50K-SG01HP3-EU-BM4": "https://www.deyeinverter.com/product/three-phase-hybrid-inverter/sun-29-9-30-35-40-50k-sg01hp3-eu-bm3-bm4-29950kw-three-phase-hybrid.html",
    # EVE
    "EP20KTR-2M3P": "N/A",
    "EP30KTR-2M3P": "N/A",
    "EP50KTR-2M3P": "N/A",
    # FOX ESS
    "P3-Pro-20.0": "https://www.fox-ess.com/product/p3-pro/",
    "T20-G3": "https://www.fox-ess.com/product/t-pro/",
    "P3-Pro-30.0": "https://www.fox-ess.com/product/p3-pro/",
    "T30-M": "https://www.fox-ess.com/product/t-pro/",
    "V50": "https://www.fox-ess.com/product/v-series/",
    # FRECON
    "F025i-4PVb": "N/A",
    # FRONIUS
    "Fronius Symo 20.0-3-M": "https://www.fronius.com/en/solar-energy/installers-partners/products/all-products/inverters/symo-200-to-500",
    "Fronius Eco 27.0-3-S": "https://www.fronius.com/en/solar-energy/installers-partners/products/all-products/inverters/eco-250-to-270",
    # GOODWE
    "GW20K-ET-L-G10": "https://en.goodwe.com/et-lv-series",
    "GW20K-SDT-20": "https://en.goodwe.com/sdt-series",
    "GW30K-MT": "https://en.goodwe.com/mt-series",
    "GW30K-SDT-C30": "https://en.goodwe.com/sdt-series",
    "GW50K-ET-10": "https://en.goodwe.com/et-series",
    "GW50K-SDT-C30": "https://en.goodwe.com/sdt-series",
    "GW50KS-MT": "https://en.goodwe.com/mt-series",
    # Growatt
    "Growatt 20000TL3-HE-TH": "https://en.growatt.com/products/mid-11-30ktl3-xh",
    "Growatt 20000UE": "https://en.growatt.com/products",
    "MID 20KTL3-X": "https://en.growatt.com/products/mid-15-30ktl3-x-x2",
    "Growatt 40000TL3-NS": "https://en.growatt.com/products/mid-30-50ktl3-x2",
    "MID 40KTL3-X": "https://en.growatt.com/products/mid-30-50ktl3-x2",
    # Haier
    "HS3P-20KA": "N/A",
    "HS3P-50KA": "N/A",
    # hoymiles
    "HIT-20L-G3": "https://www.hoymiles.com/products/hit-series/",
    "HIT-20L-G3S": "https://www.hoymiles.com/products/hit-series/",
    "HiOne-20T-G3": "https://www.hoymiles.com/products/hione-series/",
    # Huawei
    "SUN2000-20K-MB0": "https://solar.huawei.com/en/products/sun2000-12-15-17-20-25k-mb0/",
    "SUN2000-20KTL": "https://solar.huawei.com/en/products/sun2000-20-40ktl-m3/",
    "SUN2000-20KTL-M0": "https://solar.huawei.com/en/products/sun2000-20-40ktl-m0/",
    "SUN2000-20KTL-M2": "https://solar.huawei.com/en/products/sun2000-20-40ktl-m2/",
    "SUN2000-20KTL-M5": "https://solar.huawei.com/en/products/sun2000-20-40ktl-m5/",
    "SUN2000-30K-MC0": "https://solar.huawei.com/en/products/sun2000-30-40k-mc0/",
    "SUN2000-30KTL-M3": "https://solar.huawei.com/en/products/sun2000-20-40ktl-m3/",
    "SUN2000-33KTL": "https://solar.huawei.com/en/products/sun2000-20-40ktl/",
    "SUN2000-36KTL": "https://solar.huawei.com/en/products/sun2000-36ktl/",
    "SUN2000-36KTL-M3": "https://solar.huawei.com/en/products/sun2000-36ktl-m3/",
    "SUN2000-40K-MC0": "https://solar.huawei.com/en/products/sun2000-30-40k-mc0/",
    "SUN2000-40KTL-M3": "https://solar.huawei.com/en/products/sun2000-20-40ktl-m3/",
    "SUN2000-42KTL": "https://solar.huawei.com/en/products/sun2000-42ktl/",
    "SUN2000-50K-MC0": "https://solar.huawei.com/en/products/sun2000-50k-mc0/",
    "SUN2000-50KTL-M3": "https://solar.huawei.com/en/products/sun2000-50ktl-m3/",
    # HYXiPOWER
    "HYX-H25K-HT": "N/A",
    "HYX-S30K-T": "N/A",
    "HYX-S50K-T": "N/A",
    # INVT
    "iMars BG20KTR": "https://www.invt.com/products/solar-inverters",
    "iMars BG30KTR": "https://www.invt.com/products/solar-inverters",
    "iMars BG50KTR": "https://www.invt.com/products/solar-inverters",
    # ISUNA
    "Isuna 20000T": "N/A",
    # JFY
    "SUNTREE 20000TL": "N/A",
    "SUNTREE 30000TL": "N/A",
    # KACO
    "Blueplanet 20.0 TL3 M2 WM | OD IIG0": "https://www.kaco-newenergy.com/en/products/solar-inverters/blueplanet-gridsave",
    "Powador 30.0 TL3 - M – INT": "https://www.kaco-newenergy.com/en/products/solar-inverters/powador",
    "Powador 30.0 TL3 - XL - F - | INT": "https://www.kaco-newenergy.com/en/products/solar-inverters/powador",
    "Powador 30.0 TL3 - XL - F - | SPD 1+2": "https://www.kaco-newenergy.com/en/products/solar-inverters/powador",
    "Powador 30.0 TL3 - XL - INT | - SPD 1+2": "https://www.kaco-newenergy.com/en/products/solar-inverters/powador",
    "Powador 30.0 TL3 - XL – INT": "https://www.kaco-newenergy.com/en/products/solar-inverters/powador",
    "Powador 39.0 TL3-XL-INT": "https://www.kaco-newenergy.com/en/products/solar-inverters/powador",
    "Powador 60.0 TL3-XL-INT": "https://www.kaco-newenergy.com/en/products/solar-inverters/powador",
    "Blueplanet 50.0 TL3 M1 WM | OD IIGM": "https://www.kaco-newenergy.com/en/products/solar-inverters/blueplanet-gridsave",
    "Blueplanet 50.0 TL3 M1 WM | OD IIGX": "https://www.kaco-newenergy.com/en/products/solar-inverters/blueplanet-gridsave",
    # KEHUA TECH
    "SPI20K-B": "N/A",
    # KOSTAL
    "PIKO 20": "https://www.kostal.com/en/solar-electrics/solar-inverters/piko-20",
    # KSTAR
    "BluE-20KT-M1": "N/A",
    "KSG-20K": "N/A",
    "KAC50DP": "N/A",
    "KAC50DP2": "N/A",
    "KSG-50K": "N/A",
    # LANPWR
    "LANPWR 20000T": "N/A",
    "UHC 50KT-U2": "N/A",
    # LEONICS
    "APOLLO GTP-4020TL": "N/A",
    # LESSO
    "LSBH20KTL3-OC1": "N/A",
    # Litto
    "LT 20000HD": "N/A",
    # Midea
    "MEI2-HT20H": "N/A",
    "MEI2-HT30H": "N/A",
    # PEC
    "PCS-VACON-3-20G": "N/A",
    # Pixii
    "PowerShaper 2 ACU OD AUS | Pole Top": "N/A",
    # PrimeVOLT
    "PV-20000T-U": "N/A",
    # PSI
    "P250": "N/A",
    "P300": "N/A",
    "P500": "N/A",
    # RENAC
    "NAC33K-DT": "N/A",
    # SAJ
    "H2-20K-LT2": "N/A",
    "CH2-50K-T6": "N/A",
    "R6-50K-T4-32": "N/A",
    # Schneider Electric
    "Conext CL 20000 E": "https://www.se.com/ww/en/product-range/2250-solar-inverters/",
    "Conext CL 25000 E": "https://www.se.com/ww/en/product-range/2250-solar-inverters/",
    "Conext CL36": "https://www.se.com/ww/en/product-range/2250-solar-inverters/",
    # Sigenergy
    "SigenStor EC 20.0 TP": "N/A",
    "SigenStor EC 25.0 TP": "N/A",
    "Sigen PV 50M1": "N/A",
    "Sigen PV 50M1-HYA": "N/A",
    "Sigen PV 50M1-HYB": "N/A",
    # SINENG
    "SN20PT": "N/A",
    "SN30PT": "N/A",
    "SN50PT": "N/A",
    # Sinexcel
    "PMG2-50K": "N/A",
    # SMA
    "STP 20-50": "https://www.sma-solar.com/en/products/solar-inverters/sunny-tripower-x-2050",
    "STP 20000TL-30": "https://www.sma-solar.com/en/products/solar-inverters/sunny-tripower-20000tl-30",
    "STP 20000TLEE-10": "https://www.sma-solar.com/en/products/solar-inverters/sunny-tripower-20000tlee-10",
    "STP 25-50": "https://www.sma-solar.com/en/products/solar-inverters/sunny-tripower-x-2550",
    "STP 25000TL-30": "https://www.sma-solar.com/en/products/solar-inverters/sunny-tripower-25000tl-30",
    "STP 50-40": "https://www.sma-solar.com/en/products/solar-inverters/sunny-tripower-core1-stp-50-40",
    "STP 50-41": "https://www.sma-solar.com/en/products/solar-inverters/sunny-tripower-core1-stp-50-41",
    # SOFAR
    "SOFAR 20KTLX-G3": "https://www.sofarsolar.com/products/sofar-20ktlx-g3",
    "SOFAR 20KTLX2-G3P": "https://www.sofarsolar.com/products/sofar-20ktlx2-g3p",
    "SOFAR 30KTLX-G3": "https://www.sofarsolar.com/products/sofar-30ktlx-g3",
    "SOFAR 50KTLX-G3": "https://www.sofarsolar.com/products/sofar-50ktlx-g3",
    # Solar Edge
    "SE27.6K": "https://www.solaredge.com/products/inverters/single-phase-inverters",
    "SE30K": "https://www.solaredge.com/products/inverters/three-phase-inverters",
    "SE33.3K": "https://www.solaredge.com/products/inverters/three-phase-inverters",
    "PCS050": "https://www.solaredge.com/products/inverters/three-phase-inverters",
    # SOLAR ENERGY
    "Sunlight-20KTH": "N/A",
    # SOLAX POWER
    "X3-PRO-20K-G2": "https://th.solaxpower.com/x3-pro-g2",
    "X3-ULT-20K": "https://th.solaxpower.com/x3-ult",
    "X3-PRO-30K-G2": "https://th.solaxpower.com/x3-pro-g2",
    "X3-ULT-30K": "https://th.solaxpower.com/x3-ult",
    "X3-MGA-40K-G2": "https://th.solaxpower.com/x3-mga",
    "X3-MGA-50K-G2": "https://th.solaxpower.com/x3-mga",
    # SOLINTEG
    "M2HT-50K-150": "https://www.solinteg.com/m2ht",
    # Solis
    "S5-GR3P20K": "https://www.ginlong.com/solis-s5-gr3p-series/",
    "S5-GR3P20K(21A)": "https://www.ginlong.com/solis-s5-gr3p-series/",
    "S6-GC3P30K03-NV-ND": "https://www.ginlong.com/solis-s6-gc3p-series/",
    "S6-GC3P36K03-NV-ND": "https://www.ginlong.com/solis-s6-gc3p-series/",
    "S5-GC40K": "https://www.ginlong.com/solis-s5-gc-series/",
    "S6-GC3P40K04-NV-ND": "https://www.ginlong.com/solis-s6-gc3p-series/",
    "S6-EH3P50K-H": "https://www.ginlong.com/solis-s6-eh3p-series/",
    "S6-EH3P50K-H(21A)": "https://www.ginlong.com/solis-s6-eh3p-series/",
    # SUNGROW
    "SG20RT": "https://www.sungrowpower.com/en/products/string-inverter/b-sg20rt",
    "SG20RT-P2": "https://www.sungrowpower.com/en/products/string-inverter/b-sg20rt-p2",
    "SH25T": "https://www.sungrowpower.com/en/products/hybrid-inverter/sh-series",
    "SG36CX-P2": "https://www.sungrowpower.com/en/products/string-inverter/b-sg36-40-50cx-p2",
    "SG36KTL-M": "https://www.sungrowpower.com/en/products/string-inverter/sg36ktl-m",
    "SC 50HV": "https://www.sungrowpower.com/en/products/string-inverter/sc-series",
    "SG50CX": "https://www.sungrowpower.com/en/products/string-inverter/b-sg36-40-50cx-p2",
    "SG50CX-P2": "https://www.sungrowpower.com/en/products/string-inverter/b-sg36-40-50cx-p2",
    # TBEA
    "TS50KTL_PLUS": "N/A",
    # TCL
    "BlueArk X5-50k-100": "N/A",
    # Thai Tabuchi
    "TPD-T250P6-TH": "N/A",
    # TMDA
    "SUN-20K-G03": "N/A",
    "SUN-30K-G03": "N/A",
    "SUN-40K-G04": "N/A",
    "SUN-50K-G03": "N/A",
    # Trannergy
    "TRM033KTL": "N/A",
    # URECO
    "IS050B1": "N/A",
    # V SOLAR
    "VSOLAR 33K": "N/A",
    # VEICHI
    "VHT-50K-100-H": "N/A",
    # ZONERGY
    "Mars 50k": "N/A",
}

# Update Web Link column (column I = 9)
updated = 0
for row in range(2, ws.max_row + 1):
    model = ws.cell(row=row, column=3).value
    if model:
        model = model.strip()
    link = model_links.get(model, "N/A")
    ws.cell(row=row, column=9).value = link
    updated += 1

wb.save(path)
print(f"Done! Updated {updated} rows with direct model links")
