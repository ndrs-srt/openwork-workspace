import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load workbook
wb = openpyxl.load_workbook(r'E:\01.NDRS_WORKON\OpenWork\MEA_INVERTER_LIST\MEA_Inverter_list_pass_standard_20-50kW.xlsx')
ws = wb.active

# IEC 62477 compliance per brand (confirmed from datasheet searches)
iec62477 = {
    "ABB/FIMER": False,
    "Afore": False,
    "AOTAI": False,
    "ATESS": False,
    "AUXSOL": True,
    "Bonaire Energia": False,
    "CHINT POWER": True,
    "CHUPHOTIC": False,
    "CLEANLINE": False,
    "Delta": False,
    "DELTA": False,
    "Deye": True,
    "EVE": True,
    "FOX ESS": True,
    "FRECON": False,
    "FRONIUS": False,
    "GOODWE": True,
    "Growatt": True,
    "Haier": True,
    "hoymiles": True,
    "Huawei": False,
    "HUAWEI": False,
    "HYXiPOWER": True,
    "INVT": False,
    "ISUNA": True,
    "JFY": False,
    "KACO": True,
    "KEHUA TECH": True,
    "KOSTAL": True,
    "KSTAR": True,
    "LANPWR": False,
    "LEONICS": True,
    "LESSO": False,
    "Litto": False,
    "Midea": True,
    "PEC": False,
    "Pixii": True,
    "PrimeVOLT": True,
    "PSI": True,
    "RENAC": True,
    "SAJ": True,
    "Schneider Electric": False,
    "Sigenergy": True,
    "SINENG": True,
    "Sinexcel": True,
    "SMA": True,
    "SOFAR": True,
    "Solar Edge": True,
    "SOLAR ENERGY": False,
    "SOLAX POWER": True,
    "SOLINTEG": True,
    "solis": True,
    "SUNGROW": True,
    "TBEA": True,
    "TCL": True,
    "Thai Tabuchi": False,
    "TMDA": False,
    "Trannergy": False,
    "URECO": False,
    "V SOLAR": True,
    "VEICHI": True,
    "ZONERGY": True,
}

# Styles
header_font = Font(bold=True)
header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
center_align = Alignment(horizontal="center", vertical="center")

# Set column G header
cell = ws.cell(row=1, column=7)
cell.value = "IEC 62477"
cell.font = header_font
cell.fill = header_fill
cell.border = thin_border
cell.alignment = center_align

# Set column width
ws.column_dimensions["G"].width = 18

# Populate data
matched = 0
unmatched = 0
for row in range(2, ws.max_row + 1):
    brand = ws.cell(row=row, column=2).value
    if brand:
        brand = brand.strip()
    found = iec62477.get(brand, None)
    
    cell = ws.cell(row=row, column=7)
    cell.border = thin_border
    cell.alignment = center_align
    
    if found is True:
        cell.value = "✅"
        matched += 1
    elif found is False:
        cell.value = "❌"
        matched += 1
    else:
        cell.value = "—"
        unmatched += 1

# Clear old columns H-K
for col in range(8, 13):
    col_letter = openpyxl.utils.get_column_letter(col)
    if col_letter in ws.column_dimensions:
        del ws.column_dimensions[col_letter]
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=col).value = None

wb.save(r'E:\01.NDRS_WORKON\OpenWork\MEA_INVERTER_LIST\MEA_Inverter_list_pass_standard_20-50kW.xlsx')

print(f"Done! {matched} brands matched, {unmatched} unknown")
