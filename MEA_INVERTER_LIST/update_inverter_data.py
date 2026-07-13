import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

FILE_PATH = r"E:\01.NDRS_WORKON\OpenWork\MEA_INVERTER_LIST\MEA_Inverter_list_pass_standard_20-50kW.xlsx"

brand_data = {
    "ABB/FIMER": {
        "safety": "IEC/EN 62109-1, IEC/EN 62109-2",
        "emc": "EN 61000-6-1, EN 61000-6-2, EN 61000-3-11, EN 61000-3-12",
        "web_link": "https://www.fimer.com/products/solar-inverters/string-inverters/",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "Afore": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "EN/IEC 61000-6-2, EN/IEC 61000-6-3",
        "web_link": "https://www.aforenergy.com/product/three-phase-3-25kw/",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "AOTAI": {
        "safety": "IEC 62109-1/2",
        "emc": "IEC 61000-6-2/4",
        "web_link": "https://www.aotaiwelding.com/products/photovoltaic/inverter/asp_5060ktlc.html",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "ATESS": {
        "safety": "IEC 62109-1, IEC 62109-2, EN 62109-1:2010, EN 62109-2:2011",
        "emc": "EN 61000-6-2:2019, EN 61000-6-4:2019",
        "web_link": "https://www.atesspower.com/hybrid-inverter/hps30-50-100-120-150",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "AUXSOL": {
        "safety": "IEC/EN 62109-1/2",
        "emc": "EN IEC 61000-6-1/2/3/4, EN IEC 61000-3-11, EN 61000-3-12",
        "web_link": "https://www.auxsol.com/cl-inverter/on-grid-inverter-three-phase-35-75kw-en.html",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "Bonaire Energia": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "CHINT POWER": {
        "safety": "IEC/EN 62109-1/2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "https://www.chintpowersystems.com/",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "CHUPHOTIC": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "CLEANLINE": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "Delta": {
        "safety": "IEC 62109-1/-2",
        "emc": "EN 61000-6-2, EN 61000-6-3",
        "web_link": "https://solarsolutions.delta-emea.com/",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "DELTA": {
        "safety": "IEC 62109-1/-2",
        "emc": "EN 61000-6-2, EN 61000-6-3",
        "web_link": "https://solarsolutions.delta-emea.com/",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "Deye": {
        "safety": "IEC/EN 62109-1, IEC/EN 62109-2",
        "emc": "IEC/EN 61000-6-1/2/3/4",
        "web_link": "https://deye.com/product-category/inverter/three-phase-string-inverter/",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "EVE": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "FOX ESS": {
        "safety": "EN 62109-1, EN 62109-2",
        "emc": "IEC 61000-6-2, IEC 61000-6-3",
        "web_link": "https://www.fox-ess.com/",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "FRECON": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "FRONIUS": {
        "safety": "IEC 62109-1/-2",
        "emc": "EN 61000-6-2, EN 61000-6-3",
        "web_link": "https://www.fronius.com/en/solar-energy/installers-partners/products/all-products/inverters",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "GOODWE": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "https://en.goodwe.com/",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "Growatt": {
        "safety": "IEC 62109-1/-2",
        "emc": "EN 61000-6-1, EN 61000-6-2, EN 61000-6-3, EN 61000-6-4",
        "web_link": "https://en.growatt.com/",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "Haier": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "hoymiles": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "https://www.hoymiles.com/",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "Huawei": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "EN 55032, EN 55035, EN/IEC 61000-3-2, EN/IEC 61000-3-3",
        "web_link": "https://solar.huawei.com/en/products/inverters",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "HUAWEI": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "EN 55032, EN 55035, EN/IEC 61000-3-2, EN/IEC 61000-3-3",
        "web_link": "https://solar.huawei.com/en/products/inverters",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "HYXiPOWER": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "INVT": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "https://www.invt.com/products/xg-50-70kw-solar-inverter-211",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "ISUNA": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "JFY": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "KACO": {
        "safety": "IEC 62109-1/-2",
        "emc": "EN 61000-6-2, EN 61000-6-3",
        "web_link": "https://www.kaco-newenergy.com/en/products/solar-inverters",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "KEHUA TECH": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "KOSTAL": {
        "safety": "IEC 62109-1/-2",
        "emc": "EN 61000-6-2, EN 61000-6-3",
        "web_link": "https://www.kostal.com/en/solar-electrics/solar-inverters",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "KSTAR": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "LANPWR": {
        "safety": "N/A",
        "emc": "N/A",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "LEONICS": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "LESSO": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "Litto": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "Midea": {
        "safety": "IEC 62109-1, IEC 62109-2, IEC 62477-1",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "PEC": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "Pixii": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "PrimeVOLT": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "PSI": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "RENAC": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "SAJ": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "Schneider Electric": {
        "safety": "IEC 62109-1/-2",
        "emc": "IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "https://www.se.com/ww/en/product-range/2250-solar-inverters/",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "Sigenergy": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "SINENG": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "Sinexcel": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "SMA": {
        "safety": "IEC 62109-1/-2",
        "emc": "EN 61000-6-2, EN 61000-6-3, EN 61000-6-4",
        "web_link": "https://www.sma-solar.com/en/products/solar-inverters/",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "SOFAR": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "https://www.sofarsolar.com/products/",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "Solar Edge": {
        "safety": "IEC-62109-1, IEC-62109-2",
        "emc": "EN/IEC 61000-6-1, EN/IEC 61000-6-2, EN/IEC 61000-6-3, EN/IEC 61000-6-4, EN 55011",
        "web_link": "https://www.solaredge.com/products/inverters/three-phase-inverters",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "SOLAR ENERGY": {
        "safety": "N/A",
        "emc": "N/A",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "SOLAX POWER": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "https://www.solaxpower.com/product/x3-pro-g2/",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "SOLINTEG": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "https://www.solinteg.com/integ-m/",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "solis": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "https://www.ginlong.com/solis-s5-gr3p-series/",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "SUNGROW": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "https://www.sungrowpower.com/",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "TBEA": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "TCL": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "Thai Tabuchi": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "TMDA": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "Trannergy": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "URECO": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "V SOLAR": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "VEICHI": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
    "ZONERGY": {
        "safety": "IEC 62109-1, IEC 62109-2",
        "emc": "IEC 61000-6-1, IEC 61000-6-2, IEC 61000-6-3, IEC 61000-6-4",
        "web_link": "N/A",
        "where_to_buy": "N/A",
        "price": "N/A"
    },
}

# Build a normalized lookup (lowercase keys) for case-insensitive matching
brand_lookup = {k.strip().lower(): v for k, v in brand_data.items()}

# Column letters for the new data
# G = Safety, H = EMC, I = Web Link, J = Where to Buy, K = Price
HEADERS = {
    "G": "IEC 62477/62109 (Safety)",
    "H": "IEC 61000 (EMC)",
    "I": "Web Link",
    "J": "Where to Buy",
    "K": "Price",
}
COLUMN_KEYS = ["safety", "emc", "web_link", "where_to_buy", "price"]

COL_WIDTHS = {"G": 35, "H": 45, "I": 50, "J": 25, "K": 15}

# Styles
header_font = Font(bold=True)
header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
wrap_alignment = Alignment(wrap_text=True, vertical="top")


def main():
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active

    # Write headers in row 1
    for col_letter, header_text in HEADERS.items():
        cell = ws[f"{col_letter}1"]
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    matched = 0
    not_matched = []

    for row in range(2, ws.max_row + 1):
        brand_raw = ws.cell(row, 2).value  # Column B
        if brand_raw is None:
            not_matched.append((row, "EMPTY"))
            continue

        brand_key = str(brand_raw).strip().lower()
        data = brand_lookup.get(brand_key)

        if data is None:
            not_matched.append((row, str(brand_raw)))
            continue

        matched += 1
        for col_letter, key in zip(HEADERS.keys(), COLUMN_KEYS):
            cell = ws[f"{col_letter}{row}"]
            cell.value = data[key]
            cell.border = thin_border
            cell.alignment = wrap_alignment

    # Apply border to empty new cells (rows that didn't match)
    for row in range(2, ws.max_row + 1):
        for col_letter in HEADERS.keys():
            cell = ws[f"{col_letter}{row}"]
            if cell.value is None:
                cell.value = ""
            cell.border = thin_border
            cell.alignment = wrap_alignment

    wb.save(FILE_PATH)

    total_rows = ws.max_row - 1
    print(f"Done! Processed {total_rows} inverter models.")
    print(f"  Matched:    {matched}")
    print(f"  Not matched: {len(not_matched)}")
    if not_matched:
        print("\nUnmatched brands (row: brand):")
        for row, brand in not_matched:
            print(f"  Row {row}: {brand}")


if __name__ == "__main__":
    main()
