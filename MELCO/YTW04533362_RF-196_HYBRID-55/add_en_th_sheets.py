import os
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font

base = r"C:\Working_Space\OpenWork\MELCO\YTW04533362_RF-196_HYBRID-55"
path = os.path.join(base, "NDRS_shipping_template_from_image.xlsx")

wb = load_workbook(path)

# Remove old translated sheets if re-run
for name in ["English", "ภาษาไทย"]:
    if name in wb.sheetnames:
        del wb[name]

src = wb["ถอดแบบจากรูป"]

def copy_sheet_with_layout(title):
    ws = wb.copy_worksheet(src)
    ws.title = title
    # keep print settings
    ws.page_setup.paperSize = src.page_setup.paperSize
    ws.page_setup.orientation = src.page_setup.orientation
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws

def set_font_range(ws, max_row=14, max_col=8, name="Arial"):
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                f = copy(cell.font)
                f.name = name
                cell.font = f

def fill_sheet(ws, lang):
    if lang == "en":
        set_font_range(ws, name="Arial")
        ws["A1"], ws["B1"], ws["C1"], ws["D1"] = "Location", "Category", "Work details", "Item"
        ws["E1"], ws["F1"], ws["G1"], ws["H1"] = "Qty", "PIC", "Period", "Remarks"
        ws["A2"] = "NDRS Office"
        ws["B2"] = "Domestic transport"
        ws["C2"] = "Send equipment from NDRS to ME-TH (hand-carry by car)"
        ws["F2"] = "MELCO, ME-TH"
        ws["G2"] = "Day1"
        rows = [
            ("RF", 196, "Terminals brought from ME-TWN"),
            ("Hybrid", 55, "Terminals brought from ME-TWN"),
            ("MIS cable", 22, "Cable brought from ME-TWN"),
            ("Voltage label", 63, "For 8 units sent to Thailand in advance + 55 Hybrid units missed in Taiwan"),
            ("Measuring instrument", 3, "Instrument used for NBTC certification"),
            ("Instrument cable", 3, "Cable used for NBTC certification"),
            ("RF", 4, "Terminal used for NBTC certification (NBTC certification label cannot be attached?)"),
            ("Hybrid", 4, "Terminal used for NBTC certification (NBTC certification label cannot be attached?)"),
            ("Power strip", 5, "Order to NDRS"),
            ("Label", "-", "Order to NDRS"),
            ("", "", "Apply after receiving invoice; acquisition period is approx. 2 weeks"),
            ("MIS cable (without power)", 50, "Order to NDRS"),
            ("AC/USB adapter", 13, "Order to NDRS"),
        ]
        ws["C11"] = "Confirm number of USB serial cables"
    else:
        set_font_range(ws, name="Arial")
        ws["A1"], ws["B1"], ws["C1"], ws["D1"] = "สถานที่", "หมวดหมู่", "รายละเอียดงาน", "รายการ"
        ws["E1"], ws["F1"], ws["G1"], ws["H1"] = "จำนวน", "ผู้รับผิดชอบ", "ระยะเวลา", "หมายเหตุ"
        ws["A2"] = "สำนักงาน NDRS"
        ws["B2"] = "ขนส่งภายในประเทศ"
        ws["C2"] = "ส่งอุปกรณ์จาก NDRS ไป ME-TH (Hand-carry ด้วยรถ)"
        ws["F2"] = "MELCO, ME-TH"
        ws["G2"] = "Day1"
        rows = [
            ("RF", 196, "เครื่องที่นำมาจาก ME-TWN"),
            ("Hybrid", 55, "เครื่องที่นำมาจาก ME-TWN"),
            ("สาย MIS", 22, "สายเคเบิลที่นำมาจาก ME-TWN"),
            ("สติ๊กเกอร์แรงดันไฟฟ้า", 63, "สำหรับ 8 เครื่องที่ส่งไปไทยล่วงหน้า + Hybrid 55 เครื่องที่ลืมติดในไต้หวัน"),
            ("เครื่องมือวัด", 3, "เครื่องมือวัดที่ใช้ในการรับรอง NBTC"),
            ("สายเครื่องมือวัด", 3, "สายเคเบิลที่ใช้ในการรับรอง NBTC"),
            ("RF", 4, "เครื่องที่ใช้ในการรับรอง NBTC (ติดฉลากรับรอง NBTC ไม่ได้หรือ?)"),
            ("Hybrid", 4, "เครื่องที่ใช้ในการรับรอง NBTC (ติดฉลากรับรอง NBTC ไม่ได้หรือ?)"),
            ("ปลั๊กพ่วง", 5, "สั่งซื้อให้ NDRS"),
            ("ฉลาก", "-", "สั่งซื้อให้ NDRS"),
            ("", "", "หลังได้รับ Invoice ให้ยื่นคำขอ โดยระยะเวลาขออนุมัติประมาณ 2 สัปดาห์"),
            ("สาย MIS (ไม่มีไฟเลี้ยง)", 50, "สั่งซื้อให้ NDRS"),
            ("อะแดปเตอร์ AC/USB", 13, "สั่งซื้อให้ NDRS"),
        ]
        ws["C11"] = "ตรวจสอบจำนวนสาย USB Serial"

    for idx, (item, qty, remark) in enumerate(rows, start=2):
        ws.cell(idx, 4, item)
        ws.cell(idx, 5, qty)
        ws.cell(idx, 8, remark)

en = copy_sheet_with_layout("English")
fill_sheet(en, "en")

th = copy_sheet_with_layout("ภาษาไทย")
fill_sheet(th, "th")

wb.save(path)
print(path)
