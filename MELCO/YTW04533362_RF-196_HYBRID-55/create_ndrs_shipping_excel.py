import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

base = r"C:\Working_Space\OpenWork\MELCO\YTW04533362_RF-196_HYBRID-55"
out_path = os.path.join(base, "NDRS_shipping_template_from_image.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "ถอดแบบจากรูป"

# Print setup
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.25
ws.page_margins.right = 0.25
ws.page_margins.top = 0.25
ws.page_margins.bottom = 0.25

# Column widths similar to screenshot
widths = {
    "A": 16,  # 場所
    "B": 12,  # カテゴリ
    "C": 56,  # 作業内容
    "D": 24,  # 対象
    "E": 8,   # 数量
    "F": 18,  # 担当
    "G": 18,  # 期間
    "H": 58,  # 備考
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

for r in range(1, 15):
    ws.row_dimensions[r].height = 21

header_fill = PatternFill("solid", fgColor="B7DEE8")
thin = Side(style="thin", color="666666")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

headers = ["場所", "カテゴリ", "作業内容", "対象", "数量", "担当", "期間", "備考"]
for idx, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=idx, value=h)
    c.fill = header_fill
    c.font = Font(name="Yu Gothic", bold=True, size=10)
    c.alignment = center
    c.border = border

rows = [
    ("RF", 196, "ME-TWNから持ってきた端末"),
    ("Hybrid", 55, "ME-TWNから持ってきた端末"),
    ("ミスケーブル", 22, "ME-TWNから持ってきたケーブル"),
    ("電圧シール", 63, "タイへ先行送付した8台＋台湾で貼り忘れたHybrid55台分"),
    ("計器", 3, "NBTC認証で使用した計器"),
    ("計器ケーブル", 3, "NBTC認証で使用したケーブル"),
    ("RF", 4, "NBTC認証で使用した端末（NBTC認証ラベル貼れない？）"),
    ("Hybrid", 4, "NBTC認証で使用した端末（NBTC認証ラベル貼れない？）"),
    ("電源タップ", 5, "NDRS様へ発注"),
    ("ラベル", "－", "NDRS様へ発注"),
    ("", "", "インボイス受領後に申請で、取得にかかる期間は約2週間"),
    ("ミスケーブル（電源なし）", 50, "NDRS様へ発注"),
    ("AC/USBアダプター", 13, "NDRS様へ発注"),
]

start_row = 2
end_row = start_row + len(rows) - 1

ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
ws.merge_cells(start_row=start_row, start_column=3, end_row=10, end_column=3)
ws.merge_cells(start_row=start_row, start_column=6, end_row=end_row, end_column=6)
ws.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7)

ws["A2"] = "NDRSオフィス"
ws["B2"] = "国内輸送"
ws["C2"] = "NDRS様⇒ME-THへ機材送付（車でハンドキャリー）"
ws["F2"] = "MELCO, ME-TH"
ws["G2"] = "Day1"

for cell in ["A2", "B2", "C2", "F2", "G2"]:
    ws[cell].font = Font(name="Yu Gothic", size=10)
    ws[cell].alignment = center if cell != "C2" else left

for i, (target, qty, note) in enumerate(rows, start_row):
    ws.cell(i, 4, target)
    ws.cell(i, 5, qty)
    ws.cell(i, 8, note)
    ws.cell(i, 4).alignment = left
    ws.cell(i, 5).alignment = center
    ws.cell(i, 8).alignment = left
    for col in [4, 5, 8]:
        ws.cell(i, col).font = Font(name="Yu Gothic", size=10)

# Apply borders to all cells including merged areas
for row in ws.iter_rows(min_row=1, max_row=end_row, min_col=1, max_col=8):
    for cell in row:
        cell.border = border

# Blue callout note in 作業内容 area (represented with merged cells)
ws.merge_cells("C11:C13")
ws["C11"] = "USBシリアルケーブルの本数を確認"
ws["C11"].fill = PatternFill("solid", fgColor="126C8A")
ws["C11"].font = Font(name="Yu Gothic", bold=True, size=10, color="FFFFFF")
ws["C11"].alignment = center
ws["C11"].border = Border(
    left=Side("medium", color="003A4D"),
    right=Side("medium", color="003A4D"),
    top=Side("medium", color="003A4D"),
    bottom=Side("medium", color="003A4D"),
)

wb.save(out_path)
print(out_path)
