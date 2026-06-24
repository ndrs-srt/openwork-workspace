import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

wb = openpyxl.Workbook()
base = r"C:\Working_Space\OpenWork\MELCO\YTW04533362_RF-196_HYBRID-55"

thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
sub_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
blue_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
green_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
red_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

def setup_page(ws, orient="portrait"):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = orient
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToPage = True

# ============ SHEET 1: ตารางสรุป ============
ws = wb.active
ws.title = "ตารางสรุป"
setup_page(ws, "portrait")

ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 18

ws.merge_cells("A1:F1")
ws["A1"].value = "สรุปสินค้า YTW04533362 (RF-196 / HYBRID-55)"
ws["A1"].font = Font(name="Arial", bold=True, size=16, color="2F5496")
ws["A1"].alignment = Alignment(horizontal="center")

ws.merge_cells("A3:F3")
ws["A3"].value = "กล่องที่ 1: BOX_No.1_RF-145 (RF 145 ชิ้น, 4 ชั้น)"
ws["A3"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
ws["A3"].fill = blue_fill
ws["A3"].alignment = Alignment(horizontal="left")

headers = ["ชั้น", "รายละเอียด", "RF (ชิ้น)", "HYBRID (ชิ้น)", "รวม (ชิ้น)", "หมายเหตุ"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=i, value=h)
    c.font = Font(name="Arial", bold=True, size=10)
    c.fill = sub_fill; c.alignment = center; c.border = thin

box1 = [
    ("L1 (ล่าง)", "RF อย่างเดียว", 40, 0, 40, ""),
    ("L2", "RF อย่างเดียว", 40, 0, 40, ""),
    ("L3", "RF อย่างเดียว", 40, 0, 40, ""),
    ("L4 (บนสุด)", "RF อย่างเดียว", 25, 0, 25, "ชั้นบนสุด"),
]
for j, vals in enumerate(box1):
    r = 5 + j
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(name="Arial", size=10)
        c.alignment = center; c.border = thin

ws.cell(row=9, column=1, value="รวม BOX 1").font = Font(name="Arial", bold=True, size=10)
for i in range(1, 7):
    ws.cell(row=9, column=i).fill = sub_fill
    ws.cell(row=9, column=i).border = thin
    ws.cell(row=9, column=i).alignment = center
ws.cell(row=9, column=3, value=145).font = Font(name="Arial", bold=True, size=10)
ws.cell(row=9, column=4, value=0)
ws.cell(row=9, column=5, value=145).font = Font(name="Arial", bold=True, size=10)

ws.merge_cells("A11:F11")
ws["A11"].value = "กล่องที่ 2: BOX_No.2_RF-51_HYBRID-55 (RF 51 + HYBRID 55, 3 ชั้น)"
ws["A11"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
ws["A11"].fill = green_fill
ws["A11"].alignment = Alignment(horizontal="left")

for i, h in enumerate(headers, 1):
    c = ws.cell(row=12, column=i, value=h)
    c.font = Font(name="Arial", bold=True, size=10)
    c.fill = sub_fill; c.alignment = center; c.border = thin

box2 = [
    ("L1 (ล่าง)", "RF อย่างเดียว", 40, 0, 40, ""),
    ("L2", "RF + HYBRID", 6, 34, 40, ""),
    ("L3 (บนสุด)", "RF + HYBRID", 5, 21, 26, "ชั้นบนสุด"),
]
for j, vals in enumerate(box2):
    r = 13 + j
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(name="Arial", size=10)
        c.alignment = center; c.border = thin

ws.cell(row=16, column=1, value="รวม BOX 2").font = Font(name="Arial", bold=True, size=10)
for i in range(1, 7):
    ws.cell(row=16, column=i).fill = sub_fill
    ws.cell(row=16, column=i).border = thin
    ws.cell(row=16, column=i).alignment = center
ws.cell(row=16, column=3, value=51).font = Font(name="Arial", bold=True, size=10)
ws.cell(row=16, column=4, value=55).font = Font(name="Arial", bold=True, size=10)
ws.cell(row=16, column=5, value=106).font = Font(name="Arial", bold=True, size=10)

ws.merge_cells("A18:C18")
ws["A18"].value = "รวมทั้งหมด (BOX 1 + BOX 2)"
ws["A18"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
ws["A18"].fill = red_fill
ws["A18"].alignment = Alignment(horizontal="center")
ws["A18"].border = thin
for i in range(2, 4):
    ws.cell(row=18, column=i).fill = red_fill
    ws.cell(row=18, column=i).border = thin

for i, v in {4: "RF", 5: "HYBRID", 6: "รวม"}.items():
    c = ws.cell(row=18, column=i, value=v)
    c.font = Font(name="Arial", bold=True, size=10)
    c.fill = sub_fill; c.alignment = center; c.border = thin

for i, v in {4: 196, 5: 55, 6: 251}.items():
    c = ws.cell(row=19, column=i, value=v)
    c.font = Font(name="Arial", bold=True, size=14)
    c.alignment = center; c.border = thin

for r in range(1, 21):
    ws.row_dimensions[r].height = 22

# ============ SHEET 2: รูป BOX 1 ============
ws2 = wb.create_sheet("รูปกล่องที่ 1 (RF-145)")
setup_page(ws2, "portrait")
ws2.column_dimensions["A"].width = 50
ws2.column_dimensions["B"].width = 30

ws2.merge_cells("A1:B1")
ws2["A1"].value = "รูปประกอบกล่องที่ 1: BOX_No.1_RF-145 (RF 145 ชิ้น, 4 ชั้น)"
ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="2F5496")
ws2["A1"].alignment = Alignment(horizontal="center")

box1_dir = os.path.join(base, "BOX_No.1_RF-145")
box1_imgs = [
    ("BOX_No.1.jpeg", "รูป BOX No.1 (RF-145)"),
    ("L3-L2-L1_each40pcs.jpeg", "ชั้น L1-L2-L3 (ชั้นละ 40 ชิ้น)"),
    ("L4(Top)_25pcs.jpeg", "ชั้น L4 บนสุด (25 ชิ้น)"),
]

row = 3
for fname, cap in box1_imgs:
    p = os.path.join(box1_dir, fname)
    if os.path.exists(p):
        img = PILImage.open(p)
        img.thumbnail((450, 350), PILImage.LANCZOS)
        tmp = f"{p}.thumb.jpg"
        img.save(tmp)
        xl_img = Image(tmp)
        ws2.add_image(xl_img, f"A{row}")
        ws2.cell(row=row + 1, column=1, value=cap).font = Font(name="Arial", bold=True, size=10)
        ws2.cell(row=row + 1, column=1).alignment = Alignment(horizontal="center")
        row += 18

# ============ SHEET 3: รูป BOX 2 ============
ws3 = wb.create_sheet("รูปกล่องที่ 2 (RF-51_HYBRID-55)")
setup_page(ws3, "portrait")
ws3.column_dimensions["A"].width = 50
ws3.column_dimensions["B"].width = 30

ws3.merge_cells("A1:B1")
ws3["A1"].value = "รูปประกอบกล่องที่ 2: BOX_No.2_RF-51_HYBRID-55 (RF 51 + HYBRID 55, 3 ชั้น)"
ws3["A1"].font = Font(name="Arial", bold=True, size=14, color="548235")
ws3["A1"].alignment = Alignment(horizontal="center")

box2_dir = os.path.join(base, "BOX_No.2_RF-51_HYBRID-55")
box2_imgs = [
    ("BOX_No.2.jpeg", "รูป BOX No.2 (RF-51 + HYBRID-55)"),
    ("L1_RF-40pcs.jpeg", "ชั้น L1 (RF 40 ชิ้น)"),
    ("L2_RF-6pcs_HYBRID-34pcs.jpeg", "ชั้น L2 (RF 6 + HYBRID 34 ชิ้น)"),
    ("L3(Top)_RF-5pcs_HYBRID-21pcs.jpeg", "ชั้น L3 บนสุด (RF 5 + HYBRID 21 ชิ้น)"),
]

row = 3
for fname, cap in box2_imgs:
    p = os.path.join(box2_dir, fname)
    if os.path.exists(p):
        img = PILImage.open(p)
        img.thumbnail((450, 350), PILImage.LANCZOS)
        tmp = f"{p}.thumb.jpg"
        img.save(tmp)
        xl_img = Image(tmp)
        ws3.add_image(xl_img, f"A{row}")
        ws3.cell(row=row + 1, column=1, value=cap).font = Font(name="Arial", bold=True, size=10)
        ws3.cell(row=row + 1, column=1).alignment = Alignment(horizontal="center")
        row += 14

output = os.path.join(base, "YTW04533362_RF-196_HYBRID-55_Summary.xlsx")
wb.save(output)
print(f"Saved: {output}")
